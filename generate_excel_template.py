import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_okr_template(filename="OKR_Template_New.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dánh giá OKRs"

    # Define Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
    section_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
    
    header_font = Font(name='Times New Roman', size=11, bold=True, color="FFFFFF")
    section_font = Font(name='Times New Roman', size=11, bold=True, italic=True)
    item_font_bold_italic = Font(name='Times New Roman', size=11, bold=True, italic=True)
    normal_font = Font(name='Times New Roman', size=11)
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 1. Setup Columns and Widths
    ws.column_dimensions['A'].width = 5   # TT
    ws.column_dimensions['B'].width = 60  # Noi dung
    ws.column_dimensions['C'].width = 10  # Tu cham diem
    # User columns (D onwards)
    users = ["Do The Sang", "Le Thi Hong Nhung", "Tran Thanh Son", "Tran Thi Thanh Phuong", 
             "Phan Thanh Dat", "Ha Le Truc Phuong", "Ngo Thi Thuy", "Nguyen Thi Ngoc Bich", 
             "Hoang Tran", "Pham Thanh Tung"]
    
    for i, user in enumerate(users):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 15
        ws.cell(row=2, column=4 + i).value = user

    # 2. Header
    ws.merge_cells('D1:J1') # Merge for title
    title_cell = ws['D1']
    title_cell.value = "ĐÁNH GIÁ OKRs THÁNG 1/2026"
    title_cell.font = Font(name='Times New Roman', size=14, bold=True)
    title_cell.alignment = center_align

    # Row 2 Headers
    headers = ["TT", "Nội dung", "Tự chấm điểm"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Apply header style to user columns
    for i in range(len(users)):
        cell = ws.cell(row=2, column=4 + i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Content Data
    data = [
        # (TT, Noi dung, Diem, StyleType)
        # StyleType: 'section', 'item', 'subitem', 'score'
        
        ("I", "CHẤT LƯỢNG THỰC THI OKR", "", "section"),
        ("1", "Kết quả thực tế so với mục tiêu: (Dịch chuyển so với tháng trước/33,3%)", "", "item"),
        ("-", "Nhỏ hơn 25%", 5, "score"),
        ("-", "Từ 25 - 50%", 10, "score"),
        ("-", "Từ 50 - 75%", 15, "score"),
        ("-", "Từ 75 - 100%", 18, "score"),
        ("-", "Trên 100% hoặc có đột phá lớn (Cấp trên ghi nhận)", 20, "score"),
        
        ("2", "Tiến độ và tính kỷ luật", "", "item"),
        ("2.1", "Đầy đủ OKRs cá nhân được cập nhật trên Base Goal", "Yes/No", "subitem"),
        ("-", "Có Check-in trên base hàng tuần (2 điểm/lần check-in)", 8, "score"),
        ("-", "Có check-in với người khác, cấp quản lý, làm việc chung OKRs trong bộ phận", 2, "score"),
        
        ("2.2", "Chất lượng check - in và hành động trong KR", "", "subitem"),
        ("-", "Không có hành động rõ ràng", 1, "score"),
        ("-", "Chỉ báo cáo trạng thái (đang làm, đang cố, ...)", 3, "score"),
        ("-", "Có mô tả hành động rõ ràng cụ thể + hướng giải quyết", 5, "score"),
        
        ("II", "TẦM QUAN TRỌNG CỦA OKR", "", "section"),
        ("1", "Mức độ đóng góp vào mục tiêu công ty", "", "item"),
        ("-", "Mục tiêu mang tính nội bộ/cá nhân", 1, "score"),
        ("-", "Mục tiêu hỗ trợ gián tiếp Doanh thu/Khách hàng/Chất lượng", 2, "score"),
        ("-", "Mục tiêu hỗ trợ gián tiếp Doanh thu/Khách hàng/Chất lượng", 3, "score"),
        ("-", "Mục tiêu liên quan trực tiếp Doanh thu/Khách hàng/Chất lượng", 4, "score"),
        ("-", "Mục tiêu liên quan trực tiếp Doanh thu/Khách hàng/Chất lượng", 5, "score"),
        
        ("2", "Mức độ ưu tiên mục tiêu của Quý", "", "item"),
        ("-", "Bình thường", 1, "score"),
        ("-", "Quan trọng", 2, "score"),
        ("-", "Quan trọng", 3, "score"),
        ("-", "Rất quan trọng", 4, "score"),
        ("-", "Rất quan trọng", 5, "score"),
        
        ("3", "Tính khó/tầm ảnh hưởng đến hệ thống", "", "item"),
        ("-", "Tác động với cá nhân", 1, "score"),
        ("-", "Tác động nội bộ phòng ban/đội nhóm", 2, "score"),
        ("-", "Tác động nội bộ phòng ban/đội nhóm", 3, "score"),
        ("-", "Tác động nhiều phòng ban/cả công ty", 4, "score"),
        ("-", "Tác động nhiều phòng ban/cả công ty", 5, "score"),
    ]

    current_row = 3
    for tt, content, score, style_type in data:
        # A: TT
        cell_tt = ws.cell(row=current_row, column=1, value=tt)
        cell_tt.alignment = center_align
        cell_tt.border = thin_border
        
        # B: Content
        cell_content = ws.cell(row=current_row, column=2, value=content)
        cell_content.border = thin_border
        cell_content.alignment = left_align
        
        # C: Score
        cell_score = ws.cell(row=current_row, column=3, value=score)
        cell_score.alignment = center_align
        cell_score.border = thin_border

        # User columns borders
        for i in range(len(users)):
            ws.cell(row=current_row, column=4 + i).border = thin_border

        # Styling
        if style_type == "section":
            for col in range(1, 4 + len(users)):
                ws.cell(row=current_row, column=col).fill = section_fill
            cell_tt.font = Font(name='Times New Roman', size=11, bold=True, italic=True)
            cell_content.font = Font(name='Times New Roman', size=11, bold=True, italic=True)
            
        elif style_type == "item":
            cell_tt.font = Font(name='Times New Roman', size=11, bold=True, italic=True)
            cell_content.font = Font(name='Times New Roman', size=11, bold=True, italic=True)
            
        elif style_type == "subitem":
             cell_content.font = Font(name='Times New Roman', size=11, bold=True) # Maybe just bold?
             if score == "Yes/No":
                 cell_score.font = Font(name='Times New Roman', size=11, color="FF0000") # Red text
                 
        else: # score
            cell_content.font = normal_font

        current_row += 1

    # Total Row
    ws.cell(row=current_row, column=2, value="Tổng cộng OKR").font = Font(name='Times New Roman', size=11, bold=True)
    
    # Yellow Highlight for total row specific cells (from image 2 it seems columns D onwards have yellow background and bold numbers)
    # Actually image 1 shows Row 37 "Tong cong OKR" has yellow fill for D, E...
    
    for col in range(4, 4 + len(users)):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = section_fill
        cell.font = Font(name='Times New Roman', size=11, bold=True)
        cell.border = thin_border
        # Add sum formula? For now just visual.
    
    # Apply border to A and B, C on total row
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=3).border = thin_border

    wb.save(filename)
    print(f"File saved: {filename}")

if __name__ == "__main__":
    create_okr_template()
