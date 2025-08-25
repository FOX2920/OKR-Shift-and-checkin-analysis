"""
UI Components và functions cho Streamlit interface
"""
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta, date
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
from utils import DateUtils
from analysis_system import OKRAnalysisSystem
from email_report import EmailReportGenerator


def generate_data_table(users):
    """Generate data table from users"""
    data = []
    for user in users:
        data.append({
            "Name": user.name,
            "Has OKR": "Yes" if user.co_OKR == 1 else "No",
            "Check-in": "Yes" if user.checkin == 1 else "No",
            "OKR Movement": user.dich_chuyen_OKR,
            "Score": user.score
        })
    
    df = pd.DataFrame(data)
    return df


def export_to_excel(users, filename="output1.xlsx"):
    """
    Xuất dữ liệu OKRs của danh sách users ra file Excel với giao diện được cải tiến.

    Yêu cầu:
      - Mỗi user phải có các thuộc tính: name, co_OKR, checkin, dich_chuyen_OKR, score
    """
    # Tạo workbook và sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OKRs"

    # Định nghĩa style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # --- Tiêu đề chính ---
    total_columns = 3 + len(users)  # 3 cột cố định + số user
    last_col_letter = get_column_letter(total_columns)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "ĐÁNH GIÁ OKRs THÁNG"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Header (dòng 2) ---
    fixed_headers = ["TT", "Nội dung", "Tự chấm điểm"]
    user_headers = [user.name for user in users]
    headers = fixed_headers + user_headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        # Đặt độ rộng mặc định cho các cột
        col_letter = get_column_letter(col_idx)
        if col_idx == 2:
            ws.column_dimensions[col_letter].width = 70  # Nội dung dài hơn
        elif col_idx == 1:
            ws.column_dimensions[col_letter].width = 5
        else:
            ws.column_dimensions[col_letter].width = 15

    # --- Các dòng tiêu chí (bắt đầu từ dòng 3) ---
    criteria = [
        [1, "Đầy đủ OKRs cá nhân được cập nhật trên Base Goal (Mục tiêu cá nhân + Đường dẫn)", 1],
        [2, "Có Check-in trên base hàng tuần (Mỗi tuần ít nhất 1 lần check-in)", 0.5],
        [3, "Có check-in với người khác, cấp quản lý, làm việc chung OKRs trong bộ phận", 0.5],
        [4, "Tổng OKRs dịch chuyển trong tháng (so với tháng trước):", ""],
        ["", "Nhỏ hơn 10%", 0.15],
        ["", "Từ 10 - 25%", 0.25],
        ["", "Từ 26 - 30%", 0.5],
        ["", "Từ 31 - 50%", 0.75],
        ["", "Từ 51 - 80%", 1.25],
        ["", "Từ 81% - 99%", 1.5],
        ["", "100% hoặc có đột phá lớn", 2.5],
        [5, "Tổng cộng OKRs", ""]
    ]
    start_row = 3
    for i, row_data in enumerate(criteria):
        row_idx = start_row + i
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            # Đánh dấu cột loại (nếu giá trị đầu tiên là số thứ tự) với màu nền và in đậm
            if col_idx == 1 and isinstance(value, int):
                cell.fill = category_fill
                cell.font = category_font

    # --- Ghi dữ liệu của từng user ---
    # Các user sẽ được hiển thị từ cột 4 trở đi
    for idx, user in enumerate(users, start=1):
        col_idx = 3 + idx  # cột thứ 1-3 đã dành cho tiêu đề cố định
        col_letter = get_column_letter(col_idx)
        # 1. Đánh giá OKRs cá nhân (dòng 3)
        ws.cell(row=3, column=col_idx, value=1 if user.co_OKR == 1 else 0)
        # 2. Check-in hàng tuần (dòng 4)
        ws.cell(row=4, column=col_idx, value=0.5 if user.checkin == 1 else 0)
        # 3. Check-in với người khác (dòng 5)
        ws.cell(row=5, column=col_idx, value=0.5 )

        # 4. Dịch chuyển OKR:
        # Dòng 6 hiển thị % dịch chuyển, các dòng từ 7 đến 13 hiển thị điểm tương ứng
        movement = user.dich_chuyen_OKR
        ws.cell(row=6, column=col_idx, value=f"{movement}%")

        # Xác định điểm dịch chuyển dựa theo % và dòng ghi điểm:
        if movement < 10:
            score_value = 0.15
            movement_row = 7
        elif movement < 26:
            score_value = 0.25
            movement_row = 8
        elif movement < 31:
            score_value = 0.5
            movement_row = 9
        elif movement < 51:
            score_value = 0.75
            movement_row = 10
        elif movement < 81:
            score_value = 1.25
            movement_row = 11
        elif movement < 100:
            score_value = 1.5
            movement_row = 12
        else:
            score_value = 2.5
            movement_row = 13
        ws.cell(row=movement_row, column=col_idx, value=score_value)

        # 5. Tổng điểm: sử dụng công thức SUM từ dòng 3 đến dòng 13
        formula = user.score
        ws.cell(row=14, column=col_idx, value=formula)

        # Áp dụng border và căn giữa cho các ô dữ liệu của user
        for r in range(3, 15):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # --- Freeze panes để cố định header và tiêu chí ---
    ws.freeze_panes = ws["D3"]

    # --- Tự động điều chỉnh độ rộng cột (nếu cần) ---
    # Vòng lặp qua các cột để tính độ rộng dựa trên nội dung
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Return the workbook object
    return wb


def show_data_summary(df, analyzer):
    """Show data summary statistics"""
    st.subheader("📈 Data Summary")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        total_goals = df['goal_id'].nunique()
        st.metric("Total Goals", total_goals)
    
    with col2:
        total_krs = df['kr_id'].nunique()
        st.metric("Total KRs", total_krs)
    
    with col3:
        total_checkins = df['checkin_id'].nunique()
        st.metric("Total Checkins", total_checkins)
    
    with col4:
        total_users = df['goal_user_name'].nunique()
        st.metric("Total Users", total_users)
    
    with col5:
        total_filtered_members = len(analyzer.filtered_members_df) if analyzer.filtered_members_df is not None else 0
        st.metric("Filtered Members", total_filtered_members)


def show_missing_analysis_section(analyzer):
    """Show missing goals and checkins analysis"""
    
    members_without_goals, members_without_checkins, members_with_goals_no_checkins = analyzer.analyze_missing_goals_and_checkins()
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_members = len(analyzer.filtered_members_df) if analyzer.filtered_members_df is not None else 0
        st.metric("Total Members", total_members)
    
    with col2:
        members_with_goals = total_members - len(members_without_goals)
        st.metric("Members with Goals", members_with_goals)
    
    with col3:
        members_with_checkins = total_members - len(members_without_checkins)
        st.metric("Members with Checkins", members_with_checkins)
    
    with col4:
        priority_members = len(members_with_goals_no_checkins)
        st.metric("Priority (Goals but no Checkins)", priority_members)
    
    # Charts
    col1, col2 = st.columns(2)
    
    with col1:
        # Goals distribution pie chart
        fig_goals = px.pie(
            values=[len(members_without_goals), total_members - len(members_without_goals)],
            names=['Without Goals', 'With Goals'],
            title='Goal Distribution',
            color_discrete_sequence=['#ff6b6b', '#4ecdc4']
        )
        st.plotly_chart(fig_goals, use_container_width=True)
    
    with col2:
        # Checkins distribution pie chart
        fig_checkins = px.pie(
            values=[len(members_without_checkins), total_members - len(members_without_checkins)],
            names=['Without Checkins', 'With Checkins'],
            title='Checkin Distribution',
            color_discrete_sequence=['#ff9f43', '#6c5ce7']
        )
        st.plotly_chart(fig_checkins, use_container_width=True)
    
    # Detailed tables
    if members_without_goals:
        st.subheader(f"🚫 Members Without Goals ({len(members_without_goals)})")
        goals_df = pd.DataFrame(members_without_goals)
        st.dataframe(goals_df[['name', 'username', 'job']], use_container_width=True)
    
    if members_without_checkins:
        st.subheader(f"⚠️ Members Without Checkins ({len(members_without_checkins)})")
        checkins_df = pd.DataFrame(members_without_checkins)
        st.dataframe(checkins_df[['name', 'username', 'job', 'has_goal']], use_container_width=True)
    
    if members_with_goals_no_checkins:
        st.subheader(f"🎯 Priority: Goals but No Checkins ({len(members_with_goals_no_checkins)})")
        st.warning("These members have set goals but haven't been checking in. They should be prioritized for follow-up.")
        priority_df = pd.DataFrame(members_with_goals_no_checkins)
        st.dataframe(priority_df[['name', 'username', 'job']], use_container_width=True)


def show_okr_analysis(okr_shifts, reference_date, period="weekly"):
    """Show OKR analysis section"""
    if not okr_shifts:
        st.warning("No OKR data available for analysis")
        return
    
    # Summary statistics
    col1, col2, col3, col4 = st.columns(4)
    
    progress_users = len([u for u in okr_shifts if u['okr_shift'] > 0])
    stable_users = len([u for u in okr_shifts if u['okr_shift'] == 0])
    declining_users = len([u for u in okr_shifts if u['okr_shift'] < 0])
    
    with col1:
        st.metric("📈 Progress", progress_users, delta=progress_users)
    
    with col2:
        st.metric("➡️ Stable", stable_users, delta=0)
    
    with col3:
        st.metric("📉 Declining", declining_users, delta=-declining_users if declining_users > 0 else 0)
    
    with col4:
        total_shift = sum(u['okr_shift'] for u in okr_shifts)
        avg_shift = total_shift / len(okr_shifts) if okr_shifts else 0
        st.metric("📊 Average Shift", f"{avg_shift:.2f}")
    
    # OKR shifts chart
    okr_df = pd.DataFrame(okr_shifts)
    
    # Bar chart for OKR shifts
    fig = px.bar(
        okr_df.head(20), 
        x='user_name', 
        y='okr_shift',
        title=f'Top 20 OKR Shifts ({period.title()})',
        color='okr_shift',
        color_continuous_scale=['red', 'yellow', 'green'],
        labels={'okr_shift': 'OKR Shift', 'user_name': 'User'}
    )
    fig.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig, use_container_width=True)
    
    # Top performers and issues
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🏆 Top Performers")
        top_performers = [u for u in okr_shifts if u['okr_shift'] > 0][:10]
        if top_performers:
            top_df = pd.DataFrame(top_performers)
            st.dataframe(
                top_df[['user_name', 'okr_shift', 'current_value']].round(2),
                use_container_width=True
            )
        else:
            st.info("No users with positive OKR shifts")
    
    with col2:
        st.subheader("⚠️ Need Attention")
        need_attention = [u for u in okr_shifts if u['okr_shift'] < 0][:10]
        if need_attention:
            attention_df = pd.DataFrame(need_attention)
            st.dataframe(
                attention_df[['user_name', 'okr_shift', 'current_value']].round(2),
                use_container_width=True
            )
        else:
            st.success("No users with declining OKR shifts")
    
    # Detailed table
    st.subheader(f"📋 Detailed OKR Analysis ({period.title()})")
    st.dataframe(okr_df.round(2), use_container_width=True)


def show_checkin_analysis(period_checkins, overall_checkins, last_friday, quarter_start):
    """Show checkin analysis section"""
    
    if not period_checkins or not overall_checkins:
        st.warning("No checkin data available for analysis")
        return
    
    # Summary for period checkins
    st.subheader(f"📅 Period Analysis ({quarter_start.strftime('%d/%m/%Y')} - {last_friday.strftime('%d/%m/%Y')})")
    
    period_df = pd.DataFrame(period_checkins)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_period_checkins = period_df['checkin_count_period'].sum()
        st.metric("Total Checkins (Period)", total_period_checkins)
    
    with col2:
        avg_checkin_rate = period_df['checkin_rate_period'].mean()
        st.metric("Avg Checkin Rate", f"{avg_checkin_rate:.1f}%")
    
    with col3:
        active_users_period = len([u for u in period_checkins if u['checkin_count_period'] > 0])
        st.metric("Active Users (Period)", active_users_period)
    
    with col4:
        avg_checkins_per_user = period_df['checkin_count_period'].mean()
        st.metric("Avg Checkins/User", f"{avg_checkins_per_user:.1f}")
    
    # Period checkins chart
    fig_period = px.bar(
        period_df.head(15),
        x='user_name',
        y='checkin_count_period',
        title='Top 15 - Checkins in Analysis Period',
        labels={'checkin_count_period': 'Checkin Count', 'user_name': 'User'}
    )
    fig_period.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_period, use_container_width=True)
    
    # Overall checkins analysis
    st.subheader("🏆 Overall Checkin Champions")
    
    overall_df = pd.DataFrame(overall_checkins)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_overall_checkins = overall_df['total_checkins'].sum()
        st.metric("Total Checkins (All Time)", total_overall_checkins)
    
    with col2:
        avg_frequency = overall_df['checkin_frequency_per_week'].mean()
        st.metric("Avg Frequency/Week", f"{avg_frequency:.2f}")
    
    with col3:
        # Last week analysis
        last_week_active = len([u for u in overall_checkins if u['last_week_checkins'] > 0])
        st.metric("Active Last Week", last_week_active)
    
    with col4:
        highly_active = len([u for u in overall_checkins if u['checkin_frequency_per_week'] >= 2])
        st.metric("Highly Active (≥2/week)", highly_active)
    
    # Overall checkins chart
    fig_overall = px.bar(
        overall_df.head(15),
        x='user_name',
        y='total_checkins',
        title='Top 15 - Overall Checkin Count',
        labels={'total_checkins': 'Total Checkins', 'user_name': 'User'}
    )
    fig_overall.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_overall, use_container_width=True)
    
    # Frequency analysis
    fig_freq = px.scatter(
        overall_df.head(20),
        x='total_checkins',
        y='checkin_frequency_per_week',
        size='last_week_checkins',
        hover_name='user_name',
        title='Checkin Patterns: Total vs Frequency',
        labels={
            'total_checkins': 'Total Checkins',
            'checkin_frequency_per_week': 'Frequency per Week',
            'last_week_checkins': 'Last Week Checkins'
        }
    )
    st.plotly_chart(fig_freq, use_container_width=True)
    
    # Detailed tables
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📊 Period Analysis Details")
        st.dataframe(period_df.head(10), use_container_width=True)
    
    with col2:
        st.subheader("🏆 Overall Champions")
        st.dataframe(overall_df.head(10), use_container_width=True)


def show_export_options(df, okr_shifts, okr_shifts_monthly, period_checkins, overall_checkins, analyzer):
    """Show export options for data"""
    
    st.write("Choose what to export:")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Export OKR Analysis (Weekly)", use_container_width=True):
            if okr_shifts:
                export_df = pd.DataFrame(okr_shifts)
                csv = export_df.to_csv(index=False)
                st.download_button(
                    label="📥 Download Weekly OKR Analysis CSV",
                    data=csv,
                    file_name=f"okr_analysis_weekly_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv"
                )
            else:
                st.warning("No weekly OKR data to export")
        
        if st.button("📈 Export Checkin Analysis", use_container_width=True):
            if period_checkins and overall_checkins:
                period_df = pd.DataFrame(period_checkins)
                overall_df = pd.DataFrame(overall_checkins)
                
                # Combine both dataframes
                period_csv = period_df.to_csv(index=False)
                overall_csv = overall_df.to_csv(index=False)
                
                col1_sub, col2_sub = st.columns(2)
                with col1_sub:
                    st.download_button(
                        label="📥 Period Checkins CSV",
                        data=period_csv,
                        file_name=f"checkin_analysis_period_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv"
                    )
                with col2_sub:
                    st.download_button(
                        label="📥 Overall Checkins CSV",
                        data=overall_csv,
                        file_name=f"checkin_analysis_overall_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv"
                    )
            else:
                st.warning("No checkin data to export")
    
    with col2:
        if okr_shifts_monthly:
            if st.button("📅 Export OKR Analysis (Monthly)", use_container_width=True):
                export_df = pd.DataFrame(okr_shifts_monthly)
                csv = export_df.to_csv(index=False)
                st.download_button(
                    label="📥 Download Monthly OKR Analysis CSV",
                    data=csv,
                    file_name=f"okr_analysis_monthly_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv"
                )
        
        if st.button("📋 Export Complete Dataset", use_container_width=True):
            if df is not None and not df.empty:
                csv = df.to_csv(index=False)
                st.download_button(
                    label="📥 Download Complete Dataset CSV",
                    data=csv,
                    file_name=f"complete_dataset_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv"
                )
            else:
                st.warning("No complete dataset to export")


def show_okr_analysis_monthly(okr_shifts_monthly, last_month_end):
    """Show monthly OKR analysis section"""
    if not okr_shifts_monthly:
        st.warning("No monthly OKR data available for analysis")
        return
    
    st.subheader(f"📅 Monthly OKR Analysis (vs {last_month_end.strftime('%d/%m/%Y')})")
    
    # Summary statistics
    col1, col2, col3, col4 = st.columns(4)
    
    progress_users = len([u for u in okr_shifts_monthly if u['okr_shift_monthly'] > 0])
    stable_users = len([u for u in okr_shifts_monthly if u['okr_shift_monthly'] == 0])
    declining_users = len([u for u in okr_shifts_monthly if u['okr_shift_monthly'] < 0])
    
    with col1:
        st.metric("📈 Progress (Monthly)", progress_users, delta=progress_users)
    
    with col2:
        st.metric("➡️ Stable (Monthly)", stable_users, delta=0)
    
    with col3:
        st.metric("📉 Declining (Monthly)", declining_users, delta=-declining_users if declining_users > 0 else 0)
    
    with col4:
        total_shift = sum(u['okr_shift_monthly'] for u in okr_shifts_monthly)
        avg_shift = total_shift / len(okr_shifts_monthly) if okr_shifts_monthly else 0
        st.metric("📊 Average Shift (Monthly)", f"{avg_shift:.2f}")
    
    # Monthly OKR shifts chart
    okr_monthly_df = pd.DataFrame(okr_shifts_monthly)
    
    # Bar chart for monthly OKR shifts
    fig = px.bar(
        okr_monthly_df.head(20), 
        x='user_name', 
        y='okr_shift_monthly',
        title='Top 20 Monthly OKR Shifts',
        color='okr_shift_monthly',
        color_continuous_scale=['red', 'yellow', 'green'],
        labels={'okr_shift_monthly': 'Monthly OKR Shift', 'user_name': 'User'}
    )
    fig.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig, use_container_width=True)
    
    # Top performers and issues for monthly
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🏆 Top Monthly Performers")
        top_performers_monthly = [u for u in okr_shifts_monthly if u['okr_shift_monthly'] > 0][:10]
        if top_performers_monthly:
            top_df = pd.DataFrame(top_performers_monthly)
            st.dataframe(
                top_df[['user_name', 'okr_shift_monthly', 'current_value']].round(2),
                use_container_width=True
            )
        else:
            st.info("No users with positive monthly OKR shifts")
    
    with col2:
        st.subheader("⚠️ Need Monthly Attention")
        need_attention_monthly = [u for u in okr_shifts_monthly if u['okr_shift_monthly'] < 0][:10]
        if need_attention_monthly:
            attention_df = pd.DataFrame(need_attention_monthly)
            st.dataframe(
                attention_df[['user_name', 'okr_shift_monthly', 'current_value']].round(2),
                use_container_width=True
            )
        else:
            st.success("No users with declining monthly OKR shifts")
    
    # Detailed monthly table
    st.subheader("📋 Detailed Monthly OKR Analysis")
    st.dataframe(okr_monthly_df.round(2), use_container_width=True)


def run_analysis(analyzer, selected_cycle, show_missing_analysis):
    """Run the complete analysis"""
    st.header("🚀 Running OKR & Checkin Analysis")
    
    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def update_progress(message, progress):
        status_text.text(message)
        progress_bar.progress(progress)
    
    try:
        # Load and process data
        update_progress("Loading and processing data...", 0.1)
        df = analyzer.load_and_process_data(update_progress)
        
        if df is None or df.empty:
            st.error("❌ Failed to load or process data")
            return
        
        # Data summary
        show_data_summary(df, analyzer)
        
        # Missing analysis
        if show_missing_analysis:
            st.subheader("🔍 Missing Goals & Checkins Analysis")
            show_missing_analysis_section(analyzer)
        
        # Calculate and show OKR shifts (weekly)
        st.subheader("📊 OKR Analysis (Weekly)")
        with st.spinner("Calculating weekly OKR shifts..."):
            okr_shifts = analyzer.calculate_okr_shifts_by_user()
        
        if okr_shifts:
            show_okr_analysis(okr_shifts, DateUtils.get_last_friday_date(), "weekly")
        else:
            st.warning("No weekly OKR shift data available")
        
        # Calculate and show monthly OKR shifts (if applicable)
        if analyzer.should_calculate_monthly_shift():
            st.subheader("📅 OKR Analysis (Monthly)")
            with st.spinner("Calculating monthly OKR shifts..."):
                okr_shifts_monthly = analyzer.calculate_okr_shifts_by_user_monthly()
            
            if okr_shifts_monthly:
                show_okr_analysis_monthly(okr_shifts_monthly, analyzer.get_last_month_end_date())
            else:
                st.warning("No monthly OKR shift data available")
        else:
            current_month = datetime.now().month
            quarter_months = {1: "Q1", 4: "Q2", 7: "Q3", 10: "Q4"}
            st.info(f"ℹ️ Monthly OKR shift analysis is not calculated for month {current_month} (start of {quarter_months.get(current_month, 'quarter')})")
        
        # Analyze checkin behavior
        st.subheader("📝 Checkin Behavior Analysis")
        with st.spinner("Analyzing checkin behavior..."):
            period_checkins, overall_checkins = analyzer.analyze_checkin_behavior()
        
        if period_checkins and overall_checkins:
            show_checkin_analysis(period_checkins, overall_checkins, DateUtils.get_last_friday_date(), DateUtils.get_quarter_start_date())
        else:
            st.warning("No checkin data available")
        
        # Data export - Updated to include monthly data
        st.subheader("💾 Export Data")
        okr_shifts_monthly = analyzer.calculate_okr_shifts_by_user_monthly() if analyzer.should_calculate_monthly_shift() else []
        show_export_options(df, okr_shifts, okr_shifts_monthly, period_checkins, overall_checkins, analyzer)
        
        st.success("✅ Analysis completed successfully!")
        
    except Exception as e:
        st.error(f"❌ Analysis failed: {e}")
        progress_bar.empty()
        status_text.empty()


def send_email_report(analyzer, email_generator, selected_cycle, email_from, email_password, email_to):
    """Send email report including monthly data when applicable"""
    
    st.header("📧 Sending Email Report")
    
    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def update_progress(message, progress):
        status_text.text(message)
        progress_bar.progress(progress)
    
    try:
        # Load and process data
        update_progress("Loading data for email report...", 0.1)
        df = analyzer.load_and_process_data(update_progress)
        
        if df is None or df.empty:
            st.error("❌ Failed to load data for email report")
            return
        
        update_progress("Analyzing missing goals and checkins...", 0.25)
        members_without_goals, members_without_checkins, members_with_goals_no_checkins = analyzer.analyze_missing_goals_and_checkins()
        
        update_progress("Calculating weekly OKR shifts...", 0.4)
        okr_shifts = analyzer.calculate_okr_shifts_by_user()
        
        # Calculate monthly OKR shifts if applicable
        okr_shifts_monthly = []
        if analyzer.should_calculate_monthly_shift():
            update_progress("Calculating monthly OKR shifts...", 0.55)
            okr_shifts_monthly = analyzer.calculate_okr_shifts_by_user_monthly()
        
        update_progress("Creating email content...", 0.7)
        html_content = email_generator.create_email_content(
            analyzer, selected_cycle, members_without_goals, members_without_checkins,
            members_with_goals_no_checkins, okr_shifts, okr_shifts_monthly
        )
        
        update_progress("Sending email...", 0.9)
        subject = f"📊 Báo cáo tiến độ OKR & Checkin - {selected_cycle['name']} - {datetime.now().strftime('%d/%m/%Y')}"
        
        success, message = email_generator.send_email_report(
            email_from, email_password, email_to, subject, html_content
        )
        
        progress_bar.empty()
        status_text.empty()
        
        if success:
            st.success(f"✅ {message}")
            monthly_note = " (bao gồm phân tích tháng)" if okr_shifts_monthly else ""
            st.info(f"📧 Email report sent to: {email_to}{monthly_note}")
            
            # Show email preview
            if st.checkbox("📋 Show email preview", value=False):
                st.subheader("Email Preview")
                st.components.v1.html(html_content, height=600, scrolling=True)
        else:
            st.error(f"❌ {message}")
            
    except Exception as e:
        progress_bar.empty()
        status_text.empty()
        st.error(f"❌ Error sending email report: {e}")


def main():
    """Main application entry point"""
    st.title("🎯 OKR & Checkin Analysis Dashboard")
    st.markdown("---")

    # Get API tokens from environment variables
    goal_token = os.getenv("GOAL_ACCESS_TOKEN")
    account_token = os.getenv("ACCOUNT_ACCESS_TOKEN")

    # Check if tokens are available
    if not goal_token or not account_token:
        st.error("❌ API tokens not found in environment variables. Please set GOAL_ACCESS_TOKEN and ACCOUNT_ACCESS_TOKEN.")
        st.info("Make sure to set the following environment variables:")
        st.code("""
GOAL_ACCESS_TOKEN=your_goal_token_here
ACCOUNT_ACCESS_TOKEN=your_account_token_here
        """)
        return

    # Initialize analyzer
    try:
        analyzer = OKRAnalysisSystem(goal_token, account_token)
        email_generator = EmailReportGenerator()
    except Exception as e:
        st.error(f"Failed to initialize analyzer: {e}")
        return

    # Sidebar for configuration
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # Show token status
        st.subheader("🔑 API Token Status")
        st.success("✅ Goal Access Token: Loaded")
        st.success("✅ Account Access Token: Loaded")

    # Get cycles
    with st.spinner("🔄 Loading available cycles..."):
        cycles = analyzer.get_cycle_list()

    if not cycles:
        st.error("❌ Could not load cycles. Please check your API tokens and connection.")
        return

    # Cycle selection
    with st.sidebar:
        st.subheader("📅 Cycle Selection")
        cycle_options = {f"{cycle['name']} ({cycle['formatted_start_time']})": cycle for cycle in cycles}
        selected_cycle_name = st.selectbox(
            "Select Cycle",
            options=list(cycle_options.keys()),
            index=0,
            help="Choose the quarterly cycle to analyze"
        )
        
        selected_cycle = cycle_options[selected_cycle_name]
        analyzer.checkin_path = selected_cycle['path']
        
        st.info(f"🎯 **Selected Cycle:**\n\n**{selected_cycle['name']}**\n\nPath: `{selected_cycle['path']}`\n\nStart: {selected_cycle['formatted_start_time']}")

    # Analysis options
    with st.sidebar:
        st.subheader("📊 Analysis Options")
        show_missing_analysis = st.checkbox("Show Missing Goals & Checkins Analysis", value=True)

    # Email configuration
    with st.sidebar:
        st.subheader("📧 Email Report Settings")
        
        # Pre-configured email settings
        email_from = "apluscorp.hr@gmail.com"
        email_password = 'mems nctq yxss gruw'  # App password
        email_to = "xnk3@apluscorp.vn"
        
        st.info("📧 Email settings are pre-configured")
        st.text(f"From: {email_from}")
        st.text(f"To: {email_to}")
        
        # Option to override email recipient
        custom_email = st.text_input("Custom recipient (optional):", placeholder="email@example.com")
        if custom_email.strip():
            email_to = custom_email.strip()

    # Main analysis
    col1, col2 = st.columns(2)
    
    with col1:
        analyze_button = st.button("🚀 Start Analysis", type="primary", use_container_width=True)
    
    with col2:
        email_button = st.button("📧 Send Email Report", type="secondary", use_container_width=True)

    if analyze_button:
        run_analysis(analyzer, selected_cycle, show_missing_analysis)

    # Send email report
    if email_button:
        send_email_report(analyzer, email_generator, selected_cycle, email_from, email_password, email_to)
