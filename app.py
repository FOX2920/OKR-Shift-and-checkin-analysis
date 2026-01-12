import streamlit as st
import pandas as pd
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
import numpy as np
import requests
import json
import time
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, date, timezone, timedelta
from typing import Dict, List, Tuple, Optional, Any
import warnings
import os
import smtplib
from collections import defaultdict
import openpyxl
import calendar
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from email.mime.image import MIMEImage

import base64
from io import BytesIO
import plotly.io as pio
import io
import pytz
import ollama

class TableAPIClient:
    """Client for fetching data from Base Table (ID 81)"""
    
    def __init__(self):
        self.url = "https://table.base.vn/extapi/v1/table/records"
        # Using the tokens established in table.py
        self.TABLE_ACCESS_TOKEN = os.getenv('TABLE_ACCESS_TOKEN')
        if not self.TABLE_ACCESS_TOKEN:
            # Fallback to hardcoded string only for development/testing if needed, or raise warning
            # Keeping the hardcoded one commented out for reference or emergency fallback (NOT RECOMMENDED for production)
            # self.TABLE_ACCESS_TOKEN = '...' 
            pass
        
    def get_checkin_scores(self) -> dict:
        """
        Fetch all records and return a mapping of:
        {user_id}_{timestamp} -> next_action_score
        """
        payload = {'access_token_v2': self.TABLE_ACCESS_TOKEN, 'table_id': 81}
        headers = {}
        scores_map = {}

        try:
            # print("DEBUG: Fetching checkin scores from Base Table...")
            response = requests.post(self.url, headers=headers, data=payload)
            data = response.json()
            records = data.get('data', [])
            
            if not records:
                # print("DEBUG: No records found in Base Table.")
                return {}

            # Extract vals from each record
            rows = [r.get('vals', {}) for r in records]
            df = pd.DataFrame(rows)
            
            # Check if required columns exist
            required_cols = ['f2', 'f7', 'f10']
            for col in required_cols:
                if col not in df.columns:
                    continue

            for _, row in df.iterrows():
                try:
                    score_val = row.get('f2')
                    timestamp_val = row.get('f7')
                    user_id_val = row.get('f10')
                    
                    if pd.isna(score_val) or pd.isna(timestamp_val) or pd.isna(user_id_val):
                        continue
                        
                    # Clean User ID
                    user_id = str(user_id_val).strip()
                    if not user_id: continue
                    
                    # Clean Timestamp
                    try:
                         # dt is naive (representing local time 2026-...)
                         # We verified this is treated as UTC by .timestamp() resulting in +7h offset vs real UTC
                         # So we subtract 7h to get correct UTC timestamp
                         dt = pd.to_datetime(timestamp_val) - pd.Timedelta(hours=7)
                         timestamp = int(dt.timestamp())
                    except:
                        continue
                        
                    # Clean Score
                    try:
                        score = float(score_val)
                    except:
                        score = 0
                    
                    # Create Key matching goal_test.py expectation
                    key = f"{user_id}_{timestamp}"
                    scores_map[key] = score
                    
                except Exception:
                    continue
                    
            # print(f"DEBUG: Loaded {len(scores_map)} checkin scores from Base Table.")
            return scores_map

        except Exception as e:
            # print(f"Error fetching checkin scores from Table: {e}")
            return {}

# Configuration
warnings.filterwarnings('ignore')

# Constants
QUARTER_START_MONTHS = [1, 4, 7, 10]
MIN_WEEKLY_CHECKINS = 2
REQUEST_TIMEOUT = 30
MAX_PAGES_KRS = 50
MAX_PAGES_CHECKINS = 100

# Streamlit configuration
st.set_page_config(
    page_title="OKR & Checkin Analysis",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)


class OKRSheetGenerator:
    """Generates OKR Evaluation Excel Sheet"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
        self.section_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
        self.header_font = Font(name='Times New Roman', size=11, bold=True, color="FFFFFF")
        self.section_font = Font(name='Times New Roman', size=11, bold=True, italic=True)
        self.item_font_bold_italic = Font(name='Times New Roman', size=11, bold=True, italic=True)
        self.normal_font = Font(name='Times New Roman', size=11)
        self.thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def generate_excel(self, users_data: List[Dict], cycle_name: str) -> BytesIO:
        """
        Generate Excel file with user data.
        users_data: List of dicts, each containing 'name', 'stats' (derived from get_user_excel_data)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Đánh giá OKRs"

        # 1. Setup Columns
        ws.column_dimensions['A'].width = 5   # TT
        ws.column_dimensions['B'].width = 60  # Noi dung
        ws.column_dimensions['C'].width = 10  # Tu cham diem
        
        # User columns
        user_names = [u['name'] for u in users_data]
        if not user_names:
             user_names = ["Demo User"] # Fallback if empty to avoid index error in headers

        for i, user in enumerate(user_names):
            col_letter = get_column_letter(4 + i)
            ws.column_dimensions[col_letter].width = 15
            ws.cell(row=2, column=4 + i).value = user

        # 2. Header
        # Avoid error if no users
        end_col = 3 + len(user_names) if user_names else 4
        ws.merge_cells(f'D1:{get_column_letter(end_col)}1')
        title_cell = ws['D1']
        title_cell.value = f"ĐÁNH GIÁ OKRs - {cycle_name}".upper()
        title_cell.font = Font(name='Times New Roman', size=14, bold=True)
        title_cell.alignment = self.center_align

        # Row 2 Headers
        headers = ["TT", "Nội dung", "Tự chấm điểm"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = text
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Apply header style to user columns
        for i in range(len(user_names)):
            cell = ws.cell(row=2, column=4 + i)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 3. Content Template & Data Filling
        # Structure: (TT, Content, Score/Ref, Type, Key_for_filling)
        # Type: section, item, subitem, score_row
        # Key_for_filling: Identifying which stat maps to this row
        
        data_structure = [
            ("I", "CHẤT LƯỢNG THỰC THI OKR", "", "section", None),
            ("1", "Kết quả thực tế so với mục tiêu: (Dịch chuyển so với tháng trước/33,3%)", "", "item", "okr_shift_display"),
            ("-", "Nhỏ hơn 25%", 5, "score_row", "shift_lt_25"),
            ("-", "Từ 25 - 50%", 10, "score_row", "shift_25_50"),
            ("-", "Từ 50 - 75%", 15, "score_row", "shift_50_75"),
            ("-", "Từ 75 - 100%", 18, "score_row", "shift_75_100"),
            ("-", "Trên 100% hoặc có đột phá lớn (Cấp trên ghi nhận)", 20, "score_row", "shift_gt_100"),
            
            ("2", "Tiến độ và tính kỷ luật", "", "item", None),
            ("2.1", "Đầy đủ OKRs cá nhân được cập nhật trên Base Goal", "Yes/No", "subitem", "has_okrs"),
            ("-", "Có Check-in trên base hàng tuần (2 điểm/lần check-in)", 8, "score_row", "checkin_score"),
            ("-", "Có check-in với người khác, cấp quản lý, làm việc chung OKRs trong bộ phận", 2, "score_row", "collab_score"), 
            
            ("2.2", "Chất lượng check - in và hành động trong KR", "", "subitem", None),
            ("-", "Không có hành động rõ ràng", 1, "score_row", "quality_low"),
            ("-", "Chỉ báo cáo trạng thái (đang làm, đang cố, ...)", 3, "score_row", "quality_med"),
            ("-", "Có mô tả hành động rõ ràng cụ thể + hướng giải quyết", 5, "score_row", "quality_high"),
            
            ("II", "TẦM QUAN TRỌNG CỦA OKR", "", "section", None),
            ("1", "Mức độ đóng góp vào mục tiêu công ty", "", "item", None),
            ("-", "Mục tiêu mang tính nội bộ/cá nhân", 1, "score_row", "align_personal"),
            ("-", "Mục tiêu hỗ trợ gián tiếp Doanh thu/Khách hàng/Chất lượng (1)", 2, "score_row", "align_indirect_1"),
            ("-", "Mục tiêu hỗ trợ gián tiếp Doanh thu/Khách hàng/Chất lượng (2)", 3, "score_row", "align_indirect_2"),
            ("-", "Mục tiêu liên quan trực tiếp Doanh thu/Khách hàng/Chất lượng", 4, "score_row", "align_direct_1"),
            ("-", "Mục tiêu liên quan trực tiếp Doanh thu/Khách hàng/Chất lượng (High)", 5, "score_row", "align_direct_2"),
            
            ("2", "Mức độ ưu tiên mục tiêu của Quý", "", "item", None),
            ("-", "Bình thường", 1, "score_row", "prio_normal"),
            ("-", "Quan trọng", 2, "score_row", "prio_important_1"),
            ("-", "Quan trọng (High)", 3, "score_row", "prio_important_2"),
            ("-", "Rất quan trọng", 4, "score_row", "prio_very_important_1"),
            ("-", "Rất quan trọng (High)", 5, "score_row", "prio_very_important_2"),
            
            ("3", "Tính khó/tầm ảnh hưởng đến hệ thống", "", "item", None),
            ("-", "Tác động với cá nhân", 1, "score_row", "impact_personal"),
            ("-", "Tác động nội bộ phòng ban/đội nhóm", 2, "score_row", "impact_team_1"),
            ("-", "Tác động nội bộ phòng ban/đội nhóm (High)", 3, "score_row", "impact_team_2"),
            ("-", "Tác động nhiều phòng ban/cả công ty", 4, "score_row", "impact_company_1"),
            ("-", "Tác động nhiều phòng ban/cả công ty (High)", 5, "score_row", "impact_company_2"),
        ]

        current_row = 3
        
        for tt, content, score_ref, style_type, key in data_structure:
            # A: TT
            cell_tt = ws.cell(row=current_row, column=1, value=tt)
            cell_tt.alignment = self.center_align
            cell_tt.border = self.thin_border
            
            # B: Content
            cell_content = ws.cell(row=current_row, column=2, value=content)
            cell_content.border = self.thin_border
            cell_content.alignment = self.left_align
            
            # C: Score
            cell_score = ws.cell(row=current_row, column=3, value=score_ref)
            cell_score.alignment = self.center_align
            cell_score.border = self.thin_border

            # User columns
            for i, user_data in enumerate(users_data):
                col_idx = 4 + i
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = self.thin_border
                cell.alignment = self.center_align
                
                stats = user_data.get('stats', {})
                
                if key: # Modified to check key for all types
                     if style_type == "item" and key == "okr_shift_display":
                         val = stats.get(key, '')
                         if val: 
                            # Display format: e.g. "35%"
                            cell.value = val
                     elif style_type == "score_row" or style_type == "subitem":
                         if key == "checkin_score":
                             val = stats.get('checkin_score_val', '')
                             if val: cell.value = val
                         elif key == "has_okrs":
                             val = stats.get('has_okrs', '')
                             cell.value = val 
                             if val == 'No': cell.font = Font(color="FF0000")
                         elif key in stats and stats[key]:
                             if isinstance(score_ref, (int, float)):
                                cell.value = score_ref 
                             else:
                                cell.value = "x"

            # Styling
            if style_type == "section":
                for col in range(1, 4 + len(user_names)):
                    ws.cell(row=current_row, column=col).fill = self.section_fill
                cell_tt.font = self.section_font
                cell_content.font = self.section_font
                
            elif style_type == "item":
                cell_tt.font = self.item_font_bold_italic
                cell_content.font = self.item_font_bold_italic
                
            elif style_type == "subitem":
                 cell_content.font = Font(name='Times New Roman', size=11, bold=True)
                 if score_ref == "Yes/No":
                     cell_score.font = Font(name='Times New Roman', size=11, color="FF0000")
                     
            else: # score_row
                cell_content.font = self.normal_font

            current_row += 1

        # Total Row
        total_row_idx = current_row
        ws.cell(row=total_row_idx, column=2, value="Tổng cộng OKR").font = Font(name='Times New Roman', size=11, bold=True)
        ws.cell(row=total_row_idx, column=1).border = self.thin_border
        ws.cell(row=total_row_idx, column=2).border = self.thin_border
        ws.cell(row=total_row_idx, column=3).border = self.thin_border

        for i, _ in enumerate(user_names):
            col_idx = 4 + i
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = self.section_fill
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.border = self.thin_border
            # SUM formula
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}3:{col_letter}{total_row_idx-1})"

        # 4. Legend (Fixed at B40-B44 as requested)
        legend_data = [
            (40, "Mức 1 ≤ 15", "FF0000"),       # Red
            (41, "15 < Mức 2 ≤ 25", "00B0F0"),  # Cyan/Blue
            (42, "25 < Mức 3 ≤ 35", "FFFF00"),  # Yellow
            (43, "35 < Mức 4 ≤ 45", "00B050"),  # Green
            (44, "Mức 5 > 45", "7030A0")        # Purple
        ]

        for r, text, color in legend_data:
            cell = ws.cell(row=r, column=2)
            cell.value = text
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.alignment = self.center_align
            # Merge cell B with C for better look? User said B40-B44. Let's stick to B.
            # But usually legends might span. The image looks like it spans B to... well image is cut off.
            # But let's assume just B. Note column B width is 60 so it fits.
            cell.border = self.thin_border
            
            # Add border to all columns in this row for B to end_col? Use image as guide.
            # Image shows the colored bar spans the width of column B ("Nội dung").
            # It doesn't look like it spans "Tu cham diem" (C).
            # So applying to Column B only is correct.

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class DateUtils:
    """Utility class for date calculations"""
    
    @staticmethod
    def get_last_friday_date() -> datetime:
        """Get last Friday date - always returns Friday of previous week"""
        today = datetime.now()
        current_weekday = today.weekday()
        days_to_monday_current_week = current_weekday
        monday_current_week = today - timedelta(days=days_to_monday_current_week)
        monday_previous_week = monday_current_week - timedelta(days=7)
        friday_previous_week = monday_previous_week + timedelta(days=4)
        return friday_previous_week.replace(hour=23, minute=59, second=59)

    @staticmethod
    def get_quarter_start_date() -> datetime:
        """Get current quarter start date"""
        today = datetime.now()
        quarter = (today.month - 1) // 3 + 1
        quarter_start_month = (quarter - 1) * 3 + 1
        return datetime(today.year, quarter_start_month, 1)

    @staticmethod
    def get_last_month_end_date() -> datetime:
        """Get last day of previous month"""
        today = datetime.now()
        first_day_current_month = datetime(today.year, today.month, 1)
        last_day_previous_month = first_day_current_month - timedelta(days=1)
        return last_day_previous_month.replace(hour=23, minute=59, second=59)

    @staticmethod
    def convert_timestamp_to_datetime(timestamp) -> Optional[str]:
        """Convert timestamp to datetime string in Asia/Ho_Chi_Minh timezone"""
        if timestamp is None or timestamp == '' or timestamp == 0:
            return None
        import pytz # Import locally to avoid messing up top of file if not needed everywhere
        try:
            dt_utc = datetime.fromtimestamp(int(timestamp), tz=timezone.utc)
            tz_hcm = pytz.timezone('Asia/Ho_Chi_Minh')
            dt_hcm = dt_utc.astimezone(tz_hcm)
            return dt_hcm.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            return None

    @staticmethod
    def should_calculate_monthly_shift() -> bool:
        """
        Always calculate monthly shift as per new requirement
        """
        return True
    
    @staticmethod
    def is_last_week_of_month() -> bool:
        """
        Exact copy of is_last_week_of_month from checkin.py
        Kiểm tra xem hiện tại có phải tuần cuối cùng của tháng không
        """
        now = datetime.now()
        weeks = DateUtils._get_weeks_in_current_month()
        
        if not weeks:
            return False
        
        last_week = weeks[-1]
        return last_week['start_date'] <= now.date() <= last_week['end_date']
    
    @staticmethod
    def is_week_4_or_5_of_quarter_start_month() -> bool:
        """
        Kiểm tra xem hiện tại có phải tuần 4 hoặc 5 của tháng đầu quý không
        """
        now = datetime.now()
        current_month = now.month
        
        # Kiểm tra xem có phải tháng đầu quý không (1, 4, 7, 10)
        if current_month not in QUARTER_START_MONTHS:
            return False
        
        # Lấy danh sách các tuần trong tháng hiện tại
        weeks = DateUtils._get_weeks_in_current_month()
        
        if not weeks:
            return False
        
        # Tìm tuần hiện tại
        current_week_number = None
        for week in weeks:
            if week['start_date'] <= now.date() <= week['end_date']:
                current_week_number = week['week_number']
                break
        
        # Kiểm tra xem có phải tuần 4 hoặc 5 không
        return current_week_number in [4, 5]
    
    @staticmethod
    def _get_weeks_in_current_month():
        """
        Exact copy of get_weeks_in_current_month from checkin.py
        Lấy tất cả các tuần trong tháng hiện tại
        Quy tắc: Nếu ngày đầu/cuối tháng rơi vào thứ 2-6, vẫn tính là tuần của tháng đó
        """
        now = datetime.now()
        year = now.year
        month = now.month
        
        # Ngày đầu và cuối tháng
        first_day = datetime(year, month, 1)
        last_day = datetime(year, month, calendar.monthrange(year, month)[1])
        
        weeks = []
        current_date = first_day
        
        while current_date <= last_day:
            # Tìm ngày thứ 2 của tuần (hoặc ngày đầu tháng nếu tuần bắt đầu trước đó)
            week_start = current_date - timedelta(days=current_date.weekday())
            week_start = max(week_start, first_day)  # Không được trước ngày 1
            
            # Tìm ngày chủ nhật của tuần (hoặc ngày cuối tháng nếu tuần kết thúc sau đó)
            week_end = week_start + timedelta(days=6)
            week_end = min(week_end, last_day)  # Không được sau ngày cuối tháng
            
            weeks.append({
                'week_number': len(weeks) + 1,
                'start_date': week_start.date(),
                'end_date': week_end.date(),
                'week_range': f"{week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}"
            })
            
            # Chuyển sang tuần tiếp theo
            current_date = week_end + timedelta(days=1)
        
        return weeks


class User:
    """User class for OKR tracking"""
    
    def __init__(self, user_id, name, co_OKR=1, checkin=0, dich_chuyen_OKR=0, score=0, median_score=0):
        self.user_id = str(user_id)
        self.name = name
        self.co_OKR = co_OKR
        self.checkin = checkin  # 0/1: có check-in ít nhất 3/4 tuần trong tháng hiện tại
        self.dich_chuyen_OKR = dich_chuyen_OKR
        self.median_score = median_score
        self.score = score
        self.OKR = {month: 0 for month in range(1, 13)}

    def update_okr(self, month, value):
        """Update OKR for specific month"""
        if 1 <= month <= 12:
            self.OKR[month] = value

    def calculate_score(self):
        """Calculate score based on criteria: check-in, OKR and OKR movement"""
        score = 0.5
        
        # Check-in contributes 0.5 points
        # Logic tính checkin score được handle ở UserManager level
        if self.checkin == 1:
            score += 0.5
        
        # Having OKR contributes 1 point
        if self.co_OKR == 1:
            score += 1
        
        # OKR movement scoring
        movement = self.dich_chuyen_OKR
        movement_scores = [
            (10, 0.15), (25, 0.25), (30, 0.5), (50, 0.75),
            (80, 1.25), (99, 1.5), (float('inf'), 2.5)
        ]
        
        for threshold, points in movement_scores:
            if movement < threshold:
                score += points
                break
        
        self.score = round(score, 2)

    def __repr__(self):
        return (f"User(id={self.user_id}, name={self.name}, co_OKR={self.co_OKR}, "
                f"checkin={self.checkin}, dich_chuyen_OKR={self.dich_chuyen_OKR}, median={self.median_score}, score={self.score})")


class UserManager:
    """Manages user data and calculations"""
    
    def __init__(self, account_df, krs_df, checkin_df, cycle_df=None, final_df=None, users_with_okr_names=None, monthly_okr_data=None):
        self.account_df = account_df
        self.krs_df = krs_df
        self.checkin_df = checkin_df
        self.cycle_df = cycle_df
        self.final_df = final_df
        self.users_with_okr_names = users_with_okr_names or set()  # Danh sách tên users có OKR
        self.monthly_okr_data = monthly_okr_data or []  # Dữ liệu monthly OKR từ bảng "Tất cả nhân viên tiến bộ (tháng)"
        
        self.user_name_map = self._create_user_name_map()
        self.users = self._create_users()

    def _create_user_name_map(self) -> Dict[str, str]:
        """Create user_id to name mapping from account_df"""
        user_map = {}
        if not self.account_df.empty and 'id' in self.account_df.columns and 'name' in self.account_df.columns:
            for _, row in self.account_df.iterrows():
                user_map[str(row['id'])] = row.get('name', 'Unknown')
        return user_map

    def _create_users(self) -> Dict[str, User]:
        """Create User objects for ALL account members, with accurate OKR marking"""
        users = {}
        
        # Tạo users cho TẤT CẢ account_df
        if not self.account_df.empty and 'id' in self.account_df.columns and 'name' in self.account_df.columns:
            for _, row in self.account_df.iterrows():
                user_id = str(row.get('id'))
                name = row.get('name', 'Unknown')
                
                # Chỉ đánh dấu co_OKR=1 nếu tên user có trong danh sách users_with_okr_names
                has_okr = 1 if name in self.users_with_okr_names else 0
                
                users[user_id] = User(user_id, name, co_OKR=has_okr)

        return users

    def update_checkins(self, start_date=None, end_date=None):
        """Update check-in status for each user"""
        for user in self.users.values():
            if self._meets_monthly_weekly_criteria(user.user_id):
                user.checkin = 1
            else:
                user.checkin = 0

    def _get_user_checkins(self, user_id) -> List[datetime]:
        """Get all checkin dates for a user - Using checkin data from final_df like checkin.py"""
        checkins = []
        
        # Use final_df with checkin data like checkin.py does
        if self.final_df is not None and not self.final_df.empty:
            # Get user name first
            user_name = self.user_name_map.get(str(user_id), '')
            if not user_name:
                return checkins
                
            # Get checkins for this user from final_df (equivalent to checkins_df in checkin.py)
            user_checkins = self.final_df[
                (self.final_df['goal_user_name'] == user_name) &
                (self.final_df['checkin_since'].notna()) &
                (self.final_df['checkin_since'] != '')
            ]
            
            for _, entry in user_checkins.iterrows():
                try:
                    # Parse checkin_since to datetime - same format as checkin.py
                    checkin_date_str = entry.get('checkin_since', '')
                    if checkin_date_str:
                        checkin_date = pd.to_datetime(checkin_date_str)
                    checkins.append(checkin_date)
                except (ValueError, TypeError):
                    continue
        return checkins

    def _meets_monthly_weekly_criteria(self, user_id) -> bool:
        """Check if user has checkins in at least 3 weeks of current month"""
        result = self._get_monthly_weekly_criteria_details(user_id)
        return result['meets_criteria']
    
    def _get_monthly_weekly_criteria_details(self, user_id) -> dict:
        """Get detailed information about monthly weekly criteria for a user - Exact copy from checkin.py logic"""
        
        # Get user name from user_id
        user_name = self.user_name_map.get(str(user_id), '')
        if not user_name:
            return {
                'meets_criteria': False,
                'weeks_with_checkins': 0,
                'total_weeks_in_month': 0,
                'checkins_count': 0,
                'week_details': []
            }
        
        # Use exact logic from checkin.py calculate_weekly_checkin_scores
        now = datetime.now()
        
        # Get weeks in current month - exact copy from checkin.py
        current_month_weeks = self._get_weeks_in_current_month_from_checkin_py()
        current_month_year = f"{now.year}-{now.month:02d}"
        
        # Get checkins data from final_df (equivalent to checkins_df in checkin.py)
        if self.final_df is None or self.final_df.empty:
            return {
                'meets_criteria': False,
                'weeks_with_checkins': 0,
                'total_weeks_in_month': len(current_month_weeks),
                'checkins_count': 0,
                'week_details': []
            }
        
        # Filter checkins for this user and current month - exact logic from checkin.py
        user_checkins = self.final_df[
            (self.final_df['goal_user_name'] == user_name) &
            (self.final_df['checkin_since'].notna()) &
            (self.final_df['checkin_since'] != '')
        ].copy()
        
        if user_checkins.empty:
            week_details = []
            for week in current_month_weeks:
                week_details.append({
                    'week_range': week['week_range'],
                    'has_checkin': False,
                    'checkin_dates': []
                })
            
            return {
                'meets_criteria': False,
                'weeks_with_checkins': 0,
                'total_weeks_in_month': len(current_month_weeks),
                'checkins_count': 0,
                'week_details': week_details
            }
        
        # Convert checkin dates and filter by current month - exact logic from checkin.py
        user_checkins['checkin_date'] = pd.to_datetime(user_checkins['checkin_since']).dt.date
        user_checkins['checkin_month_year'] = pd.to_datetime(user_checkins['checkin_since']).dt.strftime('%Y-%m')
        
        current_month_checkins = user_checkins[user_checkins['checkin_month_year'] == current_month_year].copy()
        
        if current_month_checkins.empty:
            week_details = []
            for week in current_month_weeks:
                week_details.append({
                    'week_range': week['week_range'],
                    'has_checkin': False,
                    'checkin_dates': []
                })
            
            return {
                'meets_criteria': False,
                'weeks_with_checkins': 0,
                'total_weeks_in_month': len(current_month_weeks),
                'checkins_count': 0,
                'week_details': week_details
            }
        
        # Determine which week each checkin belongs to - exact logic from checkin.py
        def get_week_number(checkin_date):
            for week in current_month_weeks:
                if week['start_date'] <= checkin_date <= week['end_date']:
                    return week['week_number']
            return None
        
        current_month_checkins['week_number'] = current_month_checkins['checkin_date'].apply(get_week_number)
        
        # Count weeks with checkins - exact logic from checkin.py
        user_weekly_checkins = current_month_checkins.groupby(['week_number']).size().reset_index(name='checkins_count')
        weeks_with_checkins = len(user_weekly_checkins['week_number'].unique())
        total_checkins = len(current_month_checkins)
        
        # Create week details - exact logic from checkin.py
        week_details = []
        week_checkins_map = {}
        
        # Group checkins by week and count occurrences per date
        for _, row in current_month_checkins.iterrows():
            week_num = row['week_number']
            if pd.notna(week_num):
                if week_num not in week_checkins_map:
                    week_checkins_map[week_num] = {}
                date_str = row['checkin_date'].strftime('%d/%m')
                if date_str not in week_checkins_map[week_num]:
                    week_checkins_map[week_num][date_str] = 0
                week_checkins_map[week_num][date_str] += 1
        
        for week in current_month_weeks:
            week_number = week['week_number']
            has_checkin = week_number in week_checkins_map
            
            # Format checkin dates as "DD/MM-X lần"
            checkin_dates = []
            if has_checkin:
                for date_str, count in sorted(week_checkins_map[week_number].items()):
                    if count == 1:
                        checkin_dates.append(date_str)
                    else:
                        checkin_dates.append(f"{date_str}-{count} lần")
            else:
                checkin_dates = []
            
            week_details.append({
                'week_range': week['week_range'],
                'has_checkin': has_checkin,
                'checkin_dates': checkin_dates
            })
        
        # Check if meets criteria (>= 2 weeks OR total_checkins > 5)
        meets_criteria = total_checkins > 3
        
        return {
            'meets_criteria': meets_criteria,
            'weeks_with_checkins': weeks_with_checkins,
            'total_weeks_in_month': len(current_month_weeks),
            'checkins_count': total_checkins,
            'week_details': week_details
        }

    def _get_weeks_in_current_month_from_checkin_py(self):
        """
        Exact copy of get_weeks_in_current_month from checkin.py
        Lấy tất cả các tuần trong tháng hiện tại
        Quy tắc: Nếu ngày đầu/cuối tháng rơi vào thứ 2-6, vẫn tính là tuần của tháng đó
        """
        now = datetime.now()
        year = now.year
        month = now.month
        
        # Ngày đầu và cuối tháng
        first_day = datetime(year, month, 1)
        last_day = datetime(year, month, calendar.monthrange(year, month)[1])
        
        weeks = []
        current_date = first_day
        
        while current_date <= last_day:
            # Tìm ngày thứ 2 của tuần (hoặc ngày đầu tháng nếu tuần bắt đầu trước đó)
            week_start = current_date - timedelta(days=current_date.weekday())
            week_start = max(week_start, first_day)  # Không được trước ngày 1
            
            # Tìm ngày chủ nhật của tuần (hoặc ngày cuối tháng nếu tuần kết thúc sau đó)
            week_end = week_start + timedelta(days=6)
            week_end = min(week_end, last_day)  # Không được sau ngày cuối tháng
            
            weeks.append({
                'week_number': len(weeks) + 1,
                'start_date': week_start.date(),
                'end_date': week_end.date(),
                'week_range': f"{week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}"
            })
            
            # Chuyển sang tuần tiếp theo
            current_date = week_end + timedelta(days=1)
        
        return weeks

    def _get_weeks_in_current_month(self):
        """
        Lấy tất cả các tuần trong tháng hiện tại
        Quy tắc: Nếu ngày đầu/cuối tháng rơi vào thứ 2-6, vẫn tính là tuần của tháng đó
        """
        now = datetime.now()
        year = now.year
        month = now.month
        
        # Ngày đầu và cuối tháng
        first_day = datetime(year, month, 1)
        last_day = datetime(year, month, calendar.monthrange(year, month)[1])
        
        weeks = []
        current_date = first_day
        
        while current_date <= last_day:
            # Tìm ngày thứ 2 của tuần (hoặc ngày đầu tháng nếu tuần bắt đầu trước đó)
            week_start = current_date - timedelta(days=current_date.weekday())
            week_start = max(week_start, first_day)  # Không được trước ngày 1
            
            # Tìm ngày chủ nhật của tuần (hoặc ngày cuối tháng nếu tuần kết thúc sau đó)
            week_end = week_start + timedelta(days=6)
            week_end = min(week_end, last_day)  # Không được sau ngày cuối tháng
            
            weeks.append({
                'week_number': len(weeks) + 1,
                'start_date': week_start.date(),
                'end_date': week_end.date(),
                'week_range': f"{week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}"
            })
            
            # Chuyển sang tuần tiếp theo
            current_date = week_end + timedelta(days=1)
        
        return weeks
    
    def _get_week_number_for_date(self, checkin_date, month_weeks):
        """Xác định checkin_date thuộc tuần nào trong month_weeks"""
        for week in month_weeks:
            if week['start_date'] <= checkin_date <= week['end_date']:
                return week['week_number']
        return None

    def update_okr_movement(self):
        """Update OKR movement for each user - ALWAYS use monthly calculation"""
        # Luôn sử dụng monthly shift theo yêu cầu
        self._update_okr_movement_monthly()

    def _update_okr_movement_quarter_start(self):
        """Update OKR movement for quarter start months (1,4,7,10)"""
        for user in self.users.values():
            current_okr = self._calculate_current_value_for_user(user.user_id)
            user.dich_chuyen_OKR = current_okr

    def _update_okr_movement_monthly(self):
        """Update OKR movement using EXACT values from Monthly OKR Analysis table"""
        
        # Tạo mapping từ tên user đến dịch chuyển tháng từ bảng "Tất cả nhân viên tiến bộ (tháng)"
        monthly_shift_map = {}
        if self.monthly_okr_data:
            for data in self.monthly_okr_data:
                monthly_shift_map[data['user_name']] = data['okr_shift_monthly']
        
        for user in self.users.values():
            user_name = user.name
            
            if user_name in monthly_shift_map:
                # Sử dụng CHÍNH XÁC dịch chuyển tháng từ bảng "Tất cả nhân viên tiến bộ (tháng)"
                user.dich_chuyen_OKR = round(monthly_shift_map[user_name], 2)
            else:
                # Không có trong Monthly OKR Analysis => không có OKR movement
                user.dich_chuyen_OKR = 0

    def _calculate_current_value_for_user(self, user_id) -> float:
        """Calculate current OKR value for a specific user"""
        try:
            if self.final_df is None:
                return 0
            
            user_name = self.user_name_map.get(user_id, '')
            if not user_name:
                return 0
                
            user_df = self.final_df[self.final_df['goal_user_name'] == user_name].copy()
            if user_df.empty:
                return 0
                
            unique_goals = user_df.groupby('goal_name')['goal_current_value'].first().reset_index()
            unique_goals['goal_current_value'] = pd.to_numeric(unique_goals['goal_current_value'], errors='coerce').fillna(0)
            return unique_goals['goal_current_value'].mean() if len(unique_goals) > 0 else 0
            
        except Exception as e:
            st.error(f"Error calculating current value for user {user_id}: {e}")
            return 0

    def _calculate_final_okr_goal_shift_monthly_for_user(self, user_id) -> float:
        """Calculate final_okr_goal_shift_monthly for a specific user"""
        try:
            if self.final_df is None:
                return 0
                
            user_name = self.user_name_map.get(user_id, '')
            if not user_name:
                return 0
                
            user_df = self.final_df[self.final_df['goal_user_name'] == user_name].copy()
            if user_df.empty:
                return 0
            
            reference_month_end = DateUtils.get_last_month_end_date()
            unique_combinations = {}
            
            for _, row in user_df.iterrows():
                goal_name = row.get('goal_name', '')
                kr_name = row.get('kr_name', '')
                
                if not goal_name or not kr_name:
                    continue
                
                combo_key = f"{goal_name}|{kr_name}"
                kr_shift = self._calculate_kr_shift_last_month(row, reference_month_end)
                
                if combo_key not in unique_combinations:
                    unique_combinations[combo_key] = []
                unique_combinations[combo_key].append(kr_shift)
            
            final_okr_monthly_shifts = [
                sum(kr_shifts) / len(kr_shifts) 
                for kr_shifts in unique_combinations.values() 
                if kr_shifts
            ]
            
            return sum(final_okr_monthly_shifts) / len(final_okr_monthly_shifts) if final_okr_monthly_shifts else 0
            
        except Exception as e:
            st.error(f"Error calculating final_okr_goal_shift_monthly for user {user_id}: {e}")
            return 0

    def _calculate_kr_shift_last_month(self, row, reference_month_end) -> float:
        """Calculate kr_shift_last_month = kr_current_value - last_month_end_checkin_value"""
        try:
            kr_current_value = pd.to_numeric(row.get('kr_current_value', 0), errors='coerce')
            if pd.isna(kr_current_value):
                kr_current_value = 0
            
            kr_id = row.get('kr_id', '')
            if not kr_id or self.final_df is None:
                return kr_current_value
            
            quarter_start = DateUtils.get_quarter_start_date()
            
            kr_checkins = self.final_df[
                (self.final_df['kr_id'] == kr_id) & 
                (self.final_df['checkin_id'].notna()) &
                (self.final_df['checkin_name'].notna()) &
                (self.final_df['checkin_name'] != '')
            ].copy()
            
            if not kr_checkins.empty:
                kr_checkins['checkin_since_dt'] = pd.to_datetime(kr_checkins['checkin_since'], errors='coerce')
                kr_checkins = kr_checkins[
                    (kr_checkins['checkin_since_dt'] >= quarter_start) &
                    (kr_checkins['checkin_since_dt'] <= reference_month_end)
                ]
                
                if not kr_checkins.empty:
                    latest_checkin = kr_checkins.loc[kr_checkins['checkin_since_dt'].idxmax()]
                    last_month_checkin_value = pd.to_numeric(latest_checkin.get('checkin_kr_current_value', 0), errors='coerce')
                    if pd.isna(last_month_checkin_value):
                        last_month_checkin_value = 0
                else:
                    last_month_checkin_value = 0
            else:
                last_month_checkin_value = 0
            
            return kr_current_value - last_month_checkin_value
            
        except Exception as e:
            st.warning(f"Error calculating kr_shift_last_month: {e}")
            return 0

    def _calculate_last_month_value_for_user(self, user_df, last_month_end) -> float:
        """Calculate OKR value as of last month end for specific user"""
        try:
            df = user_df.copy()
            df['checkin_since_dt'] = pd.to_datetime(df['checkin_since'], errors='coerce')

            unique_krs = df['kr_id'].dropna().unique()
            goal_values_dict = {}

            for kr_id in unique_krs:
                kr_data = df[df['kr_id'] == kr_id].copy()
                kr_data['checkin_since_dt'] = pd.to_datetime(kr_data['checkin_since'], errors='coerce')

                actual_checkins_before_month_end = kr_data[
                    (kr_data['checkin_since_dt'] <= last_month_end) &
                    (kr_data['checkin_name'].notna()) &
                    (kr_data['checkin_name'] != '')
                ]

                goal_name = kr_data.iloc[0]['goal_name'] if len(kr_data) > 0 else f"Unknown_{kr_id}"

                if len(actual_checkins_before_month_end) > 0:
                    latest_checkin = actual_checkins_before_month_end.sort_values('checkin_since_dt').iloc[-1]
                    kr_value = pd.to_numeric(latest_checkin['checkin_kr_current_value'], errors='coerce')
                    kr_value = kr_value if not pd.isna(kr_value) else 0

                    if goal_name not in goal_values_dict:
                        goal_values_dict[goal_name] = []
                    goal_values_dict[goal_name].append(kr_value)
                else:
                    goal_key = f"{goal_name}_no_checkin_{kr_id}"
                    goal_values_dict[goal_key] = [0]

            goal_values = [np.mean(kr_values_list) for kr_values_list in goal_values_dict.values()]
            return np.mean(goal_values) if goal_values else 0

        except Exception as e:
            st.error(f"Error calculating last month value: {e}")
            return 0

    def calculate_scores(self):
        """Calculate scores for all users with checkin criteria (≥2 weeks OR >5 checkins)"""
        # Chỉ tính checkin score khi ở tuần cuối cùng của tháng
        is_last_week = DateUtils.is_last_week_of_month()
        
        # Collect debug info for last week of month
        debug_info = {"pass": [], "fail": [], "details": {}}
        
        for user in self.users.values():
            # Reset checkin status trước khi tính score
            if is_last_week:
                # Chỉ khi ở tuần cuối cùng mới check 2 tuần criteria
                criteria_details = self._get_monthly_weekly_criteria_details(user.user_id)
                meets_criteria = criteria_details['meets_criteria']
                user.checkin = 1 if meets_criteria else 0
                
                # Collect detailed debug info
                debug_info["details"][user.name] = criteria_details
                if meets_criteria:
                    debug_info["pass"].append(user.name)
                else:
                    debug_info["fail"].append(user.name)
            else:
                # Các tuần khác không tính điểm checkin
                user.checkin = 0
            
            user.calculate_score()
        
        # Display debug info in expander (only during last week)
        if is_last_week and (debug_info["pass"] or debug_info["fail"]):
            with st.expander(f"🔍 Chi tiết kiểm tra điều kiện checkin ({len(debug_info['pass']) + len(debug_info['fail'])} người)"):
                # Hiển thị tóm tắt
                if debug_info["pass"]:
                    pass_users_info = []
                    for user_name in debug_info["pass"]:
                        details = debug_info["details"][user_name]
                        if details['checkins_count'] > 5:
                            pass_users_info.append(f"{user_name} (>5 checkin)")
                        else:
                            pass_users_info.append(f"{user_name} (≥2 tuần)")
                    st.success(f"✅ **Đạt điều kiện ({len(debug_info['pass'])} người)**: {', '.join(pass_users_info)}")
                if debug_info["fail"]:
                    st.warning(f"⚠️ **Chưa đạt điều kiện ({len(debug_info['fail'])} người)**: {', '.join(debug_info['fail'])}")
                
                # Hiển thị chi tiết cho một vài người đầu tiên
                st.markdown("---")
                st.markdown("### 📊 Chi tiết theo tuần")
                
                sample_users = list(debug_info["details"].keys())
                for user_name in sample_users:
                    details = debug_info["details"][user_name]
                    status_icon = "✅" if details['meets_criteria'] else "❌"
                    
                    # Xác định lý do đạt/không đạt
                    reason = ""
                    if details['meets_criteria']:
                        if details['checkins_count'] > 5:
                            reason = f" (Đạt do >5 checkin)"
                        elif details['weeks_with_checkins'] >= 2:
                            reason = f" (Đạt do ≥2 tuần)"
                    
                    st.markdown(f"**{status_icon} {user_name}**: {details['weeks_with_checkins']}/{details['total_weeks_in_month']} tuần có check-in ({details['checkins_count']} lần){reason}")
                    
                    # Hiển thị chi tiết từng tuần
                    week_status = []
                    for week_detail in details['week_details']:
                        if week_detail['has_checkin']:
                            dates_str = ', '.join(week_detail['checkin_dates'])
                            week_status.append(f"🟢 {week_detail['week_range']} ({dates_str})")
                        else:
                            week_status.append(f"⚪ {week_detail['week_range']}")
                    
                    st.markdown(f"   • {' | '.join(week_status)}")
                    st.markdown("")

    def get_users(self) -> List[User]:
        """Return list of all users"""
        return list(self.users.values())
    
    def get_realtime_checkin_preview(self) -> pd.DataFrame:
        """Get real-time preview of check-in scoring without waiting for last week - Using checkin.py logic"""
        preview_data = []
        
        for user in self.users.values():
            criteria_details = self._get_monthly_weekly_criteria_details(user.user_id)
            weeks_count = criteria_details['weeks_with_checkins']
            total_weeks = criteria_details['total_weeks_in_month']
            checkins_count = criteria_details['checkins_count']
            
            # Tính điểm preview dựa trên tiêu chí 2 tuần HOẶC tổng checkin > 5
            meets_criteria = criteria_details['meets_criteria']
            projected_score = 0.5 if meets_criteria else 0
            
            # Xác định cần bao nhiêu tuần hoặc checkin
            if meets_criteria:
                weeks_needed = 0
            else:
                weeks_needed = max(0, 2 - weeks_count)
            
            # Xác định status
            if meets_criteria:
                if checkins_count > 5:
                    status = "✅ Đạt tiêu chí (>5 checkin)"
                else:
                    status = "✅ Đạt tiêu chí (≥2 tuần)"
                status_color = "success"
            elif weeks_needed == 1:
                status = "⚠️ Cần 1 tuần nữa"
                status_color = "warning"
            else:
                status = f"❌ Cần {weeks_needed} tuần nữa"
                status_color = "error"
            
            # Tính phần trăm hoàn thành
            completion_rate = (weeks_count / 2) * 100 if weeks_count <= 2 else 100
            
            preview_data.append({
                'Tên': user.name,
                'Tuần có check-in': f"{weeks_count}/{total_weeks}",
                'Tổng số check-in': checkins_count,
                'Điểm dự kiến': projected_score,
                'Hoàn thành (%)': f"{completion_rate:.0f}%",
                'Trạng thái': status,
                'Cần thêm': weeks_needed if weeks_needed > 0 else 0,
                '_status_color': status_color  # For styling
            })
        
        # Sort by completion rate descending, then by weeks_count descending
        preview_data.sort(key=lambda x: (x['Điểm dự kiến'], x['Tuần có check-in'].split('/')[0]), reverse=True)
        
        return pd.DataFrame(preview_data)
    
    def generate_checkin_alerts(self) -> List[Dict]:
        """Generate alerts for users who need more check-ins"""
        alerts = []
        today = datetime.now()
        current_week = today.isocalendar()[1]
        days_left_in_month = (datetime(today.year, today.month + 1 if today.month < 12 else today.year + 1, 1) - today - timedelta(days=1)).days
        
        for user in self.users.values():
            criteria_details = self._get_monthly_weekly_criteria_details(user.user_id)
            weeks_count = criteria_details['weeks_with_checkins']
            checkins_count = criteria_details['checkins_count']
            meets_criteria = criteria_details['meets_criteria']
            
            # Chỉ cảnh báo nếu chưa đạt cả 2 tiêu chí
            if not meets_criteria:
                weeks_needed = 2 - weeks_count
                checkins_needed = max(0, 6 - checkins_count)
                
                # Xác định mức độ khẩn cấp
                if days_left_in_month <= 7:
                    urgency = "🔴 KHẨN CẤP"
                    urgency_level = 3
                elif days_left_in_month <= 14:
                    urgency = "🟡 CẦN CHÚ Ý"
                    urgency_level = 2
                else:
                    urgency = "🟢 BÌNH THƯỜNG"
                    urgency_level = 1
                
                # Tạo message phù hợp
                if weeks_needed > 0 and checkins_needed > 0:
                    message = f"Cần check-in thêm {weeks_needed} tuần HOẶC {checkins_needed} lần nữa để đạt 0.5 điểm"
                elif weeks_needed > 0:
                    message = f"Cần check-in thêm {weeks_needed} tuần để đạt 0.5 điểm"
                else:
                    message = f"Cần check-in thêm {checkins_needed} lần để đạt 0.5 điểm"
                
                alerts.append({
                    'user_name': user.name,
                    'user_id': user.user_id,
                    'urgency': urgency,
                    'urgency_level': urgency_level,
                    'weeks_current': weeks_count,
                    'weeks_needed': weeks_needed,
                    'checkins_current': checkins_count,
                    'checkins_needed': checkins_needed,
                    'days_left': days_left_in_month,
                    'message': message,
                    'details': criteria_details
                })
        
        # Sort by urgency level (highest first), then by weeks needed (most needed first)
        alerts.sort(key=lambda x: (x['urgency_level'], x['weeks_needed']), reverse=True)
        
        return alerts

    def calculate_weekly_checkin_scores(self):
        """
        Tính điểm checkin theo tuần cho tháng hiện tại
        
        Điều kiện:
        - Có ít nhất 2 tuần check-in trong tháng hiện tại HOẶC tổng số checkin > 5 → +0.5 điểm
        - Không đủ điều kiện trên → +0 điểm
        - Chỉ hiển thị vào tuần cuối cùng của tháng
        """
        
        # Kiểm tra có phải tuần cuối cùng của tháng không
        if not self._is_last_week_of_month():
            print("⚠️  Chỉ hiển thị điểm checkin vào tuần cuối cùng của tháng")
            return None
        
        # Lấy thông tin các tuần trong tháng hiện tại
        current_month_weeks = self._get_weeks_in_current_month_from_checkin_py()
        now = datetime.now()
        current_month_year = f"{now.year}-{now.month:02d}"
        
        print(f"📅 Tính điểm checkin cho tháng {now.month}/{now.year}")
        print(f"📊 Số tuần trong tháng: {len(current_month_weeks)}")
        
        # Hiển thị thông tin các tuần
        print("\n📋 Các tuần trong tháng:")
        for week in current_month_weeks:
            print(f"  Tuần {week['week_number']}: {week['week_range']}")
        
        # Process all users with final_df data (equivalent to checkins_df in checkin.py)
        if self.final_df is None or self.final_df.empty:
            print(f"\n⚠️  Không có checkin data")
            return pd.DataFrame()
        
        # Filter checkins for current month
        checkins_df = self.final_df.copy()
        checkins_df['checkin_date'] = pd.to_datetime(checkins_df['checkin_since']).dt.date
        checkins_df['checkin_month_year'] = pd.to_datetime(checkins_df['checkin_since']).dt.strftime('%Y-%m')
        
        current_month_checkins = checkins_df[checkins_df['checkin_month_year'] == current_month_year].copy()
        
        if current_month_checkins.empty:
            print(f"\n⚠️  Không có checkin nào trong tháng {now.month}/{now.year}")
            return pd.DataFrame()
        
        print(f"\n📊 Tổng checkins trong tháng: {len(current_month_checkins)}")
        
        # Xác định checkin thuộc tuần nào
        def get_week_number(checkin_date):
            for week in current_month_weeks:
                if week['start_date'] <= checkin_date <= week['end_date']:
                    return week['week_number']
            return None
        
        current_month_checkins['week_number'] = current_month_checkins['checkin_date'].apply(get_week_number)
        
        # Đếm số tuần có checkin cho mỗi user
        user_weekly_checkins = current_month_checkins.groupby(['goal_user_name', 'week_number']).size().reset_index(name='checkins_count')
        user_weeks_with_checkins = user_weekly_checkins.groupby('goal_user_name')['week_number'].nunique().reset_index(name='weeks_with_checkins')
        
        # Tính tổng số checkin cho mỗi user
        user_total_checkins = current_month_checkins.groupby('goal_user_name').size().reset_index(name='total_checkins')
        user_weeks_with_checkins = user_weeks_with_checkins.merge(user_total_checkins, on='goal_user_name', how='left')
        user_weeks_with_checkins['total_checkins'] = user_weeks_with_checkins['total_checkins'].fillna(0)
        
        # Tính điểm dựa trên >= 2 tuần HOẶC > 5 checkins
        user_weeks_with_checkins['score'] = user_weeks_with_checkins.apply(
            lambda row: 0.5 if (row['weeks_with_checkins'] >= 2 or row['total_checkins'] > 5) else 0.0,
            axis=1
        )
        
        # Thêm thông tin chi tiết
        user_stats = []
        for _, user_row in user_weeks_with_checkins.iterrows():
            user_name = user_row['goal_user_name']
            weeks_with_checkins = user_row['weeks_with_checkins']
            total_checkins = int(user_row['total_checkins'])
            score = user_row['score']
            
            # Lấy thông tin chi tiết các tuần
            user_week_details = user_weekly_checkins[user_weekly_checkins['goal_user_name'] == user_name]
            weeks_list = sorted(user_week_details['week_number'].tolist())
            
            # Xác định status dựa trên điều kiện
            if score > 0:
                if total_checkins > 5:
                    status = 'ĐẠT (>5 checkin)'
                else:
                    status = 'ĐẠT (≥2 tuần)'
            else:
                status = 'KHÔNG ĐẠT'
            
            user_stats.append({
                'user_name': user_name,
                'total_checkins_in_month': total_checkins,
                'weeks_with_checkins': weeks_with_checkins,
                'weeks_list': weeks_list,
                'weeks_detail': ', '.join([f"T{w}" for w in weeks_list]),
                'score': score,
                'status': status
            })
        
        result_df = pd.DataFrame(user_stats)
        result_df = result_df.sort_values(['score', 'weeks_with_checkins'], ascending=[False, False])
        
        return result_df

    def _is_last_week_of_month(self):
        """
        Exact copy of is_last_week_of_month from checkin.py
        Kiểm tra xem hiện tại có phải tuần cuối cùng của tháng không
        """
        now = datetime.now()
        weeks = self._get_weeks_in_current_month_from_checkin_py()
        
        if not weeks:
            return False
        
        last_week = weeks[-1]
        return last_week['start_date'] <= now.date() <= last_week['end_date']

    def display_weekly_scoring_results(self, result_df: pd.DataFrame):
        """
        Exact copy of display_weekly_scoring_results from checkin.py
        Hiển thị kết quả tính điểm theo tuần
        """
        if result_df is None:
            return
            
        if result_df.empty:
            print("📊 Không có dữ liệu để hiển thị")
            return
        
        print("\n" + "="*80)
        print("🏆 KẾT QUẢ ĐIỂM CHECKIN THEO TUẦN - THÁNG HIỆN TẠI")
        print("="*80)
        
        # Thống kê tổng quan
        total_users = len(result_df)
        users_passed = len(result_df[result_df['score'] > 0])
        users_failed = total_users - users_passed
        
        print(f"👥 Tổng nhân viên: {total_users}")
        print(f"✅ Đạt điều kiện (≥2 tuần): {users_passed} người ({users_passed/total_users*100:.1f}%)")
        print(f"❌ Không đạt (<2 tuần): {users_failed} người ({users_failed/total_users*100:.1f}%)")
        
        # Hiển thị chi tiết từng nhân viên
        print(f"\n📋 CHI TIẾT TỪNG NHÂN VIÊN:")
        print("-" * 80)
        
        for i, row in result_df.iterrows():
            status_icon = "✅" if row['status'] == 'ĐẠT' else "❌"
            print(f"{status_icon} {row['user_name']}")
            print(f"   📊 Số tuần checkin: {row['weeks_with_checkins']}/≥2 tuần")
            print(f"   📅 Các tuần: {row['weeks_detail']}")
            print(f"   📈 Tổng checkins: {row['total_checkins_in_month']}")
            print(f"   🎯 Điểm số: +{row['score']} điểm")
            print()

    def get_monthly_checkin_report(self):
        """
        Exact copy of get_monthly_checkin_report from checkin.py
        Tạo báo cáo checkin theo tháng với điểm số
        """
        # Tính điểm weekly
        weekly_scores = self.calculate_weekly_checkin_scores()
        
        if weekly_scores is not None and not weekly_scores.empty:
            # Hiển thị kết quả
            self.display_weekly_scoring_results(weekly_scores)
            return weekly_scores
        
        return None

    def demo_weekly_scoring(self):
        """
        Exact copy of demo_weekly_scoring from checkin.py
        Demo function để test logic tính điểm
        """
        print("🔧 DEMO: Weekly Checkin Scoring System")
        
        # Hiển thị thông tin tuần trong tháng hiện tại
        weeks = self._get_weeks_in_current_month_from_checkin_py()
        now = datetime.now()
        
        print(f"\n📅 Tháng hiện tại: {now.month}/{now.year}")
        print(f"📊 Số tuần: {len(weeks)}")
        
        for week in weeks:
            print(f"  Tuần {week['week_number']}: {week['week_range']}")
        
        # Kiểm tra có phải tuần cuối không
        is_last_week = self._is_last_week_of_month()
        print(f"\n⏰ Hiện tại có phải tuần cuối cùng: {'CÓ' if is_last_week else 'KHÔNG'}")
        
        if not is_last_week:
            print("⚠️  Chỉ hiển thị điểm vào tuần cuối cùng của tháng")


class APIClient:
    """Client for handling API requests"""
    
    def __init__(self, goal_token: str, account_token: str):
        self.goal_token = goal_token
        self.account_token = account_token

    def _make_request(self, url: str, data: Dict, description: str = "") -> requests.Response:
        """Make HTTP request with error handling and retries"""
        max_retries = 3
        retry_delay = 1

        for attempt in range(max_retries):
            try:
                response = requests.post(url, data=data, timeout=REQUEST_TIMEOUT)
                response.raise_for_status()
                return response
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    st.error(f"Error {description}: {e}")
                    raise
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff

    def get_filtered_members(self) -> pd.DataFrame:
        """Get filtered members from account API"""
        url = "https://account.base.vn/extapi/v1/group/get"
        data = {'access_token_v2': self.account_token, "path": "nvvanphong"}
        
        response = self._make_request(url, data, "fetching account members")
        response_data = response.json()
        
        members = response_data.get('group', {}).get('members', [])
        
        df = pd.DataFrame([
            {
                'id': str(m.get('id', '')),
                'name': m.get('name', ''),
                'username': m.get('username', ''),
                'job': m.get('title', ''),
                'email': m.get('email', '')
            }
            for m in members
        ])
        
        return self._apply_member_filters(df)

    def _apply_member_filters(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply filters to member dataframe"""
        filtered_df = df
        filtered_df = filtered_df[filtered_df['username'] != 'HienNguyen']
        return filtered_df

    def get_cycle_list(self) -> List[Dict]:
        """Get list of quarterly cycles sorted by most recent first"""
        url = "https://goal.base.vn/extapi/v1/cycle/list"
        data = {'access_token_v2': self.goal_token}

        response = self._make_request(url, data, "fetching cycle list")
        data = response.json()

        quarterly_cycles = []
        for cycle in data.get('cycles', []):
            if cycle.get('metatype') == 'quarterly':
                try:
                    start_time = datetime.fromtimestamp(float(cycle['start_time']), tz=timezone.utc)
                    quarterly_cycles.append({
                        'name': cycle['name'],
                        'path': cycle['path'],
                        'start_time': start_time,
                        'formatted_start_time': start_time.strftime('%d/%m/%Y')
                    })
                except (ValueError, TypeError) as e:
                    st.warning(f"Error parsing cycle {cycle.get('name', 'Unknown')}: {e}")
                    continue

        return sorted(quarterly_cycles, key=lambda x: x['start_time'], reverse=True)

    def get_account_users(self) -> pd.DataFrame:
        """Get users from Account API"""
        url = "https://account.base.vn/extapi/v1/users"
        data = {'access_token_v2': self.account_token}

        response = self._make_request(url, data, "fetching account users")
        json_response = response.json()
        
        if isinstance(json_response, list) and len(json_response) > 0:
            json_response = json_response[0]

        account_users = json_response.get('users', [])
        return pd.DataFrame([{
            'id': str(user['id']),
            'name': user['name'],
            'username': user['username']
        } for user in account_users])

    def get_target_sub_goal_ids(self, target_id: str) -> List[str]:
        """Fetch sub-goal IDs for a specific target"""
        url = "https://goal.base.vn/extapi/v1/target/get"
        data = {'access_token_v2': self.goal_token, 'id': str(target_id)}
        
        try:
            # Removed separate print to reduce noise, handled in loop or debug if needed
            response = requests.post(url, data=data, timeout=REQUEST_TIMEOUT)
            if response.status_code == 200:
                response_data = response.json()
                if response_data and 'target' in response_data and response_data['target']:
                    cached_objs = response_data['target'].get('cached_objs', [])
                    if isinstance(cached_objs, list):
                        return [str(item.get('id')) for item in cached_objs if 'id' in item]
            return []
        except Exception as e:
            print(f"Error fetching sub-goal {target_id}: {e}")
            return []

    def parse_targets_data(self, cycle_path: str) -> pd.DataFrame:
        """Parse targets data from API to create target mapping"""
        url = "https://goal.base.vn/extapi/v1/cycle/get.full"
        data = {'access_token_v2': self.goal_token, 'path': cycle_path}

        response = self._make_request(url, data, "fetching targets data")
        response_data = response.json()
        
        if not response_data or 'targets' not in response_data:
            return pd.DataFrame()
        
        all_targets = []
        raw_targets = response_data.get('targets', [])
        
        # 1. Map Company Targets (Top Level scope='company')
        company_targets_map = {}
        for t in raw_targets:
            if t.get('scope') == 'company':
                company_targets_map[str(t.get('id', ''))] = {
                    'id': str(t.get('id', '')),
                    'name': t.get('name', '')
                }
        
        # Helper to extract form data matching server.py logic
        def extract_form_data(target_obj):
            # strict columns requested by user
            form_data = {
                "Mức độ đóng góp vào mục tiêu công ty": "",
                "Mức độ ưu tiên mục tiêu của Quý": "",
                "Tính khó/tầm ảnh hưởng đến hệ thống": ""
            }
            if 'form' in target_obj and isinstance(target_obj['form'], list):
                for item in target_obj['form']:
                    key = item.get('name')
                    val = item.get('value')
                    if key:
                        form_data[key] = val
            return form_data

        # 2. Iterate ALL targets to find relevant ones
        # We use a dictionary to ensure uniqueness and maximize coverage
        targets_map = {}
        
        for t in raw_targets:
            t_id = str(t.get('id', ''))
            scope = t.get('scope', '')
            parent_id = str(t.get('parent_id') or '')
            
            # Common extraction logic
            def create_base_data(obj, parent_info=None):
                data = {
                    'target_id': str(obj.get('id', '')),
                    'target_name': obj.get('name', ''),
                    'target_scope': obj.get('scope', ''),
                    'target_company_id': parent_info['id'] if parent_info else None,
                    'target_company_name': parent_info['name'] if parent_info else None,
                    'target_dept_id': None, 'target_dept_name': None,
                    'target_team_id': None, 'target_team_name': None,
                    'team_id': str(obj.get('team_id', '')),
                    'dept_id': str(obj.get('dept_id', ''))
                }
                data.update(extract_form_data(obj))
                return data

            # Case A: Detached Dept/Team Target linked to Company Parent
            if scope in ['dept', 'team'] and parent_id in company_targets_map:
                parent = company_targets_map[parent_id]
                target_data = create_base_data(t, parent)
                targets_map[t_id] = target_data

            # Case B: Company Target (inspect its cached_objs)
            elif scope == 'company':
                # Also add the company target itself if not present
                if t_id not in targets_map:
                     targets_map[t_id] = create_base_data(t, {'id': t_id, 'name': t.get('name', '')})

                # Process cached_objs
                if 'cached_objs' in t and isinstance(t['cached_objs'], list):
                    for kr in t['cached_objs']:
                        kr_id = str(kr.get('id', ''))
                        parent_info = {'id': t_id, 'name': t.get('name', '')}
                        sub_data = create_base_data(kr, parent_info)
                        targets_map[kr_id] = sub_data
            
            # Case C: Catch-all for any other target not processed yet
            elif t_id not in targets_map:
                # Try to resolve parent name if possible
                parent_info = None
                if parent_id and parent_id in company_targets_map:
                    parent_info = company_targets_map[parent_id]
                
                target_data = create_base_data(t, parent_info)
                targets_map[t_id] = target_data
        
        collected_targets = list(targets_map.values())

        # 3. Post-process: Fill columns and fetch sub-goals
        for target_data in collected_targets:
            # Fill specific columns based on scope
            if target_data['target_scope'] == 'dept':
                target_data['target_dept_id'] = target_data['target_id']
                target_data['target_dept_name'] = target_data['target_name']
            elif target_data['target_scope'] == 'team':
                target_data['target_team_id'] = target_data['target_id']
                target_data['target_team_name'] = target_data['target_name']
            
            # Fetch sub-goal IDs (Original logic preserved)
            print(f"  Fetching sub-goals for target: {target_data['target_name']}...", end='\r')
            target_data['list_goal_id'] = self.get_target_sub_goal_ids(target_data['target_id'])
            
            all_targets.append(target_data)
        
        print("\nFinished fetching all targets.")
        return pd.DataFrame(all_targets)

    def get_goals_data(self, cycle_path: str) -> pd.DataFrame:
        """Get goals data from API"""
        url = "https://goal.base.vn/extapi/v1/cycle/get.full"
        data = {'access_token_v2': self.goal_token, 'path': cycle_path}

        response = self._make_request(url, data, "fetching goals data")
        data = response.json()

        def extract_form_data(target_obj):
            # strict columns requested by user
            form_data = {
                "Mức độ đóng góp vào mục tiêu công ty": "",
                "Mức độ ưu tiên mục tiêu của Quý": "",
                "Tính khó/tầm ảnh hưởng đến hệ thống": ""
            }
            if 'form' in target_obj and isinstance(target_obj['form'], list):
                for item in target_obj['form']:
                    key = item.get('name')
                    # normalize key? User provided strict keys.
                    # Base might have subtle diffs, but user provided strict names. 
                    # Assuming keys match exactly.
                    val = item.get('value')
                    if key in form_data:
                        form_data[key] = val
            return form_data

        goals_data = []
        for goal in data.get('goals', []):
            # Extract form data
            form_info = extract_form_data(goal)
            
            goals_data.append({
                'goal_id': goal.get('id'),
                'goal_name': goal.get('name', 'Unknown Goal'),
                'goal_content': goal.get('content', ''),
                'goal_since': DateUtils.convert_timestamp_to_datetime(goal.get('since')),
                'goal_current_value': goal.get('current_value', 0),
                'goal_user_id': str(goal.get('user_id', '')),
                'alignment_score_str': form_info.get("Mức độ đóng góp vào mục tiêu công ty", ""),
                'priority_score_str': form_info.get("Mức độ ưu tiên mục tiêu của Quý", ""),
                'impact_score_str': form_info.get("Tính khó/tầm ảnh hưởng đến hệ thống", "")
            })

        return pd.DataFrame(goals_data)

    def get_krs_data(self, cycle_path: str) -> pd.DataFrame:
        """Get KRs data from API with pagination"""
        url = "https://goal.base.vn/extapi/v1/cycle/krs"
        all_krs = []
        
        for page in range(1, MAX_PAGES_KRS + 1):
            data = {'access_token_v2': self.goal_token, "path": cycle_path, "page": page}

            response = self._make_request(url, data, f"loading KRs at page {page}")
            response_data = response.json()

            if isinstance(response_data, list) and len(response_data) > 0:
                response_data = response_data[0]

            krs_list = response_data.get("krs", [])
            if not krs_list:
                break

            for kr in krs_list:
                all_krs.append({
                    'kr_id': str(kr.get('id', '')),
                    'kr_name': kr.get('name', 'Unknown KR'),
                    'kr_content': kr.get('content', ''),
                    'kr_since': DateUtils.convert_timestamp_to_datetime(kr.get('since')),
                    'kr_current_value': kr.get('current_value', 0),
                    'kr_user_id': str(kr.get('user_id', '')),
                    'goal_id': kr.get('goal_id'),
                })

        return pd.DataFrame(all_krs)

    def get_all_checkins(self, cycle_path: str) -> List[Dict]:
        """Get all checkins with pagination"""
        url = "https://goal.base.vn/extapi/v1/cycle/checkins"
        all_checkins = []
        
        for page in range(1, MAX_PAGES_CHECKINS + 1):
            data = {'access_token_v2': self.goal_token, "path": cycle_path, "page": page}

            response = self._make_request(url, data, f"loading checkins at page {page}")
            response_data = response.json()

            if isinstance(response_data, list) and len(response_data) > 0:
                response_data = response_data[0]

            checkins = response_data.get('checkins', [])
            if not checkins:
                break

            all_checkins.extend(checkins)

            if len(checkins) < 20:
                break

        return all_checkins



class AIActionEvaluator:
    """Evaluates 'Next Action' content using AI"""
    
    @staticmethod
    def evaluate_action(action_content: str) -> int:
        """
        Evaluate the quality of the 'Next Action' content.
        
        Criteria:
        - +1: No clear action / Vague / Empty
        - +3: Status report only (doing, trying...)
        - +5: Clear action + Specific solution / Concrete plan
        """
        # TẠM THỜI KHÓA: Return 0 luôn để không gọi AI
        return 0

        # Nếu không có nội dung hoặc quá ngắn (dưới 5 ký tự) -> 0 điểm, không gọi AI
        if not action_content or len(action_content.strip()) < 5:
            return 0
            
        try:
            prompt = f"""
            Bạn là một trợ lý AI đánh giá chất lượng của nội dung "Công việc tiếp theo" trong báo cáo check-in.
            
            Nội dung cần đánh giá: "{action_content}"
            
            Hãy đánh giá dựa trên tiêu chí sau và CHỈ TRẢ VỀ MỘT CON SỐ (1, 3, hoặc 5):
            - 1: Không có hành động rõ ràng, quá ngắn gọn, hoặc vô nghĩa.
            - 3: Chỉ báo cáo trạng thái (đang làm, đang cố gắng, vẫn thế...) mà không có giải pháp cụ thể.
            - 5: Có hành động rõ ràng, cụ thể, và hướng giải quyết/kế hoạch chi tiết.
            
            Output chỉ là số:
            """
            
            response = ollama.generate(
                model='gemini-3-flash-preview:cloud',
                prompt=prompt
            )
            
            result_text = response['response'].strip()
            
            # Extract number from response (handling potential extra text)
            import re
            match = re.search(r'\b(1|3|5)\b', result_text)
            if match:
                return int(match.group(1))
            
            # Fallback simple check if regex fails but simple number exists
            if '5' in result_text: return 5
            if '3' in result_text: return 3
            if '1' in result_text: return 1
            
            return 1 # Default fallback
            
        except Exception as e:
            print(f"AI Eval Error: {e}")
            return 1 # Fallback on error


class DataProcessor:
    """Handles data processing and transformations"""
    
    @staticmethod
    def extract_checkin_data(all_checkins: List[Dict]) -> pd.DataFrame:
        """Extract checkin data into DataFrame"""
        checkin_list = []

        for checkin in all_checkins:
            try:
                checkin_id = checkin.get('id', '')
                checkin_name = checkin.get('name', '')
                user_id = str(checkin.get('user_id', ''))
                since_timestamp = checkin.get('since', '')
                since_date = DataProcessor._convert_timestamp_to_datetime(since_timestamp) or ''
                
                form_data = checkin.get('form', [])
                form_value = form_data[0].get('value', '') if form_data else ''
                
                obj_export = checkin.get('obj_export', {})
                target_name = obj_export.get('name', '') if obj_export else ''
                kr_id = str(obj_export.get('id', '')) if obj_export else ''
                current_value = checkin.get('current_value', 0)
                
                checkin_list.append({
                    'checkin_id': checkin_id,
                    'checkin_name': checkin_name,
                    'checkin_since': since_date,
                    'checkin_since_timestamp': since_timestamp,
                    'cong_viec_tiep_theo': form_value,
                    'checkin_target_name': target_name,
                    'checkin_kr_current_value': current_value,
                    'kr_id': kr_id,
                    'checkin_user_id': user_id,
                    'next_action_score': AIActionEvaluator.evaluate_action(form_value)
                })
                
            except Exception as e:
                print(f"Warning: Error processing checkin {checkin.get('id', 'Unknown')}: {e}")
                continue

        if not checkin_list:
            return pd.DataFrame(columns=[
                'checkin_id', 'checkin_name', 'checkin_since', 'checkin_since_timestamp',
                'cong_viec_tiep_theo', 'checkin_target_name', 'checkin_kr_current_value',
                'kr_id', 'checkin_user_id', 'next_action_score'
            ])

        return pd.DataFrame(checkin_list)

    @staticmethod
    def _convert_timestamp_to_datetime(timestamp):
        """Convert timestamp to datetime string in Asia/Ho_Chi_Minh timezone"""
        if timestamp is None or timestamp == '' or timestamp == 0:
            return None
        try:
            # Chuyển timestamp về UTC datetime
            dt_utc = datetime.fromtimestamp(int(timestamp), tz=timezone.utc)
            # Chuyển sang timezone Asia/Ho_Chi_Minh
            tz_hcm = pytz.timezone('Asia/Ho_Chi_Minh')
            dt_hcm = dt_utc.astimezone(tz_hcm)
            return dt_hcm.strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            return None

    @staticmethod
    def clean_final_data(df: pd.DataFrame) -> pd.DataFrame:
        """Clean and prepare final dataset"""
        try:
            df['kr_current_value'] = pd.to_numeric(df['kr_current_value'], errors='coerce').fillna(0.00)
            df['checkin_kr_current_value'] = pd.to_numeric(df['checkin_kr_current_value'], errors='coerce').fillna(0.00)
            
            # Fill NaN next_action_score with 0 (no checkin or failed eval)
            if 'next_action_score' in df.columns:
                df['next_action_score'] = pd.to_numeric(df['next_action_score'], errors='coerce').fillna(0).astype(int)

            df['kr_since'] = df['kr_since'].fillna(df['goal_since'])
            df['checkin_since'] = df['checkin_since'].fillna(df['kr_since'])

            columns_to_drop = ['kr_user_id']
            existing_columns_to_drop = [col for col in columns_to_drop if col in df.columns]
            if existing_columns_to_drop:
                df = df.drop(columns=existing_columns_to_drop)

            return df
        except Exception as e:
            print(f"Error cleaning data: {e}")
            return df


class OKRCalculator:
    """Handles OKR calculations and analysis"""
    
    @staticmethod
    def calculate_current_value(df: pd.DataFrame) -> float:
        """Calculate current OKR value using average of goal_current_value for unique goal_names"""
        try:
            unique_goals = df.groupby('goal_name')['goal_current_value'].first().reset_index()
            unique_goals['goal_current_value'] = pd.to_numeric(unique_goals['goal_current_value'], errors='coerce').fillna(0)
            return unique_goals['goal_current_value'].mean() if len(unique_goals) > 0 else 0
        except Exception as e:
            st.error(f"Error calculating current value: {e}")
            return 0

    @staticmethod
    def calculate_reference_value(reference_date: datetime, df: pd.DataFrame) -> Tuple[float, List[Dict]]:
        """Calculate OKR value as of reference date (works for both Friday and month-end)"""
        try:
            df = df.copy()
            df['checkin_since_dt'] = pd.to_datetime(df['checkin_since'], errors='coerce')

            unique_krs = df['kr_id'].dropna().unique()
            goal_values_dict = {}
            kr_details = []

            for kr_id in unique_krs:
                kr_data = df[df['kr_id'] == kr_id].copy()
                
                actual_checkins_before_date = kr_data[
                    (kr_data['checkin_since_dt'] <= reference_date) &
                    (kr_data['checkin_name'].notna()) &
                    (kr_data['checkin_name'] != '')
                ]

                goal_name = kr_data.iloc[0]['goal_name'] if len(kr_data) > 0 else f"Unknown_{kr_id}"

                if len(actual_checkins_before_date) > 0:
                    latest_checkin = actual_checkins_before_date.sort_values('checkin_since_dt').iloc[-1]
                    kr_value = pd.to_numeric(latest_checkin['checkin_kr_current_value'], errors='coerce')
                    kr_value = kr_value if not pd.isna(kr_value) else 0

                    if goal_name not in goal_values_dict:
                        goal_values_dict[goal_name] = []
                    goal_values_dict[goal_name].append(kr_value)

                    kr_details.append({
                        'kr_id': kr_id,
                        'goal_name': goal_name,
                        'kr_value': kr_value,
                        'checkin_date': latest_checkin['checkin_since_dt'],
                        'source': f'checkin_before_{reference_date.strftime("%Y%m%d")}'
                    })
                else:
                    goal_key = f"{goal_name}_no_checkin_{kr_id}"
                    goal_values_dict[goal_key] = [0]

                    kr_details.append({
                        'kr_id': kr_id,
                        'goal_name': goal_name,
                        'kr_value': 0,
                        'checkin_date': None,
                        'source': f'no_checkin_before_{reference_date.strftime("%Y%m%d")}'
                    })

            # Re-normalize: Group by Goal Name proper to handle multiple KRs per goal correctly
            final_goal_values = defaultdict(list)
            for item in kr_details:
                final_goal_values[item['goal_name']].append(item['kr_value'])
            
            final_averages = []
            for g_name, vals in final_goal_values.items():
                final_averages.append(np.mean(vals))

            reference_value = np.mean(final_averages) if final_averages else 0
            
            return reference_value, kr_details

        except Exception as e:
            st.error(f"Error calculating reference value: {e}")
            return 0, []

    @staticmethod
    def calculate_kr_shift(row: pd.Series, reference_date: datetime, final_df: pd.DataFrame) -> float:
        """Calculate shift for a single KR"""
        try:
            # Lấy thông tin KR
            kr_id = row['kr_id']
            current_val = pd.to_numeric(row.get('kr_current_value'), errors='coerce')
            current_val = current_val if not pd.isna(current_val) else 0.0

            if not kr_id:
                return current_val

            quarter_start = DateUtils.get_quarter_start_date()

            # Lấy lịch sử checkin từ final_df của toàn hệ thống
            # Cần lọc đúng KR đang xét
            kr_history = final_df[final_df['kr_id'] == kr_id].copy()
            if kr_history.empty:
                 return current_val

            kr_history['checkin_since_dt'] = pd.to_datetime(kr_history['checkin_since'], errors='coerce')

            # Lọc checkin trước ngày mốc VÀ sau đầu quý
            valid_checkins = kr_history[
                (kr_history['checkin_since_dt'] <= reference_date) & 
                (kr_history['checkin_since_dt'] >= quarter_start) &
                (kr_history['checkin_name'].notna()) & 
                (kr_history['checkin_name'] != '')
            ]

            reference_val = 0.0
            if not valid_checkins.empty:
                # Lấy checkin mới nhất trước mốc thời gian
                latest_checkin = valid_checkins.sort_values('checkin_since_dt').iloc[-1]
                val = pd.to_numeric(latest_checkin.get('checkin_kr_current_value'), errors='coerce')
                reference_val = val if not pd.isna(val) else 0.0
            
            # Shift = Giá trị hiện tại - Giá trị tại mốc
            return current_val - reference_val

        except Exception as e:
            # st.warning(f"Error calculating kr_shift: {e}")
            return 0.0



class OKRAnalysisSystem:
    """Hệ thống phân tích OKR toàn diện"""
    
    def __init__(self, goal_token: str, account_token: str, checkin_path: str = None):
        self.goal_token = goal_token
        self.account_token = account_token
        self.checkin_path = checkin_path
        self.api_client = APIClient(goal_token, account_token)
        self.checkin_data = None
        self.goals_data = None
        self.krs_data = None
        self.target_data = None
        self.filtered_members_df = None
        
    def get_cycle_list(self) -> List[Dict]:
        """Lấy danh sách chu kỳ"""
        return self.api_client.get_cycle_list()
        
    def load_and_process_data(self):
        """Tải và xử lý dữ liệu từ API"""
        print(f"Loading data for cycle: {self.checkin_path}...")
        
        # 1. Load data from API
        goals_df = self.api_client.get_goals_data(self.checkin_path)
        krs_df = self.api_client.get_krs_data(self.checkin_path)
        
        # Parse targets and assign to self.target_df
        self.target_df = self.api_client.parse_targets_data(self.checkin_path)
        
        # Fetch checkins
        all_checkins = self.api_client.get_all_checkins(self.checkin_path)
        checkin_df = DataProcessor.extract_checkin_data(all_checkins)
        
        # Fetch account members
        account_df = self.api_client.get_filtered_members()
        self.filtered_members_df = account_df.copy() # Store for later use
        
        # 2. Process Member Filtering (logic from main checkin.py)
        # Assuming account_df is already filtered by get_filtered_members
        
        # Filter Goal and KR data to only include users in the filtered account list
        # Map user IDs to usernames first
        all_users_df = self.api_client.get_account_users()
        # Create map: id -> username
        user_map = dict(zip(all_users_df['id'], all_users_df['username']))
        user_name_map = dict(zip(all_users_df['id'], all_users_df['name'])) # id -> Full Name

        # Enrich DataFrames with username
        goals_df['user_id'] = goals_df['goal_user_id'] # alias
        goals_df['username'] = goals_df['user_id'].map(user_map)
        
        krs_df['user_id'] = krs_df['kr_user_id'] # alias
        krs_df['username'] = krs_df['user_id'].map(user_map)
        
        checkin_df['user_id'] = checkin_df['checkin_user_id']
        checkin_df['username'] = checkin_df['user_id'].map(user_map)
        
        # Filter by filtered_members_df
        valid_usernames = self.filtered_members_df['username'].tolist()
        
        self.goals_data = goals_df[goals_df['username'].isin(valid_usernames)].copy()
        self.krs_data = krs_df[krs_df['username'].isin(valid_usernames)].copy()
        self.checkin_data = checkin_df[checkin_df['username'].isin(valid_usernames)].copy()
        
        # Clean Final Data (Merge logic) - Similar to create_final_dataframe in legacy code
        # We need a unified DF for calculations
        # Merge Goals and KRs
        merged_df = pd.merge(
            self.krs_data, 
            self.goals_data[['goal_id', 'goal_name', 'goal_user_id', 'goal_since', 'username']], 
            on='goal_id', 
            how='left',
            suffixes=('_kr', '_goal')
        )
        
        # Merge with Checkins
        # Expand checkins for KRs
        # Right join on checkins? Or Left join on KRs?
        # We want to enable analysis of KRs with and without checkins
        final_df = pd.merge(
            merged_df,
            self.checkin_data,
            on='kr_id',
            how='left'
        )
        
        # Clean and Prepare
        self.final_df = DataProcessor.clean_final_data(final_df)
        
        # Store for internal use
        self.user_map = user_map
        self.user_name_map = user_name_map
        
        print("Data loaded and processed successfully.")
        return self.final_df

    def analyze_missing_goals_and_checkins(self) -> Tuple[List[str], List[str], List[str]]:
        """Phân tích người dùng thiếu mục tiêu và check-in"""
        if self.filtered_members_df is None:
            return [], [], []
            
        all_active_users = set(self.filtered_members_df['username'])
        users_with_goals = set(self.goals_data['username'])
        users_with_checkins = set(self.checkin_data['username'])
        
        # Users without goals
        members_without_goals = list(all_active_users - users_with_goals)
        
        # Users without checkins (but might have goals)
        members_without_checkins = list(all_active_users - users_with_checkins)
        
        # Users with goals but NO checkins (intersect)
        members_with_goals_no_checkins = list(users_with_goals - users_with_checkins)
        
        # Convert to Full Names for reporting
        def to_names(usernames):
            return [self.filtered_members_df[self.filtered_members_df['username'] == u]['name'].iloc[0] for u in usernames if u in self.filtered_members_df['username'].values]
            
        return (
            to_names(members_without_goals),
            to_names(members_without_checkins),
            to_names(members_with_goals_no_checkins)
        )

    def calculate_okr_shifts_by_user(self) -> List[Dict]:
        """Calculate OKR shifts (Weekly)"""
        return self._calculate_okr_shifts_by_period(period_type="weekly")

    def calculate_okr_shifts_by_user_monthly(self) -> List[Dict]:
        """Calculate OKR shifts (Monthly)"""
        return self._calculate_okr_shifts_by_period(period_type="monthly")
    
    def _calculate_okr_shifts_by_period(self, period_type="weekly") -> List[Dict]:
        """Generic OKR shift calculation"""
        if self.final_df is None:
            return []
            
        user_shifts = []
        unique_users = self.final_df['goal_user_name'].unique()
        
        # Determine reference date based on period
        if period_type == "weekly":
            reference_date = DateUtils.get_last_friday()
        else:
            reference_date = DateUtils.get_last_month_end_date()
            
        for user_name in unique_users:
            user_df = self.final_df[self.final_df['goal_user_name'] == user_name]
            
            if period_type == "weekly":
                shift_data = self._calculate_weekly_shift_data(user_name, user_df, reference_date)
            else:
                shift_data = self._calculate_monthly_shift_data(user_name, user_df, reference_date)
            
            if shift_data:
                user_shifts.append(shift_data)
                
        return sorted(user_shifts, key=lambda x: x.get('okr_shift', 0) if period_type=="weekly" else x.get('okr_shift_monthly', 0), reverse=True)

    def _calculate_user_shift_data(self, user_name, user_df, reference_date, period_key_suffix="") -> Dict:
        """Deprecated generic helper - see specific methods below"""
        pass

    def _calculate_weekly_shift_data(self, user_name, user_df, last_friday) -> Dict:
        """Calculate weekly shift metrics"""
        current_value = OKRCalculator.calculate_current_value(user_df)
        last_friday_value, _ = OKRCalculator.calculate_reference_value(last_friday, user_df)
        okr_shift = current_value - last_friday_value
        
        # Lấy user_id từ user_df
        user_id = user_df['checkin_user_id'].iloc[0] if not user_df.empty else None
        
        return {
            'user_name': user_name,
            'user_id': user_id,
            'current_value': current_value,
            'last_friday_value': last_friday_value,
            'okr_shift': okr_shift,
            'kr_details_count': user_df['kr_id'].nunique()
        }

    def _calculate_monthly_shift_data(self, user_name, user_df, last_month_end) -> Dict:
        """Calculate monthly shift metrics"""
        current_value = OKRCalculator.calculate_current_value(user_df)
        
        # Logic đặc biệt cho tháng đầu quý (1, 4, 7, 10)
        # Nếu đang ở tháng đầu quý, last_month_value = 0 (vì so với cuối quý trước, reset OKR)
        # Hoặc logic khác tùy business rule. Ở đây giả sử so với 0.
        
        # Logic checkin.py:
        # Nếu tháng hiện tại là tháng đầu quý (1, 4, 7, 10)
        # Thì Monthly Shift = Current Value (vì đầu kỳ là 0)
        
        if DateUtils.should_calculate_monthly_shift(): # Luôn trả về True theo code mới
            # Nhưng cần check xem last_month_end có phải là tháng của quý trước không?
            # Hàm get_last_month_end_date trả về ngày cuối của tháng trước.
            # Nếu hôm nay là tháng 4, last_month_end là 31/3.
            
            # Logic mới theo yêu cầu: Tháng 1, 4, 7, 10 -> dịch chuyển là chính giá trị hiện tại
            current_month = datetime.now().month
            if current_month in QUARTER_START_MONTHS:
                 last_month_value = 0
            else:
                 last_month_value, _ = OKRCalculator.calculate_reference_value(last_month_end, user_df)
        else:
             last_month_value, _ = OKRCalculator.calculate_reference_value(last_month_end, user_df)

        okr_shift_monthly = current_value - last_month_value
        
        # Lấy user_id từ user_df
        user_id = user_df['checkin_user_id'].iloc[0] if not user_df.empty else None

        return {
            'user_name': user_name,
            'user_id': user_id,
            'current_value': current_value,
            'last_month_value': last_month_value,
            'okr_shift_monthly': okr_shift_monthly,
            'kr_details_count': user_df['kr_id'].nunique()
        }

    def _calculate_final_okr_goal_shift(self, user_df, reference_date) -> float:
        """Calculate shift aggregated by Goal first (average KRs shift), then average Goals"""
        # Get unique combinations of Goal+KR
        # Calculate shift for each KR
        
        # Logic from server.py _calculate_final_okr_goal_shift_monthly is complex
        # For simplicity and consistency with verify_excel.py logic:
        # We assume _calculate_weekly_shift_data logic is sufficient (Mean of Goals current - Mean of Goals past)
        # But if we need exact per-KR shift aggregation:
        
        # This function might be redundant if we use OKRCalculator's approach
        return 0

    def analyze_checkin_behavior(self) -> Tuple[List[Dict], List[Dict]]:
        """Phân tích hành vi check-in"""
        if self.checkin_data is None:
            return [], []
            
        # 1. Period Checkins (This Week / Month)
        # Dùng logic tuần trước (Last Week) cho báo cáo tuần
        last_friday = DateUtils.get_last_friday()
        # Định nghĩa "tuần này" là từ sau last_friday đến hiện tại?
        # Hay "tuần trước" là từ T2-CN tuần trước?
        
        # Theo yêu cầu report: "Checkin tuần trước" - là số lượng checkin thực hiện trong tuần trước
        today = datetime.now()
        days_since_monday = today.weekday()
        monday_this_week = today - timedelta(days=days_since_monday)
        monday_last_week = monday_this_week - timedelta(days=7)
        sunday_last_week = monday_last_week + timedelta(days=6)
        
        # Filter checkins in last week range
        checkins_last_week = self.checkin_data[
            (pd.to_datetime(self.checkin_data['checkin_since_timestamp'], unit='s') >= monday_last_week) &
            (pd.to_datetime(self.checkin_data['checkin_since_timestamp'], unit='s') <= sunday_last_week)
        ]
        
        user_checkins_last_week = checkins_last_week.groupby('username').size().reset_index(name='checkin_count_period')
        user_checkins_last_week['checkin_rate_period'] = user_checkins_last_week['checkin_count_period'] # Simple count
        
        period_checkins = user_checkins_last_week.to_dict('records')
        
        # 2. Overall Checkins (Quarter to Date)
        # Tính từ đầu quý
        quarter_start = DateUtils.get_quarter_start_date()
        checkins_quarter = self.checkin_data[
            pd.to_datetime(self.checkin_data['checkin_since_timestamp'], unit='s') >= quarter_start
        ]
        
        user_total_checkins = checkins_quarter.groupby('username').size().reset_index(name='total_checkins')
        
        # Merge with last week data
        overall_data = pd.merge(user_total_checkins, user_checkins_last_week, on='username', how='left').fillna(0)
        overall_data.rename(columns={'checkin_count_period': 'last_week_checkins'}, inplace=True)
        
        # Calculate Frequency
        weeks_passed = (today - quarter_start).days / 7
        weeks_passed = max(1, weeks_passed)
        overall_data['checkin_frequency_per_week'] = overall_data['total_checkins'] / weeks_passed
        
        # Rename username to user_name for consistency
        overall_data.rename(columns={'username': 'user_name'}, inplace=True)
        
        overall_checkins = overall_data.to_dict('records')
        
        return period_checkins, overall_checkins

    def analyze_alignment_contribution(self) -> Dict:
        """Phân tích sự đóng góp và liên kết"""
        if self.final_df is None:
            return {}
            
        alignment_data = {}
        unique_users = self.final_df['goal_user_name'].unique()
        
        for user_name in unique_users:
            user_df = self.final_df[self.final_df['goal_user_name'] == user_name]
            
            # Form fields
            alignment_scores = user_df['Mức độ đóng góp vào mục tiêu công ty'].replace('', '1').astype(float)
            priority_scores = user_df['Mức độ ưu tiên mục tiêu của Quý'].replace('', '1').astype(float)
            impact_scores = user_df['Tính khó/tầm ảnh hưởng đến hệ thống'].replace('', '1').astype(float)
            
            alignment_data[user_name] = {
                'avg_alignment': alignment_scores.mean(),
                'avg_priority': priority_scores.mean(),
                'avg_impact': impact_scores.mean()
            }
            
        return alignment_data

    def generate_comprehensive_okr_report(self) -> Dict:
        """Tạo báo cáo tổng hợp toàn diện"""
        if self.final_df is None:
            self.load_and_process_data()
            
        print("Analyzing OKR performance...")
        
        # 1. Weekly Analysis
        weekly_analysis = self._analyze_weekly_okr_performance()
        
        # 2. Alerts and Warnings
        alerts = self._generate_alerts_and_warnings(weekly_analysis['user_shifts'])
        
        # 3. Organization Health
        health = self._calculate_organization_health(weekly_analysis, alerts)
        
        # 4. Detailed User Analysis
        user_details = self._create_detailed_user_analysis(weekly_analysis['user_shifts'])
        
        # 5. Summary Report
        report = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'cycle_path': self.checkin_path,
            'weekly_okr_analysis': weekly_analysis,
            'alerts_and_warnings': alerts,
            'organization_health': health,
            'detailed_user_analysis': user_details
        }
        
        # Add Executive Summary
        report['summary'] = self._create_summary_report(report)
        
        return report

    def _analyze_weekly_okr_performance(self) -> Dict:
        """Phân tích hiệu suất OKR tuần này"""
        user_shifts = self.calculate_okr_shifts_by_user()
        
        total_users = len(user_shifts)
        users_positive = len([u for u in user_shifts if u['okr_shift'] > 0])
        users_negative = len([u for u in user_shifts if u['okr_shift'] < 0])
        users_neutral = len([u for u in user_shifts if u['okr_shift'] == 0])
        
        avg_shift = np.mean([u['okr_shift'] for u in user_shifts]) if user_shifts else 0
        avg_current_value = np.mean([u['current_value'] for u in user_shifts]) if user_shifts else 0
        
        # Performance Distribution
        performance_dist = {
            'high_performers': len([u for u in user_shifts if u['okr_shift'] >= 20]),
            'medium_performers': len([u for u in user_shifts if 10 <= u['okr_shift'] < 20]),
            'low_performers': len([u for u in user_shifts if u['okr_shift'] < 0])
        }
        
        return {
            'total_users': total_users,
            'users_positive_shift': users_positive,
            'users_negative_shift': users_negative,
            'users_neutral_shift': users_neutral,
            'avg_shift': avg_shift,
            'avg_current_value': avg_current_value,
            'performance_distribution': performance_dist,
            'user_shifts': user_shifts
        }

    def _generate_alerts_and_warnings(self, user_shifts: List[Dict]) -> Dict:
        """Tạo cảnh báo và thông báo"""
        critical_issues = []
        moderate_issues = []
        opportunities = []
        
        # Get missing goals/checkins
        no_goals, no_checkins, goals_no_checkins = self.analyze_missing_goals_and_checkins()
        
        for user in no_goals:
            critical_issues.append({
                'type': 'NO_GOALS',
                'user': user,
                'message': f"Nhân viên {user} chưa thiết lập OKR nào"
            })
            
        for user in no_checkins:
            critical_issues.append({
                'type': 'NO_CHECKINS',
                'user': user,
                'message': f"Nhân viên {user} chưa thực hiện check-in nào"
            })
            
        for user in goals_no_checkins:
            moderate_issues.append({
                'type': 'GOALS_NO_CHECKINS',
                'user': user,
                'message': f"Nhân viên {user} có mục tiêu nhưng chưa check-in"
            })
            
        # Analyze Shifts for Issues
        for user_data in user_shifts:
            name = user_data['user_name']
            shift = user_data['okr_shift']
            
            if shift < -5:
                critical_issues.append({
                    'type': 'NEGATIVE_PROGRESS',
                    'user': name,
                    'message': f"{name}: Tiến độ giảm mạnh ({shift:.2f}%)"
                })
            elif shift < 0:
                moderate_issues.append({
                    'type': 'SLIGHT_NEGATIVE',
                    'user': name,
                    'message': f"{name}: Tiến độ giảm nhẹ ({shift:.2f}%)"
                })
            elif shift >= 20:
                opportunities.append({
                    'type': 'HIGH_PERFORMANCE',
                    'user': name,
                    'message': f"{name}: Hiệu suất xuất sắc (+{shift:.2f}%)"
                })
                
        return {
            'critical_issues': critical_issues,
            'moderate_issues': moderate_issues,
            'improvement_opportunities': opportunities
        }

    def _calculate_organization_health(self, weekly_analysis: Dict, alerts: Dict) -> Dict:
        """Tính điểm sức khỏe tổng thể của tổ chức"""
        # 1. OKR Performance Score (0-40)
        avg_shift = weekly_analysis.get('avg_shift', 0)
        # Target avg shift per week ~1-5%? Normalized to 40
        okr_score = min(40, max(0, (avg_shift + 5) * 4))
        
        # 2. Check-in Compliance Score (0-30)
        total_users = weekly_analysis.get('total_users', 1)
        critical_issues = len(alerts.get('critical_issues', []))
        compliance_ratio = max(0, (total_users - critical_issues) / total_users)
        checkin_score = compliance_ratio * 30
        
        # 3. Engagement Score (0-30)
        # Based on positive shifts percentage
        users_positive = weekly_analysis.get('users_positive_shift', 0)
        engagement_ratio = users_positive / total_users if total_users > 0 else 0
        engagement_score = engagement_ratio * 30
        
        overall_health = okr_score + checkin_score + engagement_score
        
        # Trends and Recommendations
        trends = self._analyze_health_trends()
        recommendations = self._generate_health_recommendations(overall_health, okr_score, checkin_score)
        
        return {
            'overall_health_score': round(overall_health, 1),
            'okr_health_score': round(okr_score * 2.5, 1), # Scale to 100
            'checkin_health_score': round(checkin_score * 3.33, 1), # Scale to 100
            'engagement_score': round(engagement_score * 3.33, 1),
            'trends': trends,
            'recommendations': recommendations
        }

    def _analyze_health_trends(self) -> List[str]:
        """Phân tích xu hướng (Placeholder - Requires historical data)"""
        return ["Dữ liệu lịch sử chưa đủ để phân tích xu hướng chi tiết"]

    def _generate_health_recommendations(self, overall: float, okr: float, checkin: float) -> List[str]:
        """Đưa ra khuyến nghị dựa trên điểm sức khỏe"""
        recs = []
        if overall < 60:
            recs.append("Cần họp khẩn cấp để rà soát lại quy trình OKR")
        if checkin < 15: # < 50%
            recs.append("Tổ chức đào tạo lại về tầm quan trọng của Check-in")
        if okr < 10:
            recs.append("Rà soát lại tính khả thi của các mục tiêu")
        return recs

    def _create_detailed_user_analysis(self, user_shifts: List[Dict]) -> List[Dict]:
        """Tạo phân tích chi tiết cho từng user"""
        user_analysis = []
        processed_checkins, overall_checkins = self.analyze_checkin_behavior()
        alignment_data = self.analyze_alignment_contribution()
        
        for shift_data in user_shifts:
            user_name = shift_data['user_name']
            
            # Find checkin data
            user_period_checkin = next((x for x in processed_checkins if x['user_name'] == user_name), {})
            user_overall_checkin = next((x for x in overall_checkins if x['user_name'] == user_name), {})
            user_alignment = alignment_data.get(user_name, {})
            
            # Format Alignment strings
            user_alignment_data = {
                'alignment_score': f"{user_alignment.get('avg_alignment', 0):.1f}/5",
                'priority_score': f"{user_alignment.get('avg_priority', 0):.1f}/5",
                'impact_score': f"{user_alignment.get('avg_impact', 0):.1f}/5"
            }
            
            analysis = {
                'user_name': user_name,
                'okr_performance': {
                    'weekly_shift': shift_data['okr_shift'],
                    'current_progress': shift_data['current_value'],
                    'performance_level': self._classify_performance(shift_data['okr_shift'])
                },
                'checkin_behavior': {
                    'period_checkins': user_period_checkin.get('checkin_count_period', 0),
                    'total_checkins': user_overall_checkin.get('total_checkins', 0),
                    'checkin_rate': user_period_checkin.get('checkin_rate_period', 0),
                    'frequency_per_week': user_overall_checkin.get('checkin_frequency_per_week', 0),
                    'last_week_checkins': user_overall_checkin.get('last_week_checkins', 0)
                },
                'alignment_contribution': user_alignment_data,
                'risk_assessment': self._assess_user_risk(shift_data, user_period_checkin, user_overall_checkin),
                'recommendations': self._generate_user_recommendations(shift_data, user_period_checkin)
            }
            
            user_analysis.append(analysis)
            
        return sorted(user_analysis, key=lambda x: x['okr_performance']['weekly_shift'], reverse=True)

    def _classify_performance(self, shift_value: float) -> str:
        """Phân loại hiệu suất"""
        if shift_value >= 20:
            return 'Xuất sắc'
        elif shift_value >= 10:
            return 'Tốt'
        elif shift_value >= 0:
            return 'Đạt yêu cầu'
        else:
            return 'Cần cải thiện'

    def _assess_user_risk(self, shift_data: Dict, period_checkin: Dict, overall_checkin: Dict) -> Dict:
        """Đánh giá rủi ro của user"""
        risk_score = 0
        risk_factors = []
        
        # Kiểm tra tiến độ OKR
        if shift_data.get('okr_shift', 0) < 0:
            risk_score += 30
            risk_factors.append('Tiến độ OKR âm')
            
        # Kiểm tra số lượng check-in
        if period_checkin.get('checkin_count_period', 0) < 2:
            risk_score += 25
            risk_factors.append('Ít check-in trong kỳ')
            
        # Kiểm tra tần suất check-in
        if overall_checkin.get('checkin_frequency_per_week', 0) < 1:
            risk_score += 20
            risk_factors.append('Tần suất check-in thấp')
            
        # Kiểm tra số KR
        if shift_data.get('kr_details_count', 0) == 0:
            risk_score += 25
            risk_factors.append('Không có KR hoạt động')
            
        # Phân loại rủi ro
        if risk_score >= 60:
            risk_level = 'Cao'
        elif risk_score >= 30:
            risk_level = 'Trung bình'
        else:
            risk_level = 'Thấp'
            
        return {
            'risk_score': risk_score,
            'risk_level': risk_level,
            'risk_factors': risk_factors
        }

    def _generate_user_recommendations(self, shift_data: Dict, period_checkin: Dict) -> List[str]:
        """Tạo đề xuất cho từng user"""
        recommendations = []
        shift = shift_data.get('okr_shift', 0)
        checkins = period_checkin.get('checkin_count_period', 0)
        
        if shift < 0:
            recommendations.append('Tập trung cải thiện tiến độ OKR')
            
        if checkins < 2:
            recommendations.append('Tăng cường tần suất check-in (ít nhất 2 lần/tuần)')
            
        if shift_data.get('kr_details_count', 0) == 0:
            recommendations.append('Thiết lập các KR cụ thể và đo lường được')
            
        if not recommendations:
            recommendations.append('Tiếp tục duy trì hiệu suất tốt')
            
        return recommendations

    def _create_summary_report(self, report: Dict) -> Dict:
        """Tạo báo cáo tổng hợp"""
        summary = {
            'report_generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'key_metrics': {},
            'top_issues': [],
            'highlights': []
        }
        
        # Các chỉ số chính
        weekly_analysis = report.get('weekly_okr_analysis', {})
        alerts = report.get('alerts_and_warnings', {})
        health = report.get('organization_health', {})
        
        summary['key_metrics'] = {
            'total_active_users': weekly_analysis.get('total_users', 0),
            'okr_health_score': health.get('okr_health_score', 0),
            'checkin_health_score': health.get('checkin_health_score', 0),
            'overall_health_score': health.get('overall_health_score', 0),
            'critical_issues': len(alerts.get('critical_issues', [])),
            'moderate_issues': len(alerts.get('moderate_issues', []))
        }
        
        # Các vấn đề hàng đầu
        top_issues = []
        for issue in alerts.get('critical_issues', [])[:5]:
            top_issues.append(f"{issue['type']}: {issue['user']}")
        for issue in alerts.get('moderate_issues', [])[:3]:
            top_issues.append(f"{issue['type']}: {issue['user']}")
        summary['top_issues'] = top_issues
        
        # Điểm nổi bật
        highlights = []
        if health.get('overall_health_score', 0) >= 80:
            highlights.append('Sức khỏe OKR tổng thể rất tốt')
        elif health.get('overall_health_score', 0) >= 65:
            highlights.append('Sức khỏe OKR tổng thể ở mức tốt')
        else:
            highlights.append('Cần cải thiện sức khỏe OKR tổng thể')
            
        perf_dist = weekly_analysis.get('performance_distribution', {})
        if perf_dist.get('high_performers', 0) > 0:
            highlights.append(f'Có {perf_dist["high_performers"]} thành viên xuất sắc')
            
        summary['highlights'] = highlights
        
        return summary


def print_report(report: Dict):
    """Hiển thị báo cáo tổng hợp một cách dễ đọc"""
    summary = report.get('summary', {})
    
    print(f"\n📊 THÔNG TIN TỔNG QUAN")
    print("-" * 50)
    print(f"Thời gian tạo báo cáo: {summary.get('report_generated', 'N/A')}")
    print(f"Tổng số thành viên: {summary.get('key_metrics', {}).get('total_active_users', 0)}")
    print(f"Điểm sức khỏe OKR: {summary.get('key_metrics', {}).get('okr_health_score', 0)}/100")
    print(f"Điểm sức khỏe Check-in: {summary.get('key_metrics', {}).get('checkin_health_score', 0)}/100")
    print(f"Điểm sức khỏe tổng thể: {summary.get('key_metrics', {}).get('overall_health_score', 0)}/100")
    
    # Điểm nổi bật
    highlights = summary.get('highlights', [])
    if highlights:
        print(f"\n✨ ĐIỂM NỔI BẬT:")
        for highlight in highlights:
            print(f"  • {highlight}")
            
    # Các vấn đề hàng đầu
    top_issues = summary.get('top_issues', [])
    if top_issues:
        print(f"\n⚠️  CÁC VẤN ĐỀ HÀNG ĐẦU:")
        for issue in top_issues:
            print(f"  • {issue['message']}")
            
    # Phân tích OKR theo tuần
    weekly_analysis = report.get('weekly_okr_analysis', {})
    if weekly_analysis:
        print(f"\n📈 PHÂN TÍCH OKR THEO TUẦN")
        print("-" * 50)
        print(f"Tổng số người dùng: {weekly_analysis.get('total_users', 0)}")
        print(f"Người dùng có tiến độ tích cực: {weekly_analysis.get('users_positive_shift', 0)}")
        print(f"Người dùng có tiến độ âm: {weekly_analysis.get('users_negative_shift', 0)}")
        print(f"Giá trị thay đổi trung bình: {weekly_analysis.get('avg_shift', 0):.2f}")
        print(f"Giá trị hiện tại trung bình: {weekly_analysis.get('avg_current_value', 0):.2f}")
        
        perf_dist = weekly_analysis.get('performance_distribution', {})
        print(f"\nPhân loại hiệu suất:")
        print(f"  • Xuất sắc (≥20): {perf_dist.get('high_performers', 0)} người")
        print(f"  • Tốt (10-19): {perf_dist.get('medium_performers', 0)} người")
        print(f"  • Đạt yêu cầu (0-9): {len([u for u in report.get('detailed_user_analysis', []) if u['okr_performance']['performance_level'] == 'Đạt yêu cầu'])} người")
        print(f"  • Cần cải thiện (<0): {perf_dist.get('low_performers', 0)} người")
        
    # Cảnh báo và thông báo
    alerts = report.get('alerts_and_warnings', {})
    if alerts:
        print(f"\n🚨 CẢNH BÁO VÀ THÔNG BÁO")
        print("-" * 50)
        
        critical_issues = alerts.get('critical_issues', [])
        if critical_issues:
            print(f"Vấn đề nghiêm trọng ({len(critical_issues)}):")
            for issue in critical_issues[:5]:  # Hiển thị tối đa 5 vấn đề
                print(f"  • {issue['message']}")
                
        moderate_issues = alerts.get('moderate_issues', [])
        if moderate_issues:
            print(f"\nVấn đề vừa phải ({len(moderate_issues)}):")
            for issue in moderate_issues[:5]:  # Hiển thị tối đa 5 vấn đề
                print(f"  • {issue['message']}")
                
        opportunities = alerts.get('improvement_opportunities', [])
        if opportunities:
            print(f"\nCơ hội cải thiện ({len(opportunities)}):")
            for issue in opportunities[:5]:
                print(f"  • {issue['message']}")

class EmailReportGenerator:
    """Generate and send email reports for OKR analysis"""
    
    def __init__(self, smtp_server="smtp.gmail.com", smtp_port=587):
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port

    def create_visual_html_chart(self, data: Dict, chart_type: str, title: str) -> str:
        """Create HTML-based visual charts"""
        if chart_type == "pie":
            return self._create_pie_chart_html(data, title)
        elif chart_type == "bar":
            return self._create_bar_chart_html(data, title)
        return f"<div class='modern-chart'><h3>{title}</h3><p>Loại biểu đồ không được hỗ trợ</p></div>"

    def _create_pie_chart_html(self, data: Dict, title: str) -> str:
        """Create pie chart HTML"""
        total = sum(data.values())
        if total == 0:
            return f"<div class='chart-fallback'><h4>{title}</h4><p>Không có dữ liệu</p></div>"
        
        colors = ['#27AE60', '#E74C3C', '#3498DB', '#F39C12', '#9B59B6']
        
        html = f"""
        <div class='modern-chart'>
            <h3 style='text-align: center; margin-bottom: 30px; color: #2c3e50; font-size: 20px;'>{title}</h3>
            <div style='display: flex; justify-content: center; align-items: center; gap: 40px; padding: 20px;'>
        """
        
        for i, (label, value) in enumerate(data.items()):
            percentage = (value / total * 100) if total > 0 else 0
            color = colors[i % len(colors)]
            circle_size = max(100, min(140, 100 + (value / total * 40)))
            font_size = max(20, min(28, 20 + (value / total * 8)))
            
            html += f"""
            <div style='text-align: center; flex: 1; max-width: 200px;'>
                <div style='width: {circle_size}px; height: {circle_size}px; border-radius: 50%; 
                            background: linear-gradient(135deg, {color}, {color}dd); 
                            margin: 0 auto 15px auto; display: flex; align-items: center; 
                            justify-content: center; color: white; font-weight: bold; 
                            font-size: {font_size}px; box-shadow: 0 8px 25px rgba(0,0,0,0.15);
                            border: 4px solid white; position: relative; overflow: hidden;'>
                    <span style='z-index: 2; position: relative;'>{value}</span>
                    <div style='position: absolute; top: 0; left: 0; right: 0; bottom: 0; 
                                background: rgba(255,255,255,0.1); border-radius: 50%;'></div>
                </div>
                <div style='font-weight: bold; margin-bottom: 8px; color: #2c3e50; font-size: 16px;'>{label}</div>
                <div style='color: #7f8c8d; font-size: 14px; background: #ecf0f1; padding: 4px 12px; 
                            border-radius: 15px; display: inline-block;'>{percentage:.1f}%</div>
            </div>
            """
        
        html += "</div></div>"
        return html

    def _create_bar_chart_html(self, data: Dict, title: str) -> str:
        """Create bar chart HTML"""
        if not data:
            return f"<div class='modern-chart'><h3>{title}</h3><p>Không có dữ liệu</p></div>"
        
        max_value = max(abs(v) for v in data.values())
        
        html = f"""
        <div class='modern-chart'>
            <h3 style='text-align: center; margin-bottom: 25px; color: #2c3e50; font-size: 20px;'>{title}</h3>
            <div style='max-height: 500px; overflow-y: auto; padding: 10px;'>
        """
        
        for name, value in list(data.items()):
            width_pct = (abs(value) / max_value * 100) if max_value > 0 else 0
            color, bg_color, icon = self._get_bar_style(value)
            
            html += f"""
            <div style='margin-bottom: 20px; padding: 15px; background: {bg_color}; 
                        border-radius: 12px; border-left: 4px solid {color};
                        box-shadow: 0 2px 8px rgba(0,0,0,0.05);'>
                <div style='display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;'>
                    <strong style='color: #2c3e50; font-size: 15px;'>{name}</strong>
                    <span style='color: {color}; font-weight: bold; font-size: 16px;'>
                        {icon} {value:.2f}
                    </span>
                </div>
                <div style='background: rgba(255,255,255,0.8); height: 12px; border-radius: 6px; overflow: hidden;'>
                    <div style='background: {color}; height: 100%; width: {width_pct}%; 
                                border-radius: 6px; transition: width 0.3s ease;
                                box-shadow: inset 0 1px 2px rgba(0,0,0,0.1);'></div>
                </div>
            </div>
            """
        
        html += "</div></div>"
        return html

    def _get_bar_style(self, value: float) -> Tuple[str, str, str]:
        """Get style properties for bar chart based on value"""
        if value > 0:
            return '#27AE60', 'rgba(39, 174, 96, 0.1)', '📈'
        elif value < 0:
            return '#E74C3C', 'rgba(231, 76, 60, 0.1)', '📉'
        else:
            return '#F39C12', 'rgba(243, 156, 18, 0.1)', '➡️'

    def _generate_table_html(self, data: List[Dict], headers: List[str], fields: List[str]) -> str:
        """Generate HTML table from data"""
        if not data:
            return "<div style='text-align: center; padding: 20px; background: #f8f9fa; border-radius: 10px; color: #7f8c8d;'><p>📭 Không có dữ liệu</p></div>"
        
        html = "<table><thead><tr>"
        for header in headers:
            html += f"<th>{header}</th>"
        html += "</tr></thead><tbody>"
        
        for i, item in enumerate(data):
            row_class = "even" if i % 2 == 0 else "odd"
            html += f"<tr class='{row_class}'>"
            for field in fields:
                value = item.get(field, "")
                if field == "has_goal":
                    value = "<span style='color: #27AE60; font-weight: bold;'>✅ Có</span>" if value else "<span style='color: #E74C3C; font-weight: bold;'>❌ Không</span>"
                html += f"<td>{value}</td>"
            html += "</tr>"
        
        html += "</tbody></table>"
        return html

    def _generate_okr_table_html(self, data: List[Dict], period_type: str = "weekly") -> str:
        """Generate HTML table for OKR data (weekly or monthly)"""
        if not data:
            return "<div style='text-align: center; padding: 20px; background: #f8f9fa; border-radius: 10px; color: #7f8c8d;'><p>📭 Không có dữ liệu</p></div>"
        
        shift_key = 'okr_shift' if period_type == "weekly" else 'okr_shift_monthly'
        value_key = 'last_friday_value' if period_type == "weekly" else 'last_month_value'
        period_label = "(tuần)" if period_type == "weekly" else "(tháng)"
        
        html = f"""
        <table>
            <thead>
                <tr>
                    <th>👤 Nhân viên</th>
                    <th>📊 Dịch chuyển {period_label}</th>
                    <th>🎯 Giá trị hiện tại</th>
                    <th>📅 Giá trị trước đó</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for i, item in enumerate(data):
            shift_value = item.get(shift_key, 0)
            shift_class = "positive" if shift_value > 0 else "negative" if shift_value < 0 else "neutral"
            shift_icon = "📈" if shift_value > 0 else "📉" if shift_value < 0 else "➡️"
            row_class = "even" if i % 2 == 0 else "odd"
            
            html += f"""
            <tr class='{row_class}'>
                <td><strong>{item.get('user_name', 'Unknown')}</strong></td>
                <td class="{shift_class}">{shift_icon} <strong>{shift_value:.2f}</strong></td>
                <td><span style='color: #3498db; font-weight: 600;'>{item.get('current_value', 0):.2f}</span></td>
                <td><span style='color: #7f8c8d;'>{item.get(value_key, 0):.2f}</span></td>
            </tr>
            """
        
        html += "</tbody></table>"
        return html

    def _generate_checkin_overview_table_html(self, overall_checkins_data: List[Dict]) -> str:
        """Generate HTML table for top overall checkin users"""
        if not overall_checkins_data:
            return "<div style='text-align: center; padding: 20px; background: #f8f9fa; border-radius: 10px; color: #7f8c8d;'><p>📭 Không có dữ liệu</p></div>"
        
        today = datetime.now()
        quarter_start = DateUtils.get_quarter_start_date()
        weeks_in_quarter = (today - quarter_start).days / 7
        weeks_in_quarter = max(weeks_in_quarter, 1)
        
        days_since_monday = today.weekday()
        monday_this_week = today - timedelta(days=days_since_monday)
        monday_last_week = monday_this_week - timedelta(days=7)
        sunday_last_week = monday_last_week + timedelta(days=6)
        
        html = f"""
        <div class="alert alert-info">
            <strong>📅 Tuần trước:</strong> {monday_last_week.strftime('%d/%m/%Y')} - {sunday_last_week.strftime('%d/%m/%Y')}<br>
            <strong>📊 Tần suất checkin:</strong> Tổng checkin ÷ {weeks_in_quarter:.1f} tuần (từ đầu quý đến nay)
        </div>
        <table>
            <thead>
                <tr>
                    <th>🏅 Hạng</th>
                    <th>👤 Nhân viên</th>
                    <th>📊 Tổng checkin</th>
                    <th>⚡ Tần suất/tuần (quý)</th>
                    <th>📅 Checkin tuần trước</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for i, item in enumerate(overall_checkins_data):
            rank_icon = self._get_rank_icon(i)
            frequency = item.get('checkin_frequency_per_week', 0)
            last_week = item.get('last_week_checkins', 0)
            total = item.get('total_checkins', 0)
            
            row_style = self._get_row_style(i)
            frequency_style = "style='color: #27AE60; font-weight: bold;'" if frequency >= 2 else ""
            last_week_style = "style='color: #3498db; font-weight: bold;'" if last_week > 0 else "style='color: #7f8c8d;'"
            
            html += f"""
            <tr {row_style}>
                <td style='text-align: center; font-size: 16px; font-weight: bold;'>{rank_icon}</td>
                <td><strong>{item.get('user_name', 'Unknown')}</strong></td>
                <td style='text-align: center; font-weight: 600; color: #2c3e50;'>{total}</td>
                <td style='text-align: center;' {frequency_style}>{frequency:.2f}</td>
                <td style='text-align: center;' {last_week_style}>{last_week}</td>
            </tr>
            """
        
        # Add summary row
        if overall_checkins_data:
            html += self._generate_checkin_summary_row(overall_checkins_data)
        
        html += "</tbody></table>"
        return html

    def _get_rank_icon(self, index: int) -> str:
        """Get rank icon for position"""
        rank_icons = {0: "🥇", 1: "🥈", 2: "🥉"}
        return rank_icons.get(index, str(index + 1))

    def _get_row_style(self, index: int) -> str:
        """Get row style for table"""
        if index < 3:
            return "style='background: linear-gradient(135deg, #fff9e6, #fffbf0); font-weight: 600;'"
        elif index % 2 == 0:
            return "style='background: #f8f9fa;'"
        return ""

    def _generate_checkin_summary_row(self, data: List[Dict]) -> str:
        """Generate summary row for checkin table"""
        total_checkins_sum = sum(item.get('total_checkins', 0) for item in data)
        avg_frequency = sum(item.get('checkin_frequency_per_week', 0) for item in data) / len(data)
        active_last_week = len([item for item in data if item.get('last_week_checkins', 0) > 0])
        
        return f"""
        <tr style='background: linear-gradient(135deg, #e8f4f8, #f0f8ff); border-top: 2px solid #3498db; font-weight: bold;'>
            <td colspan="2" style='text-align: center; color: #2c3e50;'>📊 TỔNG KẾT TẤT CẢ {len(data)} NHÂN VIÊN</td>
            <td style='text-align: center; color: #3498db;'>{total_checkins_sum}</td>
            <td style='text-align: center; color: #27AE60;'>{avg_frequency:.2f}</td>
            <td style='text-align: center; color: #e74c3c;'>{active_last_week} người</td>
        </tr>
        """

    def _get_email_styles(self) -> str:
        """Get CSS styles for email"""
        return """
        <style>
            body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; color: #2c3e50; max-width: 1200px; margin: 0 auto; padding: 20px; background: #f8f9fa; }
            .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 40px; border-radius: 15px; text-align: center; margin-bottom: 30px; box-shadow: 0 10px 30px rgba(0,0,0,0.15); }
            .header h1 { margin: 0 0 10px 0; font-size: 28px; font-weight: 700; }
            .header h2 { margin: 0 0 10px 0; font-size: 22px; font-weight: 500; opacity: 0.9; }
            .header p { margin: 0; font-size: 16px; opacity: 0.8; }
            .section { background: white; padding: 30px; margin: 25px 0; border-radius: 15px; box-shadow: 0 5px 20px rgba(0,0,0,0.08); border: 1px solid #e9ecef; }
            .section h2 { color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; margin-bottom: 25px; font-size: 22px; }
            .metrics { display: flex; justify-content: space-around; margin: 25px 0; flex-wrap: wrap; gap: 15px; }
            .metric { background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%); padding: 25px; border-radius: 12px; text-align: center; box-shadow: 0 4px 15px rgba(0,0,0,0.08); min-width: 140px; flex: 1; border: 1px solid #e9ecef; }
            .metric-value { font-size: 32px; font-weight: 700; color: #3498db; margin-bottom: 5px; }
            .metric-label { font-size: 14px; color: #7f8c8d; font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; }
            table { width: 100%; border-collapse: collapse; margin: 20px 0; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.05); }
            th { padding: 16px; text-align: left; background: linear-gradient(135deg, #3498db, #2980b9); color: white; font-weight: 600; font-size: 14px; text-transform: uppercase; letter-spacing: 0.5px; }
            td { padding: 14px 16px; border-bottom: 1px solid #ecf0f1; font-size: 14px; }
            tr:nth-child(even) { background: #f8f9fa; }
            tr:hover { background: #e8f4f8; transition: background 0.2s ease; }
            .chart-container { text-align: center; margin: 30px 0; }
            .modern-chart { background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%); padding: 30px; border-radius: 15px; box-shadow: 0 8px 25px rgba(0,0,0,0.1); margin: 25px 0; border: 1px solid #e9ecef; }
            .positive { color: #27AE60; font-weight: bold; }
            .negative { color: #E74C3C; font-weight: bold; }
            .neutral { color: #F39C12; font-weight: bold; }
            .footer { text-align: center; margin-top: 40px; padding: 25px; background: linear-gradient(135deg, #2c3e50, #34495e); color: white; border-radius: 15px; }
            .alert { padding: 18px; margin: 20px 0; border-radius: 10px; border-left: 4px solid; }
            .alert-warning { background: linear-gradient(135deg, #fff3cd, #fef8e6); border-left-color: #f39c12; color: #856404; }
            .alert-info { background: linear-gradient(135deg, #d1ecf1, #e8f5f7); border-left-color: #3498db; color: #0c5460; }
            .monthly-indicator { background: linear-gradient(135deg, #e8f5e8, #f0fff0); border: 2px solid #27AE60; border-radius: 10px; padding: 15px; margin: 20px 0; }
        </style>
        """

    def create_email_content(self, analyzer, selected_cycle, members_without_goals, 
                           members_without_checkins, members_with_goals_no_checkins, 
                           okr_shifts, okr_shifts_monthly=None) -> str:
        """Create HTML email content with all analysis data"""
        
        current_date = datetime.now().strftime("%d/%m/%Y")
        total_members = len(analyzer.filtered_members_df) if analyzer.filtered_members_df is not None else 0
        
        # Calculate statistics
        stats = self._calculate_email_stats(total_members, members_without_goals, 
                                          members_without_checkins, okr_shifts, okr_shifts_monthly)
        
        # Get checkin behavior analysis data
        period_checkins, overall_checkins = analyzer.analyze_checkin_behavior()
        
        # Create visual charts
        charts = self._create_email_charts(stats, okr_shifts, okr_shifts_monthly)
        
        # Generate tables
        tables = self._generate_email_tables(members_without_goals, members_without_checkins,
                                           members_with_goals_no_checkins, okr_shifts, 
                                           okr_shifts_monthly, overall_checkins)
        
        # Build HTML content
        html_content = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            {self._get_email_styles()}
        </head>
        <body>
            {self._generate_email_header(selected_cycle, current_date)}
            {self._generate_overview_section(stats)}
            {self._generate_monthly_indicator(okr_shifts_monthly)}
            {self._generate_checkin_section(tables['checkins'], stats)}
            {self._generate_okr_sections(charts, tables, stats)}
            {self._generate_detailed_sections(tables)}
            {self._generate_email_footer()}
        </body>
        </html>
        """
        
        return html_content

    def _calculate_email_stats(self, total_members: int, members_without_goals: List, 
                              members_without_checkins: List, okr_shifts: List, 
                              okr_shifts_monthly: Optional[List] = None) -> Dict:
        """Calculate statistics for email report"""
        members_with_goals = total_members - len(members_without_goals)
        members_with_checkins = total_members - len(members_without_checkins)
        
        stats = {
            'total_members': total_members,
            'members_with_goals': members_with_goals,
            'members_with_checkins': members_with_checkins,
            'progress_users': len([u for u in okr_shifts if u['okr_shift'] > 0]) if okr_shifts else 0,
            'stable_users': len([u for u in okr_shifts if u['okr_shift'] == 0]) if okr_shifts else 0,
            'issue_users': len([u for u in okr_shifts if u['okr_shift'] < 0]) if okr_shifts else 0
        }
        
        if okr_shifts_monthly:
            stats.update({
                'progress_users_monthly': len([u for u in okr_shifts_monthly if u['okr_shift_monthly'] > 0]),
                'stable_users_monthly': len([u for u in okr_shifts_monthly if u['okr_shift_monthly'] == 0]),
                'issue_users_monthly': len([u for u in okr_shifts_monthly if u['okr_shift_monthly'] < 0])
            })
        
        return stats

    def _create_email_charts(self, stats: Dict, okr_shifts: List, okr_shifts_monthly: Optional[List]) -> Dict:
        """Create charts for email"""
        charts = {}
        
        # Goal distribution chart
        charts['goal'] = self.create_visual_html_chart(
            {'Có OKR': stats['members_with_goals'], 'Chưa có OKR': stats['total_members'] - stats['members_with_goals']},
            'pie', 'Phân bố trạng thái OKR'
        )
        
        # Weekly OKR shifts chart
        if okr_shifts:
            okr_shifts_data = {u['user_name']: u['okr_shift'] for u in okr_shifts}
            charts['okr_weekly'] = self.create_visual_html_chart(
                okr_shifts_data, 'bar', 'Dịch chuyển OKR tuần (Tất cả NV có goal)'
            )
        
        # Monthly OKR shifts chart
        if okr_shifts_monthly:
            okr_shifts_monthly_data = {u['user_name']: u['okr_shift_monthly'] for u in okr_shifts_monthly}
            charts['okr_monthly'] = self.create_visual_html_chart(
                okr_shifts_monthly_data, 'bar', 'Dịch chuyển OKR tháng (Tất cả NV có goal)'
            )
        
        return charts

    def _generate_email_tables(self, members_without_goals: List, members_without_checkins: List,
                              members_with_goals_no_checkins: List, okr_shifts: List,
                              okr_shifts_monthly: Optional[List], overall_checkins: List) -> Dict:
        """Generate all tables for email"""
        return {
            'goals': self._generate_table_html(members_without_goals, 
                                             ["Tên", "Username", "Chức vụ"], 
                                             ["name", "username", "job"]),
            'checkins': self._generate_table_html(members_without_checkins,
                                                ["Tên", "Username", "Chức vụ", "Có OKR"],
                                                ["name", "username", "job", "has_goal"]),
            'goals_no_checkins': self._generate_table_html(members_with_goals_no_checkins,
                                                         ["Tên", "Username", "Chức vụ"],
                                                         ["name", "username", "job"]),
            'top_performers': self._generate_okr_table_html([u for u in okr_shifts if u['okr_shift'] > 0] if okr_shifts else []),
            'top_performers_monthly': self._generate_okr_table_html([u for u in okr_shifts_monthly if u['okr_shift_monthly'] > 0] if okr_shifts_monthly else [], "monthly"),
            'issue_performers': self._generate_okr_table_html([u for u in okr_shifts if u['okr_shift'] < 0] if okr_shifts else []),
            'issue_performers_monthly': self._generate_okr_table_html([u for u in okr_shifts_monthly if u['okr_shift_monthly'] < 0] if okr_shifts_monthly else [], "monthly"),
            'top_overall': self._generate_checkin_overview_table_html(overall_checkins if overall_checkins else [])
        }

    def _generate_email_header(self, selected_cycle: Dict, current_date: str) -> str:
        """Generate email header"""
        return f"""
        <div class="header">
            <h1>📊 BÁO CÁO TIẾN ĐỘ OKR & CHECKIN</h1>
            <h2>{selected_cycle['name']}</h2>
            <p>Ngày báo cáo: {current_date}</p>
        </div>
        """

    def _generate_overview_section(self, stats: Dict) -> str:
        """Generate overview metrics section"""
        monthly_metric = ""
        if 'progress_users_monthly' in stats:
            monthly_metric = f"""
            <div class="metric">
                <div class="metric-value">{stats['progress_users_monthly']}</div>
                <div class="metric-label">Tiến bộ (tháng)</div>
            </div>
            """
        
        return f"""
        <div class="section">
            <h2>📈 TỔNG QUAN</h2>
            <div class="metrics">
                <div class="metric">
                    <div class="metric-value">{stats['total_members']}</div>
                    <div class="metric-label">Tổng nhân viên</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{stats['members_with_goals']}</div>
                    <div class="metric-label">Có OKR</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{stats['members_with_checkins']}</div>
                    <div class="metric-label">Có Checkin</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{stats['progress_users']}</div>
                    <div class="metric-label">Tiến bộ (tuần)</div>
                </div>
                {monthly_metric}
            </div>
        </div>
        """

    def _generate_monthly_indicator(self, okr_shifts_monthly: Optional[List]) -> str:
        """Generate monthly indicator section"""
        if not okr_shifts_monthly:
            return ""
        
        current_month = datetime.now().month
        month_names = {2: "Tháng 2", 3: "Tháng 3", 5: "Tháng 5", 6: "Tháng 6", 
                      8: "Tháng 8", 9: "Tháng 9", 11: "Tháng 11", 12: "Tháng 12"}
        month_name = month_names.get(current_month, f"Tháng {current_month}")
        
        return f"""
        <div class="monthly-indicator">
            <strong>🗓️ {month_name}:</strong> Báo cáo này bao gồm phân tích dịch chuyển OKR theo tháng (so với cuối tháng trước)
        </div>
        """

    def _generate_checkin_section(self, checkins_table: str, stats: Dict) -> str:
        """Generate checkin analysis section"""
        checkin_pct = (stats['members_with_checkins']/stats['total_members']*100) if stats['total_members'] > 0 else 0
        
        return f"""
        <div class="section">
            <h2>📝 DANH SÁCH NHÂN VIÊN CHƯA CHECKIN</h2>
            <div class="chart-container">
                {checkins_table}
            </div>
            <div class="alert alert-info">
                <strong>Thống kê:</strong> {stats['members_with_checkins']}/{stats['total_members']} nhân viên đã có Checkin ({checkin_pct:.1f}%)
            </div>
        </div>
        """

    def _generate_okr_sections(self, charts: Dict, tables: Dict, stats: Dict) -> str:
        """Generate OKR analysis sections"""
        sections = []
        
        # Weekly OKR section
        if 'okr_weekly' in charts:
            sections.append(f"""
            <div class="section">
                <h2>📊 DỊCH CHUYỂN OKR (TUẦN)</h2>
                <div class="chart-container">
                    {charts['okr_weekly']}
                </div>
                <div class="metrics">
                    <div class="metric">
                        <div class="metric-value positive">{stats['progress_users']}</div>
                        <div class="metric-label">Tiến bộ</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value neutral">{stats['stable_users']}</div>
                        <div class="metric-label">Ổn định</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value negative">{stats['issue_users']}</div>
                        <div class="metric-label">Cần quan tâm</div>
                    </div>
                </div>
            </div>
            """)
        
        # Monthly OKR section
        if 'okr_monthly' in charts:
            sections.append(f"""
            <div class="section">
                <h2>🗓️ DỊCH CHUYỂN OKR (THÁNG)</h2>
                <div class="chart-container">
                    {charts['okr_monthly']}
                </div>
                <div class="metrics">
                    <div class="metric">
                        <div class="metric-value positive">{stats.get('progress_users_monthly', 0)}</div>
                        <div class="metric-label">Tiến bộ</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value neutral">{stats.get('stable_users_monthly', 0)}</div>
                        <div class="metric-label">Ổn định</div>
                    </div>
                    <div class="metric">
                        <div class="metric-value negative">{stats.get('issue_users_monthly', 0)}</div>
                        <div class="metric-label">Cần quan tâm</div>
                    </div>
                </div>
            </div>
            """)
        
        return ''.join(sections)

    def _generate_detailed_sections(self, tables: Dict) -> str:
        """Generate detailed analysis sections"""
        sections = []
        
        # Top overall checkin section
        sections.append(f"""
        <div class="section">
            <h2>🏆 TẤT CẢ NHÂN VIÊN HOẠT ĐỘNG CHECKIN</h2>
            <div class="alert alert-info">
                <strong>Thống kê:</strong> Xếp hạng dựa trên tổng số checkin và tần suất checkin từ đầu quý
            </div>
            {tables['top_overall']}
        </div>
        """)
        
        # Conditional sections
        section_configs = [
            (tables['goals'], "🚫 NHÂN VIÊN CHƯA CÓ OKR", "Cần hành động: Những nhân viên này cần được hỗ trợ thiết lập OKR."),
            (tables['goals_no_checkins'], "⚠️ CÓ OKR NHƯNG CHƯA CHECKIN", "Ưu tiên cao: Đã có mục tiêu nhưng chưa cập nhật tiến độ."),
            (tables['top_performers'], "🏆 TẤT CẢ NHÂN VIÊN TIẾN BỘ (TUẦN)", None),
            (tables['top_performers_monthly'], "🗓️ TẤT CẢ NHÂN VIÊN TIẾN BỘ (THÁNG)", None),
            (tables['issue_performers'], "⚠️ NHÂN VIÊN CẦN HỖ TRỢ (TUẦN)", "Cần quan tâm: OKR của những nhân viên này đang giảm hoặc không tiến triển."),
            (tables['issue_performers_monthly'], "🗓️ NHÂN VIÊN CẦN HỖ TRỢ (THÁNG)", "Cần quan tâm: OKR tháng của những nhân viên này đang giảm hoặc không tiến triển.")
        ]
        
        for table_content, title, alert_msg in section_configs:
            if table_content and not ("📭 Không có dữ liệu" in table_content):
                alert_html = f'<div class="alert alert-warning"><strong>{alert_msg}</strong></div>' if alert_msg else ''
                sections.append(f"""
                <div class="section">
                    <h2>{title}</h2>
                    {alert_html}
                    {table_content}
                </div>
                """)
        
        return ''.join(sections)

    def _generate_email_footer(self) -> str:
        """Generate email footer"""
        return """
        <div class="footer">
            <p><strong>🏢 A Plus Mineral Material Corporation</strong></p>
            <p>📊 Báo cáo được tạo tự động bởi hệ thống OKR Analysis</p>
            <p><em>📧 Đây là email tự động, vui lòng không trả lời email này.</em></p>
        </div>
        """

    def send_email_report(self, email_from: str, password: str, email_to: str, 
                         subject: str, html_content: str, company_name: str = "A Plus Mineral Material Corporation") -> Tuple[bool, str]:
        """Send single email report"""
        try:
            message = MIMEMultipart('related')
            message['From'] = f"OKR System {company_name} <{email_from}>"
            message['To'] = email_to
            message['Subject'] = subject
            
            msg_alternative = MIMEMultipart('alternative')
            message.attach(msg_alternative)
            
            html_part = MIMEText(html_content, 'html', 'utf-8')
            msg_alternative.attach(html_part)
            
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(email_from, password)
                server.send_message(message)
            
            return True, "Email sent successfully!"
            
        except smtplib.SMTPAuthenticationError:
            return False, "Lỗi xác thực: Vui lòng kiểm tra lại email và mật khẩu"
        except Exception as e:
            return False, f"Lỗi gửi email: {str(e)}"

    def send_email_report_bulk(self, email_from: str, password: str, recipient_list: List[str], 
                              subject: str, html_content: str, excel_attachment: BytesIO = None) -> Tuple[bool, str, List[str]]:
        """Send email report to multiple recipients with optional Excel attachment"""
        success_count = 0
        failed_count = 0
        errors = []
        
        try:
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(email_from, password)
                
                for email_to in recipient_list:
                    try:
                        message = self._create_email_message(email_from, email_to, subject, html_content, excel_attachment)
                        server.send_message(message)
                        success_count += 1
                        
                    except Exception as e:
                        failed_count += 1
                        errors.append(f"{email_to}: {str(e)}")
            
            return True, f"Successfully sent {success_count} emails, {failed_count} failed", errors
            
        except Exception as e:
            return False, f"Server connection error: {str(e)}", errors

    def _create_email_message(self, email_from: str, email_to: str, subject: str, html_content: str, excel_attachment: BytesIO = None) -> MIMEMultipart:
        """Create email message with optional Excel attachment"""
        message = MIMEMultipart('related')
        message['From'] = f"OKR System A Plus <{email_from}>"
        message['To'] = email_to
        message['Subject'] = subject
        
        msg_alternative = MIMEMultipart('alternative')
        message.attach(msg_alternative)
        html_part = MIMEText(html_content, 'html', 'utf-8')
        msg_alternative.attach(html_part)
        
        # Add Excel attachment if provided
        if excel_attachment:
            excel_attachment.seek(0)  # Reset buffer position
            attachment = MIMEApplication(
                excel_attachment.read(),
                _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
            attachment.add_header(
                'Content-Disposition',
                'attachment',
                filename=f'okr_report_{current_time}.xlsx'
            )
            message.attach(attachment)
        
        return message








    """Main OKR Analysis System combining all components"""

    def __init__(self, goal_access_token: str, account_access_token: str):
        self.api_client = APIClient(goal_access_token, account_access_token)
        self.data_processor = DataProcessor()
        self.okr_calculator = OKRCalculator()
        self.checkin_path = None
        self.final_df = None
        self.filtered_members_df = None

    def get_filtered_members(self) -> pd.DataFrame:
        """Get filtered members from account API"""
        self.filtered_members_df = self.api_client.get_filtered_members()
        return self.filtered_members_df

    def get_cycle_list(self) -> List[Dict]:
        """Get list of quarterly cycles"""
        return self.api_client.get_cycle_list()
    
    def get_total_account_users(self) -> pd.DataFrame:
        """Get all account users (not filtered)"""
        return self.api_client.get_account_users()

    def load_and_process_data(self, progress_callback=None):
        """Main function to load and process all data"""
        try:
            steps = [
                ("Loading filtered members...", 0.05, self._load_filtered_members),
                ("Loading users...", 0.1, self._load_users),
                ("Loading goals...", 0.2, self._load_goals),
                ("Loading KRs...", 0.4, self._load_krs),
                ("Loading checkins...", 0.6, self._load_checkins),
                ("Processing data...", 0.8, self._process_data)
            ]
            
            data_store = {}
            
            for message, progress, step_func in steps:
                if progress_callback:
                    progress_callback(message, progress)
                
                result = step_func(data_store)
                if result is None:
                    return None
            
            if progress_callback:
                progress_callback("Data processing completed!", 1.0)

            return self.final_df

        except Exception as e:
            st.error(f"Error in data processing: {e}")
            return None

    def _load_filtered_members(self, data_store: Dict):
        """Load filtered members data"""
        filtered_members = self.get_filtered_members()
        if filtered_members.empty:
            st.error("Failed to load filtered members data")
            return None
        data_store['filtered_members'] = filtered_members
        return filtered_members

    def _load_users(self, data_store: Dict):
        """Load users data"""
        users_df = self.api_client.get_account_users()
        if users_df.empty:
            st.error("Failed to load users data")
            return None
        data_store['users_df'] = users_df
        data_store['user_id_to_name'] = dict(zip(users_df['id'], users_df['name']))
        return users_df

    def _load_goals(self, data_store: Dict):
        """Load goals data"""
        goals_df = self.api_client.get_goals_data(self.checkin_path)
        if goals_df.empty:
            st.error("Failed to load goals data")
            return None
        data_store['goals_df'] = goals_df
        return goals_df

    def _load_krs(self, data_store: Dict):
        """Load KRs data"""
        krs_df = self.api_client.get_krs_data(self.checkin_path)
        data_store['krs_df'] = krs_df
        return krs_df

    def _load_checkins(self, data_store: Dict):
        """Load checkins data"""
        all_checkins = self.api_client.get_all_checkins(self.checkin_path)
        checkin_df = self.data_processor.extract_checkin_data(all_checkins)
        data_store['checkin_df'] = checkin_df
        return checkin_df

    def _process_data(self, data_store: Dict):
        """Process and join all data - Updated for checkin.py compatibility"""
        goals_df = data_store['goals_df']
        krs_df = data_store['krs_df']
        checkin_df = data_store['checkin_df']
        user_id_to_name = data_store['user_id_to_name']
        
        # Join all data
        joined_df = goals_df.merge(krs_df, on='goal_id', how='left')
        joined_df['goal_user_name'] = joined_df['goal_user_id'].map(user_id_to_name)
        self.final_df = joined_df.merge(checkin_df, on='kr_id', how='left')
        
        # Add user_name mapping for checkin user_id - needed for checkin.py compatibility
        if 'checkin_user_id' in self.final_df.columns:
            self.final_df['user_name'] = self.final_df['checkin_user_id'].map(user_id_to_name)
        
        # Ensure we have checkin_since_timestamp column for checkin.py compatibility
        if 'checkin_since_timestamp' not in self.final_df.columns and 'checkin_since' in self.final_df.columns:
            # Try to extract timestamp from checkin_since if it's in datetime format
            def extract_timestamp(since_str):
                if pd.notna(since_str) and since_str != '':
                    try:
                        dt = pd.to_datetime(since_str)
                        return int(dt.timestamp())
                    except:
                        return None
                return None
            
            self.final_df['checkin_since_timestamp'] = self.final_df['checkin_since'].apply(extract_timestamp)
        
        # Clean data
        self.final_df = self.data_processor.clean_final_data(self.final_df)
        
        # Add debug info for data structure compatibility
        if not self.final_df.empty:
            required_columns = ['goal_user_name', 'checkin_since', 'checkin_since_timestamp', 'checkin_kr_current_value', 'kr_id']
            missing_columns = [col for col in required_columns if col not in self.final_df.columns]
            if missing_columns:
                print(f"⚠️ Missing columns for checkin.py compatibility: {missing_columns}")
            else:
                print(f"✅ All required columns present for checkin.py compatibility")
        
        return self.final_df

    def analyze_missing_goals_and_checkins(self) -> Tuple[List[Dict], List[Dict], List[Dict]]:
        """Analyze members without goals and without checkins"""
        try:
            if self.filtered_members_df is None or self.final_df is None:
                return [], [], []

            users_with_goals = set(self.final_df['goal_user_name'].dropna().unique())
            users_with_checkins = self._get_users_with_checkins()
            all_members = set(self.filtered_members_df['name'].unique())
            
            missing_lists = {'goals': [], 'checkins': [], 'goals_no_checkins': []}
            
            for member_name in all_members:
                member_info = self.filtered_members_df[self.filtered_members_df['name'] == member_name].iloc[0].to_dict()
                member_data = self._create_member_dict(member_info)
                
                has_goal = member_name in users_with_goals
                has_checkin = member_name in users_with_checkins
                
                if not has_goal:
                    missing_lists['goals'].append(member_data)
                
                if not has_checkin:
                    member_data_with_goal = member_data.copy()
                    member_data_with_goal['has_goal'] = has_goal
                    missing_lists['checkins'].append(member_data_with_goal)
                
                if has_goal and not has_checkin:
                    missing_lists['goals_no_checkins'].append(member_data)
            
            return missing_lists['goals'], missing_lists['checkins'], missing_lists['goals_no_checkins']
            
        except Exception as e:
            st.error(f"Error analyzing missing goals and checkins: {e}")
            return [], [], []


    def get_user_excel_data(self, user_name: str, weekly_shifts: List[Dict], period_checkins: List[Dict], alignment_data: Dict) -> Dict:
        """
        Prepare data for Excel generation for a single user.
        Based on user's specific scoring rules.
        """
        stats = {}
        
        # 1. Weekly Shift Data
        # Formula: (Dịch chuyển tháng / 33.33) * 100
        # Classification: <25%, 25-50%, 50-75%, 75-100%, >100%
        
        # We need monthly shift data. 'weekly_shifts' passed here might be weekly.
        # Ideally we should pass monthly_shifts if available, or use the weekly sum?
        # User said "Dịch chuyển so với tháng trước".
        # Let's assume we use the 'okr_shift_monthly' if present in the weekly_shifts list (which might be enriched) 
        # OR we need the actual monthly data.
        # For now, let's use the 'okr_shift' from the passed list, assuming it's the relevant one (or we need to update the caller to pass monthly data).
        
        # Strategy: The caller (run_analysis -> show_export_options) passes 'okr_shifts'. 
        # If we want monthly, we should check if caller passed monthly data or we find it.
        # Let's try to look for 'okr_shift_monthly' key first, else fallback to 'okr_shift'.
        
        user_shift_data = next((u for u in weekly_shifts if u['user_name'] == user_name), None)
        
        shift_percentage = 0
        if user_shift_data:
            # Prefer monthly shift if available
            raw_shift = user_shift_data.get('okr_shift_monthly', user_shift_data.get('okr_shift', 0))
            
            # Calculate % based on target 33.33%
            # If raw_shift is already %, e.g. 0.1 (10%), then (10/33.33)*100 = 30%
            # If raw_shift is absolute value, we assume it's comparable to 0-1 scale? 
            # Usually OKR shift is 0.0 to 1.0 (0-100%).
            # Let's assume raw_shift is 0.XYZ (e.g. 0.15 = 15%).
            # Formula: (0.15 * 100) / 33.33 * 100 ?? No.
            # User says: "Dịch chuyển so với tháng trước/33,3%"
            # Example: Shift = 10%. (10 / 33.33) * 100 = 30%. -> 25-50% bucket.
            
            # Assuming raw_shift is float 0.0-1.0.
            shift_val = raw_shift * 100 # Convert to % first (e.g. 15.0)
            
            if shift_val != 0:
                 shift_percentage = (shift_val / 33.33) * 100
            else:
                 shift_percentage = 0

            # Classification
            if shift_percentage < 25:
                stats['shift_lt_25'] = True # +5
            elif 25 <= shift_percentage < 50:
                stats['shift_25_50'] = True # +10
            elif 50 <= shift_percentage < 75:
                stats['shift_50_75'] = True # +15
            elif 75 <= shift_percentage < 100:
                stats['shift_75_100'] = True # +18
            else: # >= 100
                stats['shift_gt_100'] = True # +20
        else:
             stats['shift_lt_25'] = True # Default 0 -> <25
                
        # 2. Check-in & Discipline
        # 2.1 Has OKRs
        user_goals = self.final_df[
             (self.final_df['goal_user_name'] == user_name) & 
             (self.final_df['goal_name'].notna())
        ]
        has_okrs = not user_goals.empty
        stats['has_okrs'] = "Yes" if has_okrs else "No"
        
        # Check-in Score: Count * 2, Max 8
        user_checkin_data = next((u for u in period_checkins if u['user_name'] == user_name), None)
        checkin_score = 0
        if user_checkin_data:
            # Use 'checkin_count_period' or calculate from raw checkins if period is too long?
            # Assuming 'period_checkins' is for the relevant month/period.
            count = user_checkin_data.get('checkin_count_period', 0)
            checkin_score = min(count * 2, 8) 
            
        stats['checkin_score'] = True 
        stats['checkin_score_val'] = checkin_score

        # Collaboration: Default 2
        stats['collab_score'] = True
        
        # 2.2 Quality: Median of next_action_scores (1, 3, 5)
        # We need to calculate/mock next_action_score logic
        
        # Get all checkins for user to calculate median
        user_checkins_rows = self.final_df[
            (self.final_df['goal_user_name'] == user_name) & 
            (self.final_df['checkin_content'].notna())
        ]
        
        scores = []
        if not user_checkins_rows.empty:
            # Heuristic for next_action_score (1, 3, 5)
            # 1: No clear action
            # 3: Status report
            # 5: Clear action
            # Proxy: Content length
            for content in user_checkins_rows['checkin_content'].astype(str):
                length = len(content)
                if length < 20: s = 1
                elif length < 80: s = 3
                else: s = 5
                scores.append(s)
        
        median_score = np.median(scores) if scores else 1
        
        # Fill bucket
        if median_score >= 4.5:
            stats['quality_high'] = True
        elif median_score >= 2.5:
            stats['quality_med'] = True
        else:
            stats['quality_low'] = True

        # 3. Section II: Median for Alignment, Priority, Impact
        # "tương tự thế phần II cũng tính trung vị cho cả 3 cái"
        
        # Get all goals for this user from final_df (unique goals)
        # final_df is joined with KRs and Checkins, so goals are duplicated.
        # We need unique goals.
        user_goals_df = self.final_df[self.final_df['goal_user_name'] == user_name].drop_duplicates('goal_id')
        
        def calculate_hybrid_score_from_str_col(col_name):
            values = []
            if col_name in user_goals_df.columns:
                for val_str in user_goals_df[col_name]:
                    if pd.notna(val_str) and isinstance(val_str, str) and len(val_str) > 0:
                        try:
                            # Take first char (e.g. "1 - ...")
                            first_char = val_str.strip()[0]
                            val = int(first_char)
                            values.append(val)
                        except:
                            pass
            
            if not values:
                return 1.0 # Default to 1 if empty/error
                
            # 1. Calculate Median
            med = np.median(values)
            
            # 2. Check if Median is integer (e.g. 3.0) or decimal (3.5)
            if med % 1 == 0:
                return float(med)
            
            # 3. If decimal, use Mode (Most frequent, max if tie)
            counts = {}
            for v in values:
                counts[v] = counts.get(v, 0) + 1
            max_count = max(counts.values())
            modes = [k for k, v in counts.items() if v == max_count]
            return float(max(modes))

        align_score = calculate_hybrid_score_from_str_col('alignment_score_str')
        prio_score = calculate_hybrid_score_from_str_col('priority_score_str')
        impact_score = calculate_hybrid_score_from_str_col('impact_score_str')
        
        # Deployment for Alignment Bucket
        # Map 1-5 to specific keys
        if align_score >= 4.5: stats['align_direct_2'] = True      # 5
        elif align_score >= 3.5: stats['align_direct_1'] = True    # 4
        elif align_score >= 2.5: stats['align_indirect_2'] = True  # 3
        elif align_score >= 1.5: stats['align_indirect_1'] = True  # 2
        else: stats['align_personal'] = True                       # 1
        
        # Priority Bucket
        if prio_score >= 4.5: stats['prio_very_important_2'] = True
        elif prio_score >= 3.5: stats['prio_very_important_1'] = True
        elif prio_score >= 2.5: stats['prio_important_2'] = True
        elif prio_score >= 1.5: stats['prio_important_1'] = True
        else: stats['prio_normal'] = True
        
        # Impact Bucket
        # Excel template seems to map 1-5: 
        # 1: Personal, 2: Team, 3: Team(High), 4: Company, 5: Company(High)
        if impact_score >= 4.5: stats['impact_company_2'] = True
        elif impact_score >= 3.5: stats['impact_company_1'] = True
        elif impact_score >= 2.5: stats['impact_team_2'] = True
        elif impact_score >= 1.5: stats['impact_team_1'] = True
        else: stats['impact_personal'] = True
        
        return {
            'name': user_name,
            'stats': stats
        }

    def _get_users_with_checkins(self) -> set:

        """Get set of users who have made checkins"""
        users_with_checkins = set()
        if 'checkin_user_id' in self.final_df.columns:
            user_id_to_name = dict(zip(self.filtered_members_df['id'], self.filtered_members_df['name']))
            checkin_user_ids = self.final_df['checkin_user_id'].dropna().unique()
            users_with_checkins = {
                user_id_to_name.get(uid, uid) 
                for uid in checkin_user_ids 
                if uid in user_id_to_name
            }
        return users_with_checkins

    def _create_member_dict(self, member_info: Dict) -> Dict:
        """Create standardized member dictionary"""
        return {
            'name': member_info.get('name', ''),
            'username': member_info.get('username', ''),
            'job': member_info.get('job', ''),
            'email': member_info.get('email', ''),
            'id': member_info.get('id', '')
        }

    def calculate_okr_shifts_by_user(self) -> List[Dict]:
        """Calculate OKR shifts for each user (weekly)"""
        return self._calculate_okr_shifts_by_period("weekly")

    def calculate_okr_shifts_by_user_monthly(self) -> List[Dict]:
        """Calculate monthly OKR shifts for each user"""
        if not DateUtils.should_calculate_monthly_shift():
            return []
        return self._calculate_okr_shifts_by_period("monthly")

    def _calculate_okr_shifts_by_period(self, period: str) -> List[Dict]:
        """Calculate OKR shifts for specified period (weekly or monthly)"""
        try:
            users = self.final_df['goal_user_name'].dropna().unique()
            user_okr_shifts = []
            
            reference_date = DateUtils.get_last_friday_date() if period == "weekly" else DateUtils.get_last_month_end_date()
            
            for user in users:
                user_df = self.final_df[self.final_df['goal_user_name'] == user].copy()
                shift_data = self._calculate_user_shift_data(user_df, reference_date, period)
                user_okr_shifts.append(shift_data)
            
            shift_key = 'okr_shift' if period == "weekly" else 'okr_shift_monthly'
            return sorted(user_okr_shifts, key=lambda x: x[shift_key], reverse=True)
            
        except Exception as e:
            st.error(f"Error calculating {period} OKR shifts: {e}")
            return []

    def _calculate_user_shift_data(self, user_df: pd.DataFrame, reference_date: datetime, period: str) -> Dict:
        """Calculate shift data for a single user"""
        user_name = user_df['goal_user_name'].iloc[0] if not user_df.empty else 'Unknown'
        
        if period == "weekly":
            return self._calculate_weekly_shift_data(user_df, user_name, reference_date)
        else:
            return self._calculate_monthly_shift_data(user_df, user_name, reference_date)

    def _calculate_weekly_shift_data(self, user_df: pd.DataFrame, user_name: str, reference_friday: datetime) -> Dict:
        """Calculate weekly shift data for user"""
        final_okr_goal_shift = self._calculate_final_okr_goal_shift(user_df, reference_friday, "weekly")
        current_value = self.okr_calculator.calculate_current_value(user_df)
        reference_value, kr_details = self.okr_calculator.calculate_reference_value(reference_friday, user_df)
        
        # Áp dụng logic mới theo yêu cầu:
        # 1. Nếu giá trị thứ 6 tuần trước > giá trị hiện tại thì giá trị thứ 6 = giá trị hiện tại - dịch chuyển tuần
        # 2. Nếu giá trị thứ 6 tuần trước < giá trị hiện tại và (giá trị hiện tại - giá trị thứ 6 tuần trước) != dịch chuyển
        #    thì dịch chuyển tuần = giá trị hiện tại - giá trị thứ 6 tuần trước
        
        adjusted_reference_value = reference_value
        adjusted_okr_shift = final_okr_goal_shift
        reference_adjustment_applied = False
        shift_adjustment_applied = False
        
        # Quy tắc 1: Nếu reference_value > current_value
        if reference_value > current_value:
            adjusted_reference_value = current_value - final_okr_goal_shift
            reference_adjustment_applied = True
        
        # Quy tắc 2: Nếu reference_value < current_value VÀ (current_value - reference_value) != shift
        elif reference_value < current_value and (current_value - reference_value) != final_okr_goal_shift:
            adjusted_okr_shift = current_value - reference_value
            shift_adjustment_applied = True
        
        legacy_okr_shift = current_value - reference_value

        return {
            'user_name': user_name,
            'okr_shift': adjusted_okr_shift,
            'original_shift': final_okr_goal_shift,
            'current_value': current_value,
            'last_friday_value': adjusted_reference_value,
            'original_last_friday_value': reference_value,
            'legacy_okr_shift': legacy_okr_shift,
            'adjustment_applied': shift_adjustment_applied,
            'reference_adjustment_applied': reference_adjustment_applied,
            'kr_details_count': len(kr_details),
            'reference_friday': reference_friday.strftime('%d/%m/%Y')
        }

    def _calculate_monthly_shift_data(self, user_df: pd.DataFrame, user_name: str, reference_month_end: datetime) -> Dict:
        """Calculate monthly shift data for user"""
        final_okr_goal_shift_monthly = self._calculate_final_okr_goal_shift(user_df, reference_month_end, "monthly")
        current_value = self.okr_calculator.calculate_current_value(user_df)
        reference_value, kr_details = self.okr_calculator.calculate_reference_value(reference_month_end, user_df)
        
        # Kiểm tra xem có phải tuần 4 hoặc 5 của tháng đầu quý không
        # Nếu đúng thì tính chuyển động tháng = điểm số hiện tại so với 0
        if DateUtils.is_week_4_or_5_of_quarter_start_month():
            adjusted_okr_shift = current_value  # current_value - 0
            adjusted_reference_value = 0
            reference_adjustment_applied = True
            shift_adjustment_applied = True
            legacy_okr_shift = current_value
        else:
            # Áp dụng logic mới theo yêu cầu:
            # 1. Nếu giá trị cuối tháng trước > giá trị hiện tại thì giá trị cuối tháng = giá trị hiện tại - dịch chuyển tháng
            # 2. Nếu giá trị cuối tháng trước < giá trị hiện tại và (giá trị hiện tại - giá trị cuối tháng trước) != dịch chuyển
            #    thì dịch chuyển tháng = giá trị hiện tại - giá trị cuối tháng trước
            
            adjusted_reference_value = reference_value
            adjusted_okr_shift = final_okr_goal_shift_monthly
            reference_adjustment_applied = False
            shift_adjustment_applied = False
            
            # Quy tắc 1: Nếu reference_value > current_value
            if reference_value > current_value:
                adjusted_reference_value = current_value - final_okr_goal_shift_monthly
                reference_adjustment_applied = True
            
            # Quy tắc 2: Nếu reference_value < current_value VÀ (current_value - reference_value) != shift
            elif reference_value < current_value and (current_value - reference_value) != final_okr_goal_shift_monthly:
                adjusted_okr_shift = current_value - reference_value
                shift_adjustment_applied = True
            
            legacy_okr_shift = current_value - reference_value

        return {
            'user_name': user_name,
            'okr_shift_monthly': adjusted_okr_shift,
            'original_shift_monthly': final_okr_goal_shift_monthly,
            'current_value': current_value,
            'last_month_value': adjusted_reference_value,
            'original_last_month_value': reference_value,
            'legacy_okr_shift_monthly': legacy_okr_shift,
            'adjustment_applied': shift_adjustment_applied,
            'reference_adjustment_applied': reference_adjustment_applied,
            'kr_details_count': len(kr_details),
            'reference_month_end': reference_month_end.strftime('%d/%m/%Y')
        }

    def _calculate_final_okr_goal_shift(self, user_df: pd.DataFrame, reference_date: datetime, period: str) -> float:
        """Calculate final OKR goal shift for specified period"""
        try:
            unique_combinations = {}
            
            for _, row in user_df.iterrows():
                goal_name = row.get('goal_name', '')
                kr_name = row.get('kr_name', '')
                
                if not goal_name or not kr_name:
                    continue
                
                combo_key = f"{goal_name}|{kr_name}"
                kr_shift = self.okr_calculator.calculate_kr_shift(row, reference_date, self.final_df)
                
                if combo_key not in unique_combinations:
                    unique_combinations[combo_key] = []
                unique_combinations[combo_key].append(kr_shift)
            
            final_shifts = [
                sum(kr_shifts) / len(kr_shifts) 
                for kr_shifts in unique_combinations.values() 
                if kr_shifts
            ]
            
            return sum(final_shifts) / len(final_shifts) if final_shifts else 0
            
        except Exception as e:
            st.error(f"Error calculating final_okr_goal_shift: {e}")
            return 0

    def analyze_checkin_behavior(self) -> Tuple[List[Dict], List[Dict]]:
        """Analyze checkin behavior for both period and overall"""
        try:
            last_friday = DateUtils.get_last_friday_date()
            quarter_start = DateUtils.get_quarter_start_date()

            df = self.final_df.copy()
            df['checkin_since_dt'] = pd.to_datetime(df['checkin_since'], errors='coerce')

            # Period analysis (quarter start to last Friday)
            mask_period = (df['checkin_since_dt'] >= quarter_start) & (df['checkin_since_dt'] <= last_friday)
            period_df = df[mask_period].copy()

            # Overall analysis (all time)
            all_time_df = df[df['checkin_id'].notna()].copy()

            all_users = df['goal_user_name'].dropna().unique()

            period_checkins = self._analyze_period_checkins(period_df, all_users)
            overall_checkins = self._analyze_overall_checkins(all_time_df, all_users)

            return period_checkins, overall_checkins

        except Exception as e:
            st.error(f"Error analyzing checkin behavior: {e}")
            return [], []

    def _analyze_period_checkins(self, period_df: pd.DataFrame, all_users: List[str]) -> List[Dict]:
        """Analyze checkins in the reference period"""
        period_checkins = []

        for user in all_users:
            try:
                user_period_data = period_df[period_df['goal_user_name'] == user]
                
                user_period_checkins = user_period_data[
                    (user_period_data['checkin_name'].notna()) &
                    (user_period_data['checkin_name'] != '')
                ]['checkin_id'].nunique()

                user_krs_in_period = user_period_data['kr_id'].nunique()
                checkin_rate = (user_period_checkins / user_krs_in_period * 100) if user_krs_in_period > 0 else 0

                user_checkin_dates = user_period_data[
                    (user_period_data['checkin_name'].notna()) &
                    (user_period_data['checkin_name'] != '')
                ]['checkin_since_dt'].dropna()

                first_checkin = user_checkin_dates.min() if len(user_checkin_dates) > 0 else None
                last_checkin = user_checkin_dates.max() if len(user_checkin_dates) > 0 else None
                days_between = (last_checkin - first_checkin).days if first_checkin and last_checkin else 0

                period_checkins.append({
                    'user_name': user,
                    'checkin_count_period': user_period_checkins,
                    'kr_count_period': user_krs_in_period,
                    'checkin_rate_period': checkin_rate,
                    'first_checkin_period': first_checkin,
                    'last_checkin_period': last_checkin,
                    'days_between_checkins': days_between
                })
            except Exception as e:
                st.warning(f"Error analyzing period checkins for {user}: {e}")
                continue

        return sorted(period_checkins, key=lambda x: x['checkin_count_period'], reverse=True)

    def _analyze_overall_checkins(self, all_time_df: pd.DataFrame, all_users: List[str]) -> List[Dict]:
        """Analyze overall checkin behavior"""
        overall_checkins = []
        
        today = datetime.now()
        quarter_start = DateUtils.get_quarter_start_date()
        weeks_in_quarter = max((today - quarter_start).days / 7, 1)
        
        # Calculate last week boundaries
        days_since_monday = today.weekday()
        monday_this_week = today - timedelta(days=days_since_monday)
        monday_last_week = monday_this_week - timedelta(days=7)
        sunday_last_week = monday_last_week + timedelta(days=6, hours=23, minutes=59, seconds=59)
    
        for user in all_users:
            try:
                user_data = all_time_df[all_time_df['goal_user_name'] == user]
                
                user_total_checkins = user_data['checkin_id'].nunique()
                user_total_krs = self.final_df[self.final_df['goal_user_name'] == user]['kr_id'].nunique()
                checkin_rate = (user_total_checkins / user_total_krs * 100) if user_total_krs > 0 else 0
    
                user_checkins_dates = user_data['checkin_since_dt'].dropna()
                first_checkin = user_checkins_dates.min() if len(user_checkins_dates) > 0 else None
                last_checkin = user_checkins_dates.max() if len(user_checkins_dates) > 0 else None
                days_active = (last_checkin - first_checkin).days if first_checkin and last_checkin else 0
    
                checkin_frequency = user_total_checkins / weeks_in_quarter
                
                # Count last week checkins
                last_week_checkins = user_checkins_dates[
                    (user_checkins_dates >= monday_last_week) & 
                    (user_checkins_dates <= sunday_last_week)
                ] if len(user_checkins_dates) > 0 else []
                
                user_last_week_checkins = len(last_week_checkins)
    
                overall_checkins.append({
                    'user_name': user,
                    'total_checkins': user_total_checkins,
                    'total_krs': user_total_krs,
                    'checkin_rate': checkin_rate,
                    'first_checkin': first_checkin,
                    'last_checkin': last_checkin,
                    'days_active': days_active,
                    'checkin_frequency_per_week': checkin_frequency,
                    'last_week_checkins': user_last_week_checkins,
                    'weeks_in_quarter': weeks_in_quarter
                })
            except Exception as e:
                st.warning(f"Error analyzing overall checkins for {user}: {e}")
                continue
    
        return sorted(overall_checkins, key=lambda x: x['total_checkins'], reverse=True)

# ==================== UTILITY FUNCTIONS ====================

def create_user_manager_with_monthly_calculation(analyzer):
    """Create UserManager using EXACT data from Monthly OKR Analysis - Updated for checkin.py compatibility"""
    
    # Sử dụng CHÍNH XÁC dữ liệu từ Monthly OKR Analysis
    if 'monthly_okr_data' in st.session_state and st.session_state['monthly_okr_data']:
        monthly_okr_data = st.session_state['monthly_okr_data']
        monthly_user_names = [data['user_name'] for data in monthly_okr_data]
        
        # Create account_df - MUST include ALL users from filtered_members_df for proper checkin calculation
        if analyzer.filtered_members_df is not None and not analyzer.filtered_members_df.empty:
            # Use ALL filtered members, not just monthly OKR users - this is needed for checkin.py logic
            account_df = analyzer.filtered_members_df.copy()
            
            # Add missing users from monthly_okr_data if any
            existing_names = set(account_df['name'].tolist())
            missing_names = set(monthly_user_names) - existing_names
            
            if missing_names:
                missing_records = []
                for name in missing_names:
                    missing_records.append({
                        'name': name,
                        'username': name.lower().replace(' ', ''),
                        'email': f"{name.lower().replace(' ', '')}@company.com",
                        'job': 'N/A',
                        'id': f"okr_{hash(name) % 10000}"
                    })
                missing_df = pd.DataFrame(missing_records)
                account_df = pd.concat([account_df, missing_df], ignore_index=True)
        else:
            # Tạo account_df từ monthly_okr_data nếu không có filtered_members_df
            account_records = []
            for name in monthly_user_names:
                account_records.append({
                    'name': name,
                    'username': name.lower().replace(' ', ''),
                    'email': f"{name.lower().replace(' ', '')}@company.com",
                    'job': 'N/A',
                    'id': f"okr_{hash(name) % 10000}"
                })
            account_df = pd.DataFrame(account_records)
        
        # Extract data using checkin.py compatible format
        krs_df = _extract_krs_data_for_user_manager_checkin_style(analyzer)
        checkin_df = _extract_checkin_data_for_user_manager_checkin_style(analyzer)
        
        # UserManager needs final_df to work like checkins_df from checkin.py
        monthly_users_set = set(monthly_user_names)
        return UserManager(account_df, krs_df, checkin_df, analyzer.final_df, analyzer.final_df, monthly_users_set, monthly_okr_data)
    
    else:
        st.warning("⚠️ No Monthly OKR Analysis data found. Please run Monthly OKR Analysis first!")
        return UserManager(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), None, None, set(), [])

def _extract_krs_data_for_user_manager_checkin_style(analyzer) -> pd.DataFrame:
    """Extract KRs data for UserManager - checkin.py compatible format"""
    # UserManager doesn't actually use this data for checkin calculation
    # It uses final_df directly, so we can return minimal structure
    return pd.DataFrame(columns=['user_id', 'kr_id', 'current_value'])

def _extract_checkin_data_for_user_manager_checkin_style(analyzer) -> pd.DataFrame:
    """Extract checkin data for UserManager - checkin.py compatible format"""
    # UserManager doesn't actually use this data for checkin calculation with new logic
    # It uses final_df directly like checkin.py, so we can return minimal structure
    return pd.DataFrame(columns=['user_id', 'day', 'checkin_id'])

def _extract_krs_data_for_user_manager(analyzer) -> pd.DataFrame:
    """Extract KRs data for UserManager from final_df - DEPRECATED"""
    krs_data = []
    for _, row in analyzer.final_df.iterrows():
        if pd.notna(row.get('kr_id')):
            user_name = row.get('goal_user_name', '')
            user_id = _get_user_id_from_name(analyzer, user_name)
            
            if user_id:
                krs_data.append({
                    'user_id': user_id,
                    'kr_id': row.get('kr_id'),
                    'current_value': row.get('kr_current_value', 0)
                })
    
    return pd.DataFrame(krs_data)

def _extract_checkin_data_for_user_manager(analyzer) -> pd.DataFrame:
    """Extract checkin data for UserManager from final_df - DEPRECATED"""
    checkin_data = []
    for _, row in analyzer.final_df.iterrows():
        if pd.notna(row.get('checkin_id')):
            user_name = row.get('goal_user_name', '')
            user_id = _get_user_id_from_name(analyzer, user_name)
            
            if user_id and pd.notna(row.get('checkin_since')):
                try:
                    checkin_datetime = pd.to_datetime(row['checkin_since'])
                    timestamp = checkin_datetime.timestamp()
                    
                    checkin_data.append({
                        'user_id': user_id,
                        'day': timestamp,
                        'checkin_id': row.get('checkin_id')
                    })
                except:
                    continue
    
    return pd.DataFrame(checkin_data)

def _get_user_id_from_name(analyzer, user_name: str) -> Optional[str]:
    """Get user_id from user_name using filtered_members_df"""
    if not user_name:
        return None
    
    for uid, name in analyzer.filtered_members_df.set_index('id')['name'].items():
        if name == user_name:
            return uid
    return None

def export_to_excel(users: List[User], filename: str = "output1.xlsx") -> openpyxl.Workbook:
    """Export user OKR data to Excel with improved formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OKRs"

    # Define styles
    styles = {
        'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'header_font': Font(bold=True, color="FFFFFF"),
        'category_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        'category_font': Font(bold=True),
        'thin_border': Border(left=Side(style="thin"), right=Side(style="thin"), 
                             top=Side(style="thin"), bottom=Side(style="thin"))
    }

    # Create title and headers
    _create_excel_title_and_headers(ws, users, styles)
    
    # Create criteria rows
    _create_excel_criteria_rows(ws, styles)
    
    # Fill user data
    _fill_excel_user_data(ws, users, styles)
    
    # Apply formatting
    _apply_excel_formatting(ws, users)
    
    return wb

def _create_excel_title_and_headers(ws, users: List[User], styles: Dict):
    """Create title and headers for Excel file"""
    total_columns = 3 + len(users)
    last_col_letter = get_column_letter(total_columns)
    
    # Title
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "ĐÁNH GIÁ OKRs THÁNG"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    fixed_headers = ["TT", "Nội dung", "Tự chấm điểm"]
    user_headers = [user.name for user in users]
    headers = fixed_headers + user_headers
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles['thin_border']

def _create_excel_criteria_rows(ws, styles: Dict):
    """Create criteria rows in Excel"""
    criteria = [
        [1, "Đầy đủ OKRs cá nhân được cập nhật trên Base Goal (Mục tiêu cá nhân + Đường dẫn)", 1],
        [2, "Có Check-in trên base hàng tuần (Mỗi tuần ít nhất 1 lần check-in)", 0.5],
        [3, "Có check-in với người khác, cấp quản lý, làm việc chung OKRs trong bộ phận", 0.5],
        [4, "Tổng OKRs dịch chuyển trong tháng (so với tháng trước):", ""],
        ["", "Nhỏ hơn 10%", 0.15],
        ["", "Từ 10 - 25%", 0.25],
        ["", "Từ 26 - 30%", 0.5],
        ["", "Từ 31 - 50%", 0.75],
        ["", "Từ 51 - 80%", 1.25],
        ["", "Từ 81% - 99%", 1.5],
        ["", "100% hoặc có đột phá lớn", 2.5],
        [5, "Tổng cộng OKRs", ""]
    ]
    
    start_row = 3
    for i, row_data in enumerate(criteria):
        row_idx = start_row + i
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = styles['thin_border']
            
            if col_idx == 1 and isinstance(value, int):
                cell.fill = styles['category_fill']
                cell.font = styles['category_font']

def _fill_excel_user_data(ws, users: List[User], styles: Dict):
    """Fill user data in Excel"""
    for idx, user in enumerate(users, start=1):
        col_idx = 3 + idx
        
        # Basic scores
        ws.cell(row=3, column=col_idx, value=1 if user.co_OKR == 1 else 0)
        ws.cell(row=4, column=col_idx, value=0.5 if user.checkin == 1 else 0)
        ws.cell(row=5, column=col_idx, value=0.5)

        # Movement percentage and score
        movement = user.dich_chuyen_OKR
        ws.cell(row=6, column=col_idx, value=f"{movement}%")

        # Determine movement score and row
        movement_score, movement_row = _get_movement_score_and_row(movement)
        ws.cell(row=movement_row, column=col_idx, value=movement_score)

        # Total score
        ws.cell(row=14, column=col_idx, value=user.score)

        # Apply formatting
        for r in range(3, 15):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = styles['thin_border']

def _get_movement_score_and_row(movement: float) -> Tuple[float, int]:
    """Get movement score and corresponding row number"""
    movement_ranges = [
        (10, 0.15, 7), (26, 0.25, 8), (31, 0.5, 9), (51, 0.75, 10),
        (81, 1.25, 11), (100, 1.5, 12), (float('inf'), 2.5, 13)
    ]
    
    for threshold, score, row in movement_ranges:
        if movement < threshold:
            return score, row
    
    return 2.5, 13  # Default case

def _apply_excel_formatting(ws, users: List[User]):
    """Apply final formatting to Excel worksheet"""
    # Set column widths
    column_widths = {1: 5, 2: 70}  # TT and Nội dung columns
    for col_idx in range(1, 4 + len(users)):
        col_letter = get_column_letter(col_idx)
        width = column_widths.get(col_idx, 15)
        ws.column_dimensions[col_letter].width = width

    # Freeze panes
    ws.freeze_panes = ws["D3"]

def get_email_list(analyzer) -> List[str]:
    """Get email list from filtered members"""
    if analyzer.filtered_members_df is not None:
        email_list = analyzer.filtered_members_df['email'].dropna().tolist()
        return [email for email in email_list if email.strip()]
    return []

def get_default_recipients() -> List[str]:
    """Get default special recipients"""
    return ["tts122403@gmail.com"]


# ==================== STREAMLIT UI FUNCTIONS ====================

def show_realtime_checkin_preview(analyzer):
    """Show real-time check-in preview without waiting for last week"""
    st.subheader("📈 Real-time Check-in Preview")
    
    try:
        # Tạo user manager để lấy dữ liệu preview
        user_manager = create_user_manager_with_monthly_calculation(analyzer)
        if not user_manager:
            st.warning("⚠️ Không thể tạo user manager. Vui lòng chạy Monthly OKR Analysis trước.")
            return
            
        user_manager.update_checkins()
        
        # Lấy preview data
        preview_df = user_manager.get_realtime_checkin_preview()
        alerts = user_manager.generate_checkin_alerts()
        
        if preview_df.empty:
            st.info("📊 Không có dữ liệu check-in để hiển thị")
            return
        
        # Hiển thị metrics tổng quan
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_users = len(preview_df)
            st.metric("Tổng số người", total_users)
        
        with col2:
            achieved_users = len(preview_df[preview_df['Điểm dự kiến'] >= 0.5])
            st.metric("Đạt tiêu chí", f"{achieved_users}/{total_users}")
        
        with col3:
            avg_weeks = preview_df['Tuần có check-in'].str.split('/').str[0].astype(int).mean()
            st.metric("TB tuần check-in", f"{avg_weeks:.1f}")
        
        with col4:
            urgent_alerts = len([a for a in alerts if a['urgency_level'] >= 3])
            st.metric("Cảnh báo khẩn", urgent_alerts)
        
        # Hiển thị bảng preview
        st.markdown("### 📊 Bảng điểm dự kiến")
        
        # Tạo styled dataframe
        display_df = preview_df.drop(columns=['_status_color'])
        
        # Style dataframe với màu sắc dựa trên trạng thái
        def style_preview_table(val, col_name):
            if col_name == 'Trạng thái':
                if '✅' in str(val):
                    return 'background-color: #d4edda; color: #155724;'
                elif '⚠️' in str(val):
                    return 'background-color: #fff3cd; color: #856404;'
                elif '❌' in str(val):
                    return 'background-color: #f8d7da; color: #721c24;'
            elif col_name == 'Điểm dự kiến':
                if float(val) >= 0.5:
                    return 'background-color: #d4edda; font-weight: bold;'
                else:
                    return 'background-color: #f8d7da;'
            return ''
        
        # Apply styling
        styled_df = display_df.style.applymap(lambda x: style_preview_table(x, display_df.columns[display_df.iloc[0].tolist().index(x) if x in display_df.iloc[0].tolist() else -1]) if hasattr(display_df, 'columns') else '', subset=None)
        
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        # Hiển thị alerts nếu có
        if alerts:
            st.markdown("### 🚨 Cảnh báo Check-in")
            
            # Group alerts by urgency
            urgent_alerts = [a for a in alerts if a['urgency_level'] >= 3]
            warning_alerts = [a for a in alerts if a['urgency_level'] == 2]
            info_alerts = [a for a in alerts if a['urgency_level'] == 1]
            
            if urgent_alerts:
                with st.expander(f"🔴 KHẨN CẤP ({len(urgent_alerts)} người)", expanded=True):
                    for alert in urgent_alerts:
                        st.error(f"**{alert['user_name']}**: {alert['message']} (còn {alert['days_left']} ngày)")
            
            if warning_alerts:
                with st.expander(f"🟡 CẦN CHÚ Ý ({len(warning_alerts)} người)"):
                    for alert in warning_alerts:
                        st.warning(f"**{alert['user_name']}**: {alert['message']} (còn {alert['days_left']} ngày)")
            
            if info_alerts:
                with st.expander(f"🟢 BÌNH THƯỜNG ({len(info_alerts)} người)"):
                    for alert in info_alerts:
                        st.info(f"**{alert['user_name']}**: {alert['message']} (còn {alert['days_left']} ngày)")
        
        # Export options
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📊 Export Preview Data"):
                csv = preview_df.drop(columns=['_status_color']).to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name=f"checkin_preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
        
        with col2:
            if st.button("🔄 Refresh Data"):
                st.rerun()
                
        # Thêm thông tin giải thích
        st.markdown("---")
        st.markdown("""
        ### ℹ️ Giải thích
        - **Điểm dự kiến**: Điểm check-in sẽ nhận được nếu duy trì hiện trạng đến cuối tháng
        - **Tuần có check-in**: Số tuần đã có ít nhất 1 lần check-in / Tổng số tuần trong tháng  
        - **Cần thêm**: Số tuần check-in cần bổ sung để đạt 0.5 điểm (hoặc đạt >5 checkin)
        - **Điều kiện đạt điểm**: Có ≥2 tuần check-in HOẶC tổng số checkin > 5
        - **Cảnh báo**: Dựa trên số ngày còn lại trong tháng và tuần/checkin cần bổ sung
        """)
        
        return preview_df
        
    except Exception as e:
        st.error(f"Lỗi hiển thị preview: {e}")
        return pd.DataFrame()

def prepare_okr_excel_data(analyzer, user_manager, table_scores_map) -> List[Dict]:
    """Prepare data structure for OKRSheetGenerator"""
    users_data = []
    monthly_shifts = user_manager.monthly_okr_data 
    
    # Calculate checkin counts
    checkin_counts = defaultdict(int)
    if hasattr(user_manager, 'checkin_df') and user_manager.checkin_df is not None and not user_manager.checkin_df.empty:
         if 'checkin_user_id' in user_manager.checkin_df.columns:
             counts = user_manager.checkin_df.groupby('checkin_user_id').size()
             for uid, count in counts.items():
                 checkin_counts[str(uid)] = count
         elif 'user_id' in user_manager.checkin_df.columns:
             counts = user_manager.checkin_df.groupby('user_id').size()
             for uid, count in counts.items():
                 checkin_counts[str(uid)] = count
                 
    for user_id, user in user_manager.get_users().items():
        stats = {}
        
        # 1. Monthly Shift
        user_shift_entry = next((u for u in monthly_shifts if str(u.get('user_id', '')) == str(user_id)), None)
        shift_val = 0
        if user_shift_entry:
            shift_val = user_shift_entry.get('okr_shift_monthly', 0)
        
        if shift_val != 0:
             shift_percentage = (shift_val / 33.33) * 100
        else:
             shift_percentage = 0
             
        stats['okr_shift_display'] = f"{shift_percentage:.1f}%"
        
        if shift_percentage < 25: stats['shift_lt_25'] = True
        elif 25 <= shift_percentage < 50: stats['shift_25_50'] = True
        elif 50 <= shift_percentage < 75: stats['shift_50_75'] = True
        elif 75 <= shift_percentage < 100: stats['shift_75_100'] = True
        else: stats['shift_gt_100'] = True
        
        # 2. Checkins
        stats['has_okrs'] = "Yes" if user.co_OKR == 1 else "No"
        count = checkin_counts.get(str(user_id), 0)
        checkin_score = min(count * 2, 8)
        stats['checkin_score'] = True
        stats['checkin_score_val'] = checkin_score
        stats['collab_score'] = True
        
        # 3. Quality (Median)
        scores = []
        user_checkins = pd.DataFrame()
        if hasattr(user_manager, 'checkin_df') and user_manager.checkin_df is not None and not user_manager.checkin_df.empty:
             if 'checkin_user_id' in user_manager.checkin_df.columns:
                 user_checkins = user_manager.checkin_df[user_manager.checkin_df['checkin_user_id'].astype(str) == str(user_id)]
        
        if not user_checkins.empty:
            ts_col = 'checkin_since_timestamp'
            if ts_col not in user_checkins.columns and 'checkin_since' in user_checkins.columns: 
                 ts_col = 'checkin_since'
            
            if ts_col in user_checkins.columns:
                for _, row in user_checkins.iterrows():
                    ts_val = row[ts_col]
                    try:
                        ts_int = int(float(ts_val))
                        key = f"{user_id}_{ts_int}"
                        if key in table_scores_map:
                            scores.append(table_scores_map[key])
                    except:
                        pass

        median_q = np.median(scores) if scores else 0
        user.median_score = median_q # Store for reference

        if median_q >= 4.5: stats['quality_high'] = True
        elif median_q >= 2.5: stats['quality_med'] = True
        elif median_q > 0: stats['quality_low'] = True 
        
        # 4. Alignment / Priority / Impact (Section II)
        # Calculate scores using Hybrid Logic (Median -> Mode)
        col_map = {
            'align': 'Mức độ đóng góp vào mục tiêu công ty',
            'prio': 'Mức độ ưu tiên mục tiêu của Quý',
            'impact': 'Tính khó/tầm ảnh hưởng đến hệ thống'
        }
        
        # Get user goals df
        user_col = 'goal_user_id'
        if hasattr(analyzer, 'final_df') and analyzer.final_df is not None:
             if 'pro_goal_user_id' in analyzer.final_df.columns:
                 user_col = 'pro_goal_user_id'
             
             goal_id_col = 'goal_id'
             if 'pro_goal_id' in analyzer.final_df.columns:
                 goal_id_col = 'pro_goal_id'
                 
             user_goals_df = analyzer.final_df[analyzer.final_df[user_col] == user_id].drop_duplicates(goal_id_col)
             
             def calculate_hybrid_score(col_name):
                values = []
                if col_name in user_goals_df.columns:
                    for val_str in user_goals_df[col_name]:
                        if pd.notna(val_str) and isinstance(val_str, str) and len(val_str) > 0:
                            try:
                                first_char = val_str.strip()[0]
                                val = int(first_char)
                                values.append(val)
                            except:
                                pass
                
                if not values:
                    return 1.0 # Default if no data
                    
                # 1. Calculate Median
                med = np.median(values)
                
                # 2. Check if Median is integer (e.g. 3.0) or decimal (3.5)
                if med % 1 == 0:
                    return float(med)
                    
                # 3. If decimal, use Mode
                counts = {}
                for v in values:
                    counts[v] = counts.get(v, 0) + 1
                max_count = max(counts.values())
                modes = [k for k, v in counts.items() if v == max_count]
                return float(max(modes))

             align_score = calculate_hybrid_score(col_map['align'])
             prio_score = calculate_hybrid_score(col_map['prio'])
             impact_score = calculate_hybrid_score(col_map['impact'])
             
             # Map scores to keys
             # Alignment
             if align_score >= 1: stats['align_personal'] = True 
             if align_score >= 2: stats['align_indirect_1'] = True
             if align_score >= 3: stats['align_indirect_2'] = True
             if align_score >= 4: stats['align_direct_1'] = True
             if align_score >= 5: stats['align_direct_2'] = True
             
             # Use specific key based on score exact match or range? 
             # The sheet uses checkboxes for each level. Usually distinct or cumulative?
             # Based on goal_test logic, it seemed to calculate a single score.
             # The excel generator structure has distinct rows for each score 1-5.
             # So we should probably check only the row corresponding to the score.
             # Re-mapping:
             stats['align_personal'] = (align_score == 1)
             stats['align_indirect_1'] = (align_score == 2)
             stats['align_indirect_2'] = (align_score == 3)
             stats['align_direct_1'] = (align_score == 4)
             stats['align_direct_2'] = (align_score >= 5)

             stats['prio_normal'] = (prio_score == 1)
             stats['prio_important_1'] = (prio_score == 2)
             stats['prio_important_2'] = (prio_score == 3)
             stats['prio_very_important_1'] = (prio_score == 4)
             stats['prio_very_important_2'] = (prio_score >= 5)

             stats['impact_personal'] = (impact_score == 1)
             stats['impact_team_1'] = (impact_score == 2)
             stats['impact_team_2'] = (impact_score == 3)
             stats['impact_company_1'] = (impact_score == 4)
             stats['impact_company_2'] = (impact_score >= 5)
        
        users_data.append({'name': user.name, 'stats': stats})
        
    return users_data

def show_user_score_analysis(analyzer):
    """Show user score analysis using integrated monthly calculation"""
    st.subheader("🏆 Điểm số người dùng")
    
    try:
        # Kiểm tra có dữ liệu Monthly OKR Analysis
        if not ('monthly_okr_count' in st.session_state and st.session_state['monthly_okr_count']):
            st.warning("⚠️ No Monthly OKR Analysis data available. Please run Monthly OKR Analysis first!")
            return
        
        user_manager = create_user_manager_with_monthly_calculation(analyzer)
        user_manager.update_checkins()
        user_manager.update_okr_movement()
        user_manager.calculate_scores()
        
        users = user_manager.get_users()
        scores_df = _create_user_scores_dataframe(users)
        
        if not scores_df.empty:
            _display_score_metrics(scores_df)
            _display_score_distribution(scores_df)
            
            # Thông báo logic tính điểm checkin
            st.markdown("### 📋 Logic tính điểm Checkin:")
            st.markdown("""
            - 📅 **Điều kiện**: Nhân viên có ít nhất **2 tuần check-in** trong tháng hiện tại **HOẶC tổng số checkin > 5**
            - 🎯 **Điểm số**: Đủ điều kiện → **+0.5 điểm**, không đủ → **+0 điểm**
            - ⏰ **Thời điểm hiển thị**: Chỉ vào **tuần cuối cùng của tháng**
            """)
            
            # Chỉ hiển thị score tables khi ở tuần cuối cùng của tháng
            if DateUtils.is_last_week_of_month():
                st.success("✅ **Đang ở tuần cuối cùng của tháng** - Hiển thị điểm checkin dựa trên tiêu chí ≥2 tuần HOẶC >5 checkin")
                _display_score_tables(scores_df)
                _display_score_export_options(scores_df, list(users.values()) if isinstance(users, dict) else users, analyzer, user_manager)
            else:
                st.warning("⏳ **Chưa phải tuần cuối tháng** - Score tables sẽ hiển thị với điểm checkin thực tế vào tuần cuối cùng")
                # Vẫn hiển thị score tables nhưng với thông báo
                st.info("💡 **Bảng điểm hiện tại** (chưa tính điểm checkin 2 tuần):")
                _display_score_tables(scores_df)
                _display_score_export_options(scores_df, list(users.values()) if isinstance(users, dict) else users, analyzer, user_manager)
            
            return scores_df
        else:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"Error in user score analysis: {e}")
        return pd.DataFrame()

def _create_user_scores_dataframe(users: List[User]) -> pd.DataFrame:
    """Create dataframe from users for score analysis"""
    user_data = []
    for user in users:
        user_data.append({
            'Name': user.name,
            'Has OKR': 'Yes' if user.co_OKR == 1 else 'No',
            'Check-in': 'Yes' if user.checkin == 1 else 'No',  # Yes/No theo tiêu chí 3/4 tuần
            'OKR Movement (Monthly)': user.dich_chuyen_OKR,  # Lấy từ dịch chuyển tháng
            'Score': user.score
        })
    return pd.DataFrame(user_data)

def _display_score_metrics(scores_df: pd.DataFrame):
    """Display score summary metrics"""
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        avg_score = scores_df['Score'].mean()
        st.metric("Average Score", f"{avg_score:.2f}")
    
    with col2:
        high_performers = len(scores_df[scores_df['Score'] >= 3.0])
        st.metric("High Performers (≥3.0)", high_performers)
    
    with col3:
        median_score = scores_df['Score'].median()
        st.metric("Median Score", f"{median_score:.2f}")
    
    with col4:
        has_okr_count = len(scores_df[scores_df['Has OKR'] == 'Yes'])
        st.metric("Has OKR", f"{has_okr_count}/{len(scores_df)}")

def _display_score_distribution(scores_df: pd.DataFrame):
    """Display score distribution chart"""
    fig_scores = px.histogram(
        scores_df, 
        x='Score',
        nbins=20,
        title="Distribution of User Scores (with Monthly OKR Calculation)",
        labels={'Score': 'User Score', 'count': 'Number of Users'}
    )
    st.plotly_chart(fig_scores, use_container_width=True)

def _display_score_tables(scores_df: pd.DataFrame):
    """Display score tables"""
    # All performers sorted by score
    st.subheader("📊 Tất cả nhân viên")
    all_performers = scores_df.sort_values('Score', ascending=False)
    st.dataframe(all_performers, use_container_width=True, hide_index=True)

def _display_score_export_options(scores_df: pd.DataFrame, users: List[User], analyzer=None, user_manager=None):
    """Display export options for scores"""
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Export User Scores"):
            csv = scores_df.to_csv(index=False)
            st.download_button(
                label="Download User Scores CSV",
                data=csv,
                file_name=f"user_scores_monthly_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
    
    with col2:
        if st.button("📋 Export to Excel Format"):
            try:
                # 1. Fetch data from Table (Median Scores)
                with st.spinner("Fetching data from Base Table..."):
                     client = TableAPIClient()
                     table_scores = client.get_checkin_scores()
                     
                # 2. Prepare Data
                if analyzer and user_manager:
                     users_data = prepare_okr_excel_data(analyzer, user_manager, table_scores)
                     
                     # 3. Generate
                     generator = OKRSheetGenerator()
                     excel_buffer = generator.generate_excel(users_data, "Current Cycle")
                     
                     st.download_button(
                        label="Download Excel Report",
                        data=excel_buffer.getvalue(),
                        file_name=f"okr_report_monthly_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                     )
                     st.success(f"✅ Excel file ready with {len(users)} users")
                else:
                    st.error("Missing analyzer or user_manager")
            except Exception as e:
                st.error(f"Error generating Excel: {e}")

def show_data_summary(df: pd.DataFrame, analyzer):
    """Show data summary statistics"""
    st.subheader("📈 Tổng kết")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    metrics = [
        ("Total Goals", df['goal_id'].nunique()),
        ("Total KRs", df['kr_id'].nunique()),
        ("Total Checkins", df['checkin_id'].nunique()),
        ("Total Users", df['goal_user_name'].nunique()),
        ("Filtered Members", len(analyzer.filtered_members_df) if analyzer.filtered_members_df is not None else 0)
    ]
    
    for col, (label, value) in zip([col1, col2, col3, col4, col5], metrics):
        with col:
            st.metric(label, value)

def show_missing_analysis_section(analyzer):
    """Show missing goals and checkins analysis"""
    members_without_goals, members_without_checkins, members_with_goals_no_checkins = analyzer.analyze_missing_goals_and_checkins()
    
    _display_missing_summary_metrics(analyzer, members_without_goals, members_without_checkins, members_with_goals_no_checkins)
    _display_missing_visualizations(analyzer, members_without_goals, members_without_checkins, members_with_goals_no_checkins)

def _display_missing_summary_metrics(analyzer, members_without_goals: List, members_without_checkins: List, members_with_goals_no_checkins: List):
    """Display summary metrics for missing analysis"""
    col1, col2, col3, col4 = st.columns(4)
    
    total_members = len(analyzer.filtered_members_df) if analyzer.filtered_members_df is not None else 0
    
    metrics = [
        ("Total Filtered Members", total_members, None),
        ("Members Without Goals", len(members_without_goals), f"{len(members_without_goals)/total_members*100:.1f}%" if total_members > 0 else "0%"),
        ("Members Without Checkins", len(members_without_checkins), f"{len(members_without_checkins)/total_members*100:.1f}%" if total_members > 0 else "0%"),
        ("Has Goals but No Checkins", len(members_with_goals_no_checkins), f"{len(members_with_goals_no_checkins)/total_members*100:.1f}%" if total_members > 0 else "0%")
    ]
    
    for col, (label, value, delta) in zip([col1, col2, col3, col4], metrics):
        with col:
            st.metric(label, value, delta=delta)

def _display_missing_visualizations(analyzer, members_without_goals: List, members_without_checkins: List, members_with_goals_no_checkins: List):
    """Display visualizations for missing analysis"""
    st.subheader("📊 Dữ liệu thiếu")
    
    col1, col2 = st.columns(2)
    
    with col1:
        _display_goals_pie_chart_and_table(analyzer, members_without_goals)
    
    with col2:
        _display_checkins_pie_chart_and_table(members_with_goals_no_checkins)

def _display_goals_pie_chart_and_table(analyzer, members_without_goals: List):
    """Display goals pie chart and table"""
    total_members = len(analyzer.filtered_members_df) if analyzer.filtered_members_df is not None else 0
    members_with_goals = total_members - len(members_without_goals)
    
    goal_data = pd.DataFrame({
        'Status': ['Have Goals', 'No Goals'],
        'Count': [members_with_goals, len(members_without_goals)]
    })
    
    fig_goals = px.pie(
        goal_data, 
        values='Count', 
        names='Status',
        title="Goal Status Distribution",
        color_discrete_map={'Have Goals': '#00CC66', 'No Goals': '#FF6B6B'}
    )
    st.plotly_chart(fig_goals, use_container_width=True)
    
    # Members Without Goals table
    st.subheader("🚫 Thiếu mục tiêu")
    if members_without_goals:
        no_goals_df = pd.DataFrame(members_without_goals)
        st.dataframe(no_goals_df[['name', 'username', 'job', 'email']], use_container_width=True, height=300)
        
        csv_no_goals = no_goals_df.to_csv(index=False)
        st.download_button(
            label="📥 Download Members Without Goals",
            data=csv_no_goals,
            file_name=f"members_without_goals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key="download_no_goals"
        )
    else:
        st.success("✅ All filtered members have goals!")

def _display_checkins_pie_chart_and_table(members_with_goals_no_checkins: List):
    """Display checkins analysis"""
    if members_with_goals_no_checkins:
        st.subheader("⚠️ Thiếu checkin")
        st.warning("These members have set up goals but haven't made any checkins yet. They may need guidance or reminders.")
        
        goals_no_checkins_df = pd.DataFrame(members_with_goals_no_checkins)
        st.dataframe(goals_no_checkins_df[['name', 'username', 'job', 'email']], use_container_width=True, height=300)
        
        csv_goals_no_checkins = goals_no_checkins_df.to_csv(index=False)
        st.download_button(
            label="📥 Download Members with Goals but No Checkins",
            data=csv_goals_no_checkins,
            file_name=f"members_goals_no_checkins_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key="download_goals_no_checkins"
        )
    else:
        st.success("✅ All members with goals have made checkins!")

def _show_okr_user_selection(analyzer) -> List[str]:
    """Show interface to select specific OKR users"""
    try:
        # Get users with OKRs and their details
        if analyzer.final_df is None or analyzer.filtered_members_df is None:
            st.warning("Data not loaded yet. Please run analysis first.")
            return []
        
        users_with_goals = set(analyzer.final_df['goal_user_name'].dropna().unique())
        
        # Create options list with name and email
        okr_user_options = []
        okr_user_emails = {}
        
        for _, member in analyzer.filtered_members_df.iterrows():
            if member['name'] in users_with_goals and pd.notna(member['email']) and member['email'].strip():
                display_name = f"{member['name']} ({member['email']})"
                okr_user_options.append(display_name)
                okr_user_emails[display_name] = member['email'].strip()
        
        if not okr_user_options:
            st.warning("No OKR users with valid email addresses found")
            return []
        
        # Show multiselect
        st.write(f"Select recipients from {len(okr_user_options)} OKR users:")
        selected_users = st.multiselect(
            "Choose specific OKR users:",
            options=okr_user_options,
            default=okr_user_options,  # Select all by default
            key="okr_user_selection"
        )
        
        # Return selected emails
        return [okr_user_emails[user] for user in selected_users]
        
    except Exception as e:
        st.error(f"Error showing OKR user selection: {e}")
        return []

def show_okr_analysis(okr_shifts: List[Dict], reference_date: datetime, period: str = "weekly"):
    """Show OKR shift analysis with reference date"""
    period_label = "tuần" if period == "weekly" else "tháng"
    shift_key = 'okr_shift' if period == "weekly" else 'okr_shift_monthly'
    last_value_key = 'last_friday_value' if period == "weekly" else 'last_month_value'
    
    reference_label = f"thứ 6 {period_label} trước" if period == "weekly" else f"cuối {period_label} trước"
    
    # Display user count and reference information

    
    # Summary metrics
    _display_okr_summary_metrics(okr_shifts, shift_key)
    
    # OKR shift chart
    _display_okr_shift_chart(okr_shifts, shift_key, period_label, reference_label, reference_date)
    
    # Tables
    _display_okr_tables(okr_shifts, shift_key, last_value_key, period_label, reference_label)

def _display_okr_summary_metrics(okr_shifts: List[Dict], shift_key: str):
    """Display OKR summary metrics"""
    col1, col2, col3, col4 = st.columns(4)
    
    progress_users = len([u for u in okr_shifts if u[shift_key] > 0])
    stable_users = len([u for u in okr_shifts if u[shift_key] == 0])
    issue_users = len([u for u in okr_shifts if u[shift_key] < 0])
    avg_shift = np.mean([u[shift_key] for u in okr_shifts])
    
    total_users = len(okr_shifts)
    
    with col1:
        st.metric("Tiến bộ", progress_users, delta=f"{progress_users/total_users*100:.1f}%" if total_users > 0 else "0%")
    
    with col2:
        st.metric("Ổn định", stable_users, delta=f"{stable_users/total_users*100:.1f}%" if total_users > 0 else "0%")
    
    with col3:
        st.metric("Cần hỗ trợ", issue_users, delta=f"{issue_users/total_users*100:.1f}%" if total_users > 0 else "0%")
    
    with col4:
        st.metric("Dịch chuyển TB", f"{avg_shift:.2f}")

def _display_okr_shift_chart(okr_shifts: List[Dict], shift_key: str, period_label: str, reference_label: str, reference_date: datetime):
    """Display OKR shift chart"""
    okr_df = pd.DataFrame(okr_shifts)
    
    fig = px.bar(
        okr_df, 
        x='user_name', 
        y=shift_key,
        title=f"Dịch chuyển OKR so với {reference_label} ({reference_date.strftime('%d/%m/%Y')})",
        color=shift_key,
        color_continuous_scale=['red', 'yellow', 'green'],
        labels={
            'user_name': 'Nhân viên',
            shift_key: f'Dịch chuyển OKR ({period_label})'
        }
    )
    fig.update_xaxes(tickangle=45)
    fig.update_layout(height=500)
    st.plotly_chart(fig, use_container_width=True)

def _display_okr_tables(okr_shifts: List[Dict], shift_key: str, last_value_key: str, period_label: str, reference_label: str):
    """Display OKR performance tables"""
    okr_df = pd.DataFrame(okr_shifts)
    
    # All performers with positive shift
    st.subheader(f"📈 Tất cả nhân viên tiến bộ ({period_label})")
    positive_performers = okr_df[okr_df[shift_key] > 0]
    if not positive_performers.empty:
        display_cols = ['user_name', shift_key, 'current_value', last_value_key]
        display_df = positive_performers[display_cols].round(2)
        display_df.columns = ['Nhân viên', f'Dịch chuyển ({period_label})', 'Giá trị hiện tại', f'Giá trị {reference_label}']
        st.dataframe(display_df, use_container_width=True, hide_index=True)
    else:
        pass
    
    # Issues
    issue_users = okr_df[okr_df[shift_key] < 0]
    if not issue_users.empty:
        st.subheader(f"⚠️ Nhân viên cần hỗ trợ ({period_label})")
        display_cols = ['user_name', shift_key, 'current_value', last_value_key]
        display_df = issue_users[display_cols].round(2)
        display_df.columns = ['Nhân viên', f'Dịch chuyển ({period_label})', 'Giá trị hiện tại', f'Giá trị {reference_label}']
        st.dataframe(display_df, use_container_width=True, hide_index=True)

def show_checkin_analysis(period_checkins: List[Dict], overall_checkins: List[Dict], 
                         last_friday: datetime, quarter_start: datetime):
    """Show checkin behavior analysis"""
    period_df = pd.DataFrame(period_checkins)
    overall_df = pd.DataFrame(overall_checkins)
    
    # Period analysis
    st.subheader(f"📅 Phân tích kỳ ({quarter_start.strftime('%d/%m/%Y')} - {last_friday.strftime('%d/%m/%Y')})")
    _display_period_checkin_metrics(period_checkins)
    _display_checkin_distribution_chart(period_checkins)
    
    # Overall analysis
    st.subheader("🏆 Hoạt động tích cực nhất")
    _display_overall_checkin_analysis(overall_checkins, quarter_start)

def _display_period_checkin_metrics(period_checkins: List[Dict]):
    """Display period checkin metrics"""
    col1, col2, col3, col4 = st.columns(4)
    
    active_users = len([u for u in period_checkins if u['checkin_count_period'] > 0])
    avg_checkins = np.mean([u['checkin_count_period'] for u in period_checkins]) if period_checkins else 0
    max_checkins = max([u['checkin_count_period'] for u in period_checkins]) if period_checkins else 0
    avg_rate = np.mean([u['checkin_rate_period'] for u in period_checkins]) if period_checkins else 0
    
    with col1:
        total_users = len(period_checkins)
        st.metric("Active Users", active_users, delta=f"{active_users/total_users*100:.1f}%" if total_users > 0 else "0%")
    
    with col2:
        st.metric("Avg Checkins/User", f"{avg_checkins:.1f}")
    
    with col3:
        st.metric("Max Checkins", max_checkins)
    
    with col4:
        st.metric("Avg Checkin Rate", f"{avg_rate:.1f}%")

def _display_checkin_distribution_chart(period_checkins: List[Dict]):
    """Display checkin distribution chart"""
    checkin_counts = [u['checkin_count_period'] for u in period_checkins]
    
    fig = go.Figure()
    fig.add_trace(go.Histogram(x=checkin_counts, nbinsx=10, name="Checkin Distribution"))
    fig.update_layout(
        title="Distribution of Checkins per User (Period)",
        xaxis_title="Number of Checkins",
        yaxis_title="Number of Users",
        height=400
    )
    st.plotly_chart(fig, use_container_width=True)

def _display_overall_checkin_analysis(overall_checkins: List[Dict], quarter_start: datetime):
    """Display overall checkin analysis"""
    overall_df = pd.DataFrame(overall_checkins)
    
    today = datetime.now()
    weeks_in_quarter = max((today - quarter_start).days / 7, 1)
    
    # Calculate last week boundaries
    days_since_monday = today.weekday()
    monday_this_week = today - timedelta(days=days_since_monday)
    monday_last_week = monday_this_week - timedelta(days=7)
    sunday_last_week = monday_last_week + timedelta(days=6)
    

    
    # Display table - all employees sorted by total checkins
    all_overall = overall_df.sort_values('total_checkins', ascending=False).copy()
    display_df = all_overall[[
        'user_name', 'total_checkins', 'checkin_frequency_per_week', 'last_week_checkins'
    ]].copy()
    
    display_df.columns = ['👤 Nhân viên', '📊 Tổng checkin', '⚡ Tần suất/tuần (quý)', '📅 Checkin tuần trước']
    display_df['⚡ Tần suất/tuần (quý)'] = display_df['⚡ Tần suất/tuần (quý)'].round(2)
    
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    
    # Summary metrics
    _display_overall_checkin_summary_metrics(overall_df, quarter_start)

def _display_overall_checkin_summary_metrics(overall_df: pd.DataFrame, quarter_start: datetime):
    """Display overall checkin summary metrics"""
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_last_week = overall_df['last_week_checkins'].sum()
        st.metric("🗓️ Tổng checkin tuần trước", total_last_week)
    
    with col2:
        active_last_week = len(overall_df[overall_df['last_week_checkins'] > 0])
        st.metric("👥 Người hoạt động tuần trước", active_last_week)
    
    with col3:
        avg_frequency_quarter = overall_df['checkin_frequency_per_week'].mean()
        st.metric("📈 Tần suất TB/tuần (quý)", f"{avg_frequency_quarter:.2f}")
    
    with col4:
        max_frequency_quarter = overall_df['checkin_frequency_per_week'].max()
        st.metric("🏆 Tần suất cao nhất/tuần", f"{max_frequency_quarter:.2f}")

def show_export_options(df: pd.DataFrame, okr_shifts: List, okr_shifts_monthly: List, 
                       period_checkins: List, overall_checkins: List, analyzer, cycle_name: str = "Quarter"):
    """Show data export options"""
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    
    export_configs = [
        (col1, "📊 Export Full Dataset", df, "okr_full_dataset"),
        (col2, "🎯 Export Weekly OKR Shifts", pd.DataFrame(okr_shifts), "okr_shifts_weekly"),
        (col4, "📝 Export Period Checkins", pd.DataFrame(period_checkins), "period_checkins"),
        (col5, "📈 Export Overall Checkins", pd.DataFrame(overall_checkins), "overall_checkins"),
        (col6, "👥 Export Filtered Members", analyzer.filtered_members_df, "filtered_members")
    ]
    
    for col, label, data, filename_prefix in export_configs:
        with col:
            if st.button(label):
                if data is not None and not data.empty:
                    csv = data.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
    
    # Monthly export (conditional)
    with col3:
        if okr_shifts_monthly and st.button("🗓️ Export Monthly OKR Shifts"):
            csv = pd.DataFrame(okr_shifts_monthly).to_csv(index=False)
            st.download_button(
                label="Download CSV",
                data=csv,
                file_name=f"okr_shifts_monthly_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
            
    # New Excel Format Export
    st.markdown("---")
    st.subheader("📊 Xuất báo cáo Excel (Mới)")
    col_ex_1, col_ex_2 = st.columns(2)
    
    with col_ex_1:
        if st.button("📥 Generate Excel Report (New Format)"):
            with st.spinner("Generating Excel report..."):
                try:
                    # 1. Ensure Alignment Data is available
                    alignment_analysis = analyzer.analyze_alignment_contribution()
                    
                    # 2. Collect all user data
                    # We iterate over filtered members to ensure we cover everyone relevant, 
                    # or just those with Data?
                    # Let's use users present in okr_shifts (Weekly) as base, as they have OKR activity
                    # OR use filtered_members_df if we want to show empty columns for everyone
                    
                    # Logic: Use users from Weekly Shifts (Active OKR users)
                    # users_to_export = [u['user_name'] for u in okr_shifts] 
                    
                    # Better: Use all users with Goals
                    users_with_goals = df['goal_user_name'].dropna().unique()
                    
                    excel_users_data = []
                    for user_name in users_with_goals:
                        stats = analyzer.get_user_excel_data(user_name, okr_shifts, period_checkins, alignment_analysis)
                        excel_users_data.append(stats)
                        
                    # 3. Generate Excel
                    if excel_users_data:
                        generator = OKRSheetGenerator()
                        
                        excel_file = generator.generate_excel(excel_users_data, cycle_name)
                        
                        st.download_button(
                            label="📥 Click here to Download Excel Report",
                            data=excel_file.getvalue(),
                            file_name=f"OKR_Report_New_Format_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success(f"✅ Generated report for {len(excel_users_data)} users!")
                    else:
                        st.warning("No user data found to generate report.")
                        
                except Exception as e:
                    st.error(f"Error generating Excel: {e}")


def run_analysis(analyzer, selected_cycle: Dict, show_missing_analysis: bool):
    """Run the main analysis"""
    st.header(f"📊 Kết quả - {selected_cycle['name']}")
    
    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def update_progress(message, progress):
        status_text.text(message)
        progress_bar.progress(progress)
    
    try:
        # Load data
        df = analyzer.load_and_process_data(update_progress)
        
        if df is None or df.empty:
            st.error("❌ Failed to load data. Please check your API tokens and try again.")
            return
            
        progress_bar.empty()
        status_text.empty()
        
        # Run analysis sections
        show_data_summary(df, analyzer)
        
        if show_missing_analysis:
            st.subheader("🚨 Thiếu dữ liệu")
            with st.spinner("Analyzing missing goals and checkins..."):
                show_missing_analysis_section(analyzer)
        
        # Weekly OKR Analysis
        st.subheader("🎯 OKR tuần")
        with st.spinner("Calculating weekly OKR shifts..."):
            okr_shifts = analyzer.calculate_okr_shifts_by_user()
        
        if okr_shifts:
            show_okr_analysis(okr_shifts, DateUtils.get_last_friday_date(), "weekly")
        else:
            st.warning("No weekly OKR shift data available")
        
        # Monthly OKR Analysis
        okr_shifts_monthly = []
        if DateUtils.should_calculate_monthly_shift():
            st.subheader("🗓️ OKR tháng")
            with st.spinner("Calculating monthly OKR shifts..."):
                okr_shifts_monthly = analyzer.calculate_okr_shifts_by_user_monthly()
            
            if okr_shifts_monthly:
                # Lưu TOÀN BỘ dữ liệu Monthly OKR Analysis (tên + dịch chuyển tháng)
                monthly_okr_data = []
                for shift in okr_shifts_monthly:
                    monthly_okr_data.append({
                        'user_name': shift['user_name'],
                        'okr_shift_monthly': shift.get('okr_shift_monthly', 0)
                    })
                
                st.session_state['monthly_okr_data'] = monthly_okr_data
                st.session_state['monthly_okr_users'] = set([shift['user_name'] for shift in okr_shifts_monthly])
                st.session_state['monthly_okr_count'] = len(monthly_okr_data)
                
                show_okr_analysis(okr_shifts_monthly, DateUtils.get_last_month_end_date(), "monthly")
                st.success(f"✅ 'Tất cả nhân viên tiến bộ (tháng)' table: {len(monthly_okr_data)} users saved → will be used for User Score Analysis")
            else:
                st.warning("No monthly OKR shift data available")
        else:
            current_month = datetime.now().month
            quarter_months = {1: "Q1", 4: "Q2", 7: "Q3", 10: "Q4"}

        
        # User Score Analysis (sau khi đã có Monthly OKR Analysis data)  
        st.subheader("🏆 Điểm số")
        with st.spinner("Calculating user scores with monthly OKR movement..."):
            show_user_score_analysis(analyzer)
        
        # Checkin Analysis
        st.subheader("📝 Phân tích checkin")
        with st.spinner("Analyzing checkin behavior..."):
            period_checkins, overall_checkins = analyzer.analyze_checkin_behavior()
        
        if period_checkins and overall_checkins:
            show_checkin_analysis(period_checkins, overall_checkins, DateUtils.get_last_friday_date(), DateUtils.get_quarter_start_date())
        else:
            st.warning("No checkin data available")
        
        # Export options
        st.subheader("💾 Xuất dữ liệu")
        show_export_options(df, okr_shifts, okr_shifts_monthly, period_checkins, overall_checkins, analyzer, selected_cycle['name'])
        
        st.success("✅ Analysis completed successfully!")
        
    except Exception as e:
        st.error(f"❌ Analysis failed: {e}")
    finally:
        progress_bar.empty()
        status_text.empty()

def send_email_report(analyzer, email_generator: EmailReportGenerator, selected_cycle: Dict, 
                     email_from: str, email_password: str, email_to: str):
    """Send single email report including monthly data when applicable"""
    st.header("📧 Báo cáo email")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    def update_progress(message, progress):
        status_text.text(message)
        progress_bar.progress(progress)
    
    try:
        # Load data
        update_progress("Loading data for email report...", 0.1)
        df = analyzer.load_and_process_data(update_progress)
        
        if df is None or df.empty:
            st.error("❌ Failed to load data for email report")
            return
        
        # Analyze data
        update_progress("Analyzing missing goals and checkins...", 0.25)
        members_without_goals, members_without_checkins, members_with_goals_no_checkins = analyzer.analyze_missing_goals_and_checkins()
        
        update_progress("Calculating weekly OKR shifts...", 0.4)
        okr_shifts = analyzer.calculate_okr_shifts_by_user()
        
        # Calculate monthly if applicable
        okr_shifts_monthly = []
        if DateUtils.should_calculate_monthly_shift():
            update_progress("Calculating monthly OKR shifts...", 0.55)
            okr_shifts_monthly = analyzer.calculate_okr_shifts_by_user_monthly()
        
        # Create and send email
        update_progress("Creating email content...", 0.7)
        html_content = email_generator.create_email_content(
            analyzer, selected_cycle, members_without_goals, members_without_checkins,
            members_with_goals_no_checkins, okr_shifts, okr_shifts_monthly
        )
        
        update_progress("Sending email...", 0.9)
        subject = f"📊 Báo cáo tiến độ OKR & Checkin - {selected_cycle['name']} - {datetime.now().strftime('%d/%m/%Y')}"
        
        success, message = email_generator.send_email_report(
            email_from, email_password, email_to, subject, html_content
        )
        
        # Display results
        if success:
            st.success(f"✅ {message}")
            monthly_note = " (bao gồm phân tích tháng)" if okr_shifts_monthly else ""
            st.success(f"📧 Email report sent to: {email_to}{monthly_note}")
            
            if st.checkbox("📋 Show email preview", value=False):
                st.subheader("Xem trước email")
                st.components.v1.html(html_content, height=600, scrolling=True)
        else:
            st.error(f"❌ {message}")
            
    except Exception as e:
        st.error(f"❌ Error sending email report: {e}")
    finally:
        progress_bar.empty()
        status_text.empty()

def send_email_report_enhanced(analyzer, email_generator: EmailReportGenerator, selected_cycle: Dict,
                              email_from: str, email_password: str, recipient_option: str, 
                              selected_okr_emails: Optional[List[str]] = None):
    """Enhanced email sending with bulk capability and Excel attachment"""
    st.header("📧 Email nâng cao")
    
    # Determine recipients
    recipients = _get_email_recipients(analyzer, recipient_option, selected_okr_emails)
    if not recipients:
        return
    
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # Load and analyze data (reuse existing logic but simplified)
        status_text.text("Loading data...")
        progress_bar.progress(0.1)
        
        df = analyzer.load_and_process_data()
        if df is None or df.empty:
            st.error("❌ Failed to load data for email report")
            return
        
        # Get analysis data
        status_text.text("Analyzing data...")
        progress_bar.progress(0.4)
        
        members_without_goals, members_without_checkins, members_with_goals_no_checkins = analyzer.analyze_missing_goals_and_checkins()
        okr_shifts = analyzer.calculate_okr_shifts_by_user()
        okr_shifts_monthly = analyzer.calculate_okr_shifts_by_user_monthly() if DateUtils.should_calculate_monthly_shift() else []
        
        # Save users from Monthly OKR Analysis for reference
        if okr_shifts_monthly:
            # Lưu TOÀN BỘ dữ liệu Monthly OKR Analysis cho email export
            monthly_okr_data = []
            for shift in okr_shifts_monthly:
                monthly_okr_data.append({
                    'user_name': shift['user_name'],
                    'okr_shift_monthly': shift.get('okr_shift_monthly', 0)
                })
            
            st.session_state['monthly_okr_data'] = monthly_okr_data
            monthly_okr_users = set([shift['user_name'] for shift in okr_shifts_monthly])
            st.session_state['monthly_okr_users'] = monthly_okr_users
            st.session_state['monthly_okr_count'] = len(monthly_okr_users)

        
        # Create Excel attachment if it's last week of month
        excel_attachment = None
        if DateUtils.is_last_week_of_month():
            status_text.text("Creating Excel attachment for last week of month...")
            progress_bar.progress(0.6)
            
            try:
                # Create UserManager to get user scores for Excel export
                user_manager = create_user_manager_with_monthly_calculation(analyzer)
                if user_manager:
                    user_manager.update_checkins()
                    user_manager.update_okr_movement()
                    user_manager.calculate_scores()
                    
                    users = user_manager.get_users()
                    if users:
                        wb = export_to_excel(users)
                        excel_buffer = BytesIO()
                        wb.save(excel_buffer)
                        excel_buffer.seek(0)
                        excel_attachment = excel_buffer
                        st.info(f"📊 Excel attachment created with {len(users)} users (last week of month)")
            except Exception as e:
                st.warning(f"⚠️ Could not create Excel attachment: {e}")
                excel_attachment = None
        
        # Create email content
        status_text.text("Creating email content...")
        progress_bar.progress(0.7)
        
        html_content = email_generator.create_email_content(
            analyzer, selected_cycle, members_without_goals, members_without_checkins,
            members_with_goals_no_checkins, okr_shifts, okr_shifts_monthly
        )
        
        # Send emails
        status_text.text("Sending emails...")
        progress_bar.progress(0.8)
        
        subject = f"📊 Báo cáo tiến độ OKR & Checkin - {selected_cycle['name']} - {datetime.now().strftime('%d/%m/%Y')}"
        
        success, message, errors = email_generator.send_email_report_bulk(
            email_from, email_password, recipients, subject, html_content, excel_attachment
        )
        
        # Display results
        progress_bar.progress(1.0)
        if success:
            st.success(f"✅ {message}")
            recipient_msg = f"📧 Sent to {len(recipients)} recipients"
            if excel_attachment:
                recipient_msg += " (với file Excel đính kèm)"
            st.success(recipient_msg)
            
            if errors:
                st.warning("⚠️ Some emails failed:")
                for error in errors:
                    st.text(f"- {error}")
        else:
            st.error(f"❌ {message}")
            
    except Exception as e:
        st.error(f"❌ Error: {e}")
    finally:
        progress_bar.empty()
        status_text.empty()


def _get_email_recipients(analyzer, recipient_option: str, selected_okr_emails: Optional[List[str]] = None) -> List[str]:
    """Get email recipients based on option"""
    if recipient_option == "all":
        recipients = get_email_list(analyzer)
        if not recipients:
            st.error("No email addresses found in member data")
            return []
    elif recipient_option == "special":
        recipients = get_default_recipients()
    elif recipient_option == "all_with_goals":
        # Lấy email từ những người thực sự có goal (goal_user_name)
        recipients = get_emails_from_total_users_in_summary(analyzer)
        
        # Đảm bảo admin email luôn có trong danh sách
        admin_email = "xnk3@apluscorp.vn"
        if admin_email not in recipients:
            recipients.append(admin_email)
    elif recipient_option == "okr_users":
        recipients = get_emails_of_total_users_with_okr(analyzer)
        if not recipients:
            st.warning("No OKR user emails found in total users")
            st.error("No email addresses found in total user data")
            return []
    elif recipient_option == "select_okr_users":
        if not selected_okr_emails:
            st.error("No OKR users selected")
            return []
        recipients = selected_okr_emails
    else:
        st.error("Invalid recipient option")
        return []
    
    return recipients



def setup_sidebar_configuration():
    """Setup sidebar configuration"""
    with st.sidebar:
        st.header("⚙️ Cài đặt")
        
        # Token status
        st.subheader("🔑 Trạng thái API")
        goal_token = os.getenv("GOAL_ACCESS_TOKEN")
        account_token = os.getenv("ACCOUNT_ACCESS_TOKEN")
        
        if goal_token:
            st.success("✅ Goal Access Token: Loaded")
        else:
            st.error("❌ Goal Access Token: Missing")
            
        if account_token:
            st.success("✅ Account Access Token: Loaded")
        else:
            st.error("❌ Account Access Token: Missing")
        
        return goal_token, account_token

def get_emails_from_total_users_in_summary(analyzer) -> List[str]:
    """Get email list from Total Users shown in Data Summary with fallback"""
    try:
        if analyzer.filtered_members_df is None:
            st.warning("⚠️ Filtered members data not loaded yet")
            return []
            
        if analyzer.final_df is None:
            st.warning("⚠️ Final data not loaded yet, returning all valid emails")
            return get_email_list(analyzer)  # Fallback to all filtered members
        
        # Check email column
        if 'email' not in analyzer.filtered_members_df.columns:
            st.error("❌ No email column found in member data")
            return []
        
        # Get Total Users - đúng như trong Data Summary: df['goal_user_name'].nunique()
        total_users_with_goals = set(analyzer.final_df['goal_user_name'].dropna().unique())
        st.info(f"🎯 Found {len(total_users_with_goals)} users with goals: {list(total_users_with_goals)[:5]}...")
        
        # Get all valid emails from filtered members
        all_member_emails = []
        for _, member in analyzer.filtered_members_df.iterrows():
            member_email = member.get('email', '')
            if pd.notna(member_email) and str(member_email).strip() and '@' in str(member_email):
                all_member_emails.append(str(member_email).strip())
        
        # Match by name and get emails from filtered members with goals
        total_users_emails = []
        for _, member in analyzer.filtered_members_df.iterrows():
            member_name = member.get('name', '')
            member_email = member.get('email', '')
            
            # Chỉ lấy email của những người có tên trong Total Users
            if (member_name in total_users_with_goals and 
                pd.notna(member_email) and 
                str(member_email).strip() and 
                '@' in str(member_email)):
                
                total_users_emails.append(str(member_email).strip())
        
        # Always add default admin email to users with goals
        admin_email = "xnk3@apluscorp.vn"
        
        # Return users with goals emails, or fallback to all if none found
        if total_users_emails:
            # Add admin email to the list
            total_users_emails.append(admin_email)
            final_emails = list(set(total_users_emails))  # Remove duplicates
            st.success(f"✅ Found {len(final_emails)} emails for users with goals (including admin)")
            return final_emails
        else:
            # Add admin email to fallback list too
            all_member_emails.append(admin_email)
            final_emails = list(set(all_member_emails))
            st.warning(f"⚠️ No email matches found for users with goals. Fallback to all {len(final_emails)} member emails (including admin)")
            return final_emails
        
    except Exception as e:
        st.error(f"❌ Error getting emails: {str(e)}")
        # Ultimate fallback - always include admin email
        fallback_emails = get_email_list(analyzer)
        fallback_emails.append("xnk3@apluscorp.vn")
        return list(set(fallback_emails))

def get_emails_of_okr_users(analyzer) -> List[str]:
    """Get email list of users who have OKRs (legacy function, now uses Total Users)"""
    return get_emails_from_total_users_in_summary(analyzer)

def setup_cycle_selection(analyzer) -> Dict:
    """Setup cycle selection in sidebar"""
    with st.sidebar:
        st.subheader("📅 Chu kỳ")
        
        with st.spinner("🔄 Loading available cycles..."):
            cycles = analyzer.get_cycle_list()

        if not cycles:
            st.error("❌ Could not load cycles. Please check your API tokens and connection.")
            return None

        # Determine default index based on current date
        default_index = 0
        now = datetime.now()
        hcm_tz = timezone(timedelta(hours=7)) # Ensure timezone definition if not global, but assuming global or re-import if needed. 
        # Actually hcm_tz might be defined globally or we use offset-naive comparison if cycles use consistent naive approach in this app context. 
        # In goal.py we used hcm_tz. goal_app.py has `from datetime import datetime, timedelta, timezone` likely.
        # Let's check how `get_cycle_list` returns dates. Usually API returns strings or parsed datetimes.
        
        for i, cycle in enumerate(cycles):
            # cycle['start_time'] assumed to be parsed datetime or we parse it
            # Safest is to use the logic from goal.py
            try:
                start_time = cycle.get('start_time')
                if isinstance(start_time, str):
                    # Parse if string (fallback if not already parsed)
                     # 2024-01-01T00:00:00Z format usually
                    pass # goal_app.py might have it parsed in get_cycle_list. 
                    # Let's rely on standard logic similar to goal.py but robust.
                
                # Simplified check: if cycle['name'] contains current quarter hints? 
                # Better: Use date logic if start_time is reliable.
                if isinstance(start_time, datetime):
                     # Convert to HCM or naive
                     start_hcm = start_time.astimezone(timezone(timedelta(hours=7))).replace(tzinfo=None)
                     end_hcm = start_hcm + timedelta(days=100)
                     if start_hcm <= now <= end_hcm:
                         default_index = i
                         break
            except Exception:
                continue

        cycle_options = {f"{cycle['name']} ({cycle['formatted_start_time']})": cycle for cycle in cycles}
        selected_cycle_name = st.selectbox(
            "Select Cycle",
            options=list(cycle_options.keys()),
            index=default_index,
            help="Choose the quarterly cycle to analyze"
        )
        
        selected_cycle = cycle_options[selected_cycle_name]
        analyzer.checkin_path = selected_cycle['path']
        
        # Auto-load filtered members để tránh lỗi trong sidebar
        if analyzer.filtered_members_df is None:
            with st.spinner("🔄 Loading filtered members..."):
                try:
                    analyzer.get_filtered_members()
                    st.success(f"✅ Loaded {len(analyzer.filtered_members_df)} filtered members")
                except Exception as e:
                    st.error(f"❌ Failed to load filtered members: {e}")
        
        st.info(f"🎯 **Selected:** {selected_cycle['name']}")
        
        return selected_cycle

def get_total_users_emails_count(analyzer) -> int:
    """Get count of Total Users emails (for Data Summary)"""
    try:
        return len(get_emails_from_total_users_in_summary(analyzer))
    except Exception as e:
        return 0

def get_all_member_emails_count(analyzer) -> int:
    """Get count of all member emails for sidebar display"""
    try:
        if analyzer.filtered_members_df is None:
            return 0
        
        if 'email' not in analyzer.filtered_members_df.columns:
            return 0
        
        valid_email_count = 0
        for _, member in analyzer.filtered_members_df.iterrows():
            email = member.get('email', '')
            if pd.notna(email) and str(email).strip() and '@' in str(email):
                valid_email_count += 1
        
        return valid_email_count
        
    except Exception:
        return 0

def get_total_user_emails_count(analyzer) -> int:
    """Get count of all valid email addresses in total account users (not filtered)"""
    try:
        total_users_df = analyzer.get_total_account_users()
        if total_users_df is None or total_users_df.empty:
            return 0
        
        if 'email' not in total_users_df.columns:
            return 0
        
        valid_emails = 0
        for _, user in total_users_df.iterrows():
            user_email = user.get('email', '')
            if pd.notna(user_email) and str(user_email).strip() and '@' in str(user_email):
                valid_emails += 1
        
        return valid_emails
    except Exception as e:
        return 0

def get_emails_of_total_users_with_okr(analyzer) -> List[str]:
    """Get email list of total users (not filtered) who have OKRs"""
    try:
        # Get total account users
        total_users_df = analyzer.get_total_account_users()
        if total_users_df is None or total_users_df.empty:
            return []
        
        if 'email' not in total_users_df.columns:
            return []
        
        # Get all valid emails from total users
        all_total_emails = []
        for _, user in total_users_df.iterrows():
            user_email = user.get('email', '')
            if pd.notna(user_email) and str(user_email).strip() and '@' in str(user_email):
                all_total_emails.append(str(user_email).strip())
        
        # If no OKR data loaded yet, return all total emails
        if analyzer.final_df is None or analyzer.final_df.empty:
            return all_total_emails
        
        # Get unique users who have goals
        users_with_goals = set()
        for _, row in analyzer.final_df.iterrows():
            goal_user_name = row.get('goal_user_name', '')
            if goal_user_name:
                users_with_goals.add(goal_user_name)
        
        # Match by name and get emails from total users
        okr_users_emails = []
        for _, user in total_users_df.iterrows():
            user_name = user.get('name', '')
            user_email = user.get('email', '')
            
            if (user_name in users_with_goals and 
                pd.notna(user_email) and 
                str(user_email).strip() and
                '@' in str(user_email)):
                
                okr_users_emails.append(str(user_email).strip())
        
        # Return OKR user emails if found, otherwise all total emails
        return okr_users_emails if okr_users_emails else all_total_emails
        
    except Exception as e:
        return []

def setup_analysis_options():
    """Setup analysis options in sidebar"""
    with st.sidebar:
        st.subheader("📊 Tùy chọn")
        return st.checkbox("Show Missing Goals & Checkins Analysis", value=True)



def setup_enhanced_email_configuration(analyzer):
    """Setup enhanced email configuration in sidebar"""
    with st.sidebar:
        st.subheader("📧 Cấu hình email")
        
        # Recipient options - mặc định chọn all_with_goals
        recipient_option = st.radio(
            "Send emails to:",
            ["all_with_goals", "special", "all", "okr_users"],
            format_func=lambda x: {
                "special": "Special recipients only (tts122403@gmail.com)",
                "all": "All filtered members (backup option)",
                "all_with_goals": "Nhân viên có Goal + Admin",
                "okr_users": "People with OKRs (legacy option)"
            }[x],
            index=1  # Mặc định chọn special (theo yêu cầu)
        )
        
        # Display recipient info với analyzer
        _display_recipient_email_list(recipient_option, analyzer)
        
        return recipient_option, None

def _display_recipient_email_list(recipient_option: str, analyzer=None, selected_okr_emails: Optional[List[str]] = None):
    """Display recipient email list in sidebar"""
    st.markdown("---")
    st.markdown("📧 **Danh sách email sẽ gửi:**")
    
    try:
        # Get actual email recipients
        recipients = _get_email_recipients(analyzer, recipient_option, selected_okr_emails)
        
        # Đảm bảo admin email luôn có trong danh sách "all_with_goals"
        admin_email = "xnk3@apluscorp.vn"
        if recipient_option == "all_with_goals" and admin_email not in recipients:
            recipients.append(admin_email)
            st.info(f"✅ Đã thêm admin email: {admin_email}")
        
        if not recipients:
            st.warning("⚠️ Không tìm thấy email nào để gửi")
            return
            
        # Display email count
        st.success(f"📊 Tổng số: **{len(recipients)} email**")
        
        # Display email list in scrollable container
        with st.container():
            # Show first few emails directly
            display_count = min(5, len(recipients))
            for i, email in enumerate(recipients[:display_count]):
                st.text(f"{i+1}. {email}")
            
            # Show remaining emails in expander if more than 5
            if len(recipients) > 5:
                with st.expander(f"👀 Xem thêm {len(recipients) - 5} email khác"):
                    for i, email in enumerate(recipients[5:], start=6):
                        st.text(f"{i}. {email}")
                        
    except Exception as e:
        st.error(f"❌ Lỗi khi lấy danh sách email: {str(e)}")

def _display_recipient_info_with_count(recipient_option: str, analyzer=None, selected_okr_emails: Optional[List[str]] = None):
    """Display recipient information with email count (legacy function - deprecated)"""
    pass

def main():
    """Main application entry point"""
    st.title("🎯 Phân tích OKR")
    st.markdown("---")

    # Setup configuration
    goal_token, account_token = setup_sidebar_configuration()
    
    if not goal_token or not account_token:
        st.error("❌ API tokens not found in environment variables. Please set GOAL_ACCESS_TOKEN and ACCOUNT_ACCESS_TOKEN.")

        st.code("""
GOAL_ACCESS_TOKEN=your_goal_token_here
ACCOUNT_ACCESS_TOKEN=your_account_token_here
        """)
        return

    # Initialize system
    try:
        analyzer = OKRAnalysisSystem(goal_token, account_token)
        email_generator = EmailReportGenerator()
        # Lưu analyzer vào session state để sử dụng trong sidebar
        st.session_state.analyzer = analyzer
    except Exception as e:
        st.error(f"Failed to initialize analyzer: {e}")
        return

    # Setup sidebar options
    selected_cycle = setup_cycle_selection(analyzer)
    if not selected_cycle:
        return
    
    show_missing_analysis = setup_analysis_options()
    recipient_option, custom_emails = setup_enhanced_email_configuration(analyzer)

    # Auto-run analysis để đảm bảo data được load
    auto_run_key = f"auto_analysis_{selected_cycle['path']}"
    if auto_run_key not in st.session_state:
        st.session_state[auto_run_key] = True
        with st.spinner("🚀 Auto-running analysis..."):
            run_analysis(analyzer, selected_cycle, show_missing_analysis)

    # Main action buttons
    col1, col2 = st.columns(2)
    
    with col1:
        analyze_button = st.button("🔄 Re-run Analysis", type="primary", use_container_width=True)
    
    with col2:
        # Email report button text
        if recipient_option == "okr_users":
            button_text = "📧 Send Email Report to OKR Users"
        else:
            button_text = "📧 Send Enhanced Email Report"
        email_button = st.button(button_text, type="secondary", use_container_width=True)
    
    # Handle button clicks
    if analyze_button:
        run_analysis(analyzer, selected_cycle, show_missing_analysis)
    
    if email_button:
        send_email_report_enhanced(
            analyzer, email_generator, selected_cycle, 
            "apluscorp.hr@gmail.com", 'mems nctq yxss gruw', 
            recipient_option, custom_emails
        )


if __name__ == "__main__":
    main()
